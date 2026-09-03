import contextlib
import io
import json
import os
import re
import subprocess
import tempfile
import threading
import time
import traceback
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, send_file, send_from_directory
from openpyxl import Workbook

import attachments as attachments_mod
import barby
import barby_events
import db
import discord_bot
import filters as watcher_filters
import haku
import import_jerujam
import kupat
import kupat_credits
import kupat_events
import dice
import kupat_pdf
import mail_intake
import market
import matcher
import notify
import edm_events
import pacha_events
import pacha_tickets
import scraper
import series
import tickchak
import tickchak_pdf
import ticketmaster
import tm_discover
import tm_events
import todos as todos_mod
import vault
import viagogo_listing
import viagogo_market_sales
import viagogo_pricer
import crowdvolt_pricer

# Drop-checker sources keyed by the value stored in tm_watchers.source.
# Each module exposes parse_url, perf_url, fetch_selectable_seats,
# seat_key, get_labels, event_summary. Adding a new ticketing site is a
# matter of writing one module and adding a row here.
WATCHER_SOURCES = {
    "ticketmaster": ticketmaster,
    "kupat": kupat,
    "tickchak": tickchak,
    "barby": barby,
    "dice": dice,
    "haku": haku,
    "tmdiscover": tm_discover,
}


def _detect_source(url):
    """Pick the right source module from a URL. Returns (source_name, module).

    Falls back to ticketmaster for shorthand 'ABC123/001' with letters,
    kupat for shorthand '1358/51596' (digits only — both kupat IDs are
    numeric), so existing callers keep working.
    """
    s = (url or "").strip().lower()
    if "barby.co.il" in s:
        return "barby", barby
    if "dice.fm" in s:
        return "dice", dice
    # haku race registration (Houston Marathon): the platform's own hosts
    # plus the marathon site, whose URLs parse_url maps to the known key.
    if "hakuapp.com" in s or "haku.ly" in s or "houstonmarathon" in s:
        return "haku", haku
    if "tickchak.co.il" in s:
        return "tickchak", tickchak
    if "kupat.co.il" in s:
        return "kupat", kupat
    # TM-IL "series" landing pages (discover.ticketmaster.co.il/event/<slug>)
    # — a separate site from the ticketing SPA, so this must run BEFORE the
    # ticketmaster.co.il rule below, which would otherwise swallow it.
    if "discover.ticketmaster.co.il" in s or re.match(r"\s*(?:tm)?discover\s*/", s or ""):
        return "tmdiscover", tm_discover
    if "ticketmaster.co.il" in s:
        return "ticketmaster", ticketmaster
    if re.fullmatch(r"\s*\d+\s*/\s*\d+\s*", s or ""):
        return "kupat", kupat
    # Bare 24-hex shorthand is a DICE internal event id (must run before
    # the tickchak bare-slug rule, which would also match it).
    if re.fullmatch(r"\s*[a-f0-9]{24}\s*", s or ""):
        return "dice", dice
    # Bare 20-hex is a haku event key (also before the tickchak slug rule).
    if re.fullmatch(r"\s*[a-f0-9]{20}\s*", s or ""):
        return "haku", haku
    # Bare slug shorthand with no slash, scheme, or query string — most
    # likely a tickchak event slug (e.g. "mada26", "103350").
    if re.fullmatch(r"\s*[a-z0-9_\-]{2,}\s*", s or "") and "/" not in s and "?" not in s:
        return "tickchak", tickchak
    return "ticketmaster", ticketmaster

load_dotenv()

app = Flask(__name__)
# Cap multipart uploads (file + form overhead) — slightly above the per-file
# MAX_BYTES in attachments.py so legitimate uploads aren't rejected by Flask
# before our route handler can return a friendly error.
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024
db.init()
# Seed the three originally-requested EDM events on a fresh DB. One-way
# latch inside the helper, so removing one doesn't resurrect it at boot.
db.edm_seed_tracked(edm_events.DEFAULT_TRACKED,
                    datetime.now(timezone.utc).isoformat())

_last_run = {"at": None, "count": 0, "error": None, "running": False}
_run_lock = threading.Lock()

BACKUP_DIR = Path(os.getenv("KARTIS_BACKUP_DIR") or (Path(__file__).parent / "backups"))
BACKUP_KEEP_DAYS = int(os.getenv("KARTIS_BACKUP_KEEP_DAYS") or 30)
_last_backup = {"at": None, "path": None, "error": None, "running": False}
_backup_lock = threading.Lock()

_last_jerujam = {"at": None, "count": None, "error": None, "running": False}
_jerujam_lock = threading.Lock()

# Pacha ticket fetch is on-demand from the Tools page. Synchronous like the
# kupat/tickchak PDF tools, but guarded by a lock so two clicks can't hit Gmail
# at once.
_pacha_lock = threading.Lock()

_last_tm = {"at": None, "checked": 0, "drops": 0, "errors": 0, "running": False}
_tm_lock = threading.Lock()

# Viagogo auto-pricer — full-dashboard mode only (needs the CDP Chrome).
# watcher_only.py must never grow this job.
_last_pricer = {"at": None, "changed": 0, "paused": 0, "skipped": 0,
                "errors": 0, "eligible": 0, "dry_run": None, "error": None,
                "running": False}
_pricer_lock = threading.Lock()
PRICER_INTERVAL_MINUTES = int(os.getenv("KARTIS_PRICER_INTERVAL_MINUTES") or 15)
# The pricer job wakes on the FAST cadence and decides per tick whether it
# owes a full pass (every PRICER_INTERVAL_MINUTES) or a fast-lane-only pass
# over the events in the `pricer_fast_events` setting. One job rather than
# two so a fast tick can never collide with — or starve — the full pass.
PRICER_FAST_INTERVAL_MINUTES = max(
    1, int(os.getenv("KARTIS_PRICER_FAST_INTERVAL_MINUTES") or 5))
_last_pricer_full_at = None

# CrowdVolt auto-pricer — same rules, CV's JSON APIs through the CDP Chrome.
_last_cv_pricer = {"at": None, "changed": 0, "paused": 0, "skipped": 0,
                   "errors": 0, "eligible": 0, "listings": 0, "dry_run": None,
                   "error": None, "running": False}
_cv_pricer_lock = threading.Lock()
CV_PRICER_INTERVAL_MINUTES = int(os.getenv("KARTIS_CV_PRICER_INTERVAL_MINUTES") or 15)
TM_CHECK_INTERVAL_SECONDS = int(os.getenv("TM_CHECK_INTERVAL_SECONDS") or 60)
# Per-seat notification cool-down (minutes). A hot seat that flaps in and out
# of the buyable feed — kupat VIP seats cycling through carts/holds — would
# otherwise re-ping every few minutes; after a seat pings we suppress further
# pings of that exact seat for this long. 0 disables. (Keep in sync with
# watcher_only.py.)
SEAT_COOLDOWN_SECONDS = int(os.getenv("KARTIS_SEAT_COOLDOWN_MINUTES") or 30) * 60
# DICE restock-snipe fast poll: while a dice watcher is SOLD OUT we poll it
# every few seconds so a restock is caught in seconds, not up to a minute.
# Only sold-out dice watchers are fetched (usually 1-2), keeping the
# request footprint on DICE's anonymous API tiny. Set _ENABLED=0 to disable.
DICE_SNIPER_SECONDS = int(os.getenv("KARTIS_DICE_SNIPER_SECONDS") or 15)
DICE_SNIPER_ENABLED = (os.getenv("KARTIS_DICE_SNIPER_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")

# Fast poll for the TM-IL sources. Unlike dice's sniper this is not restock-
# only: every TM source here is cheap pure HTTP (a whole event-level NEXT
# check is ~1.5-3.5s of small JSON) and the thing being raced is a return
# landing on a sold-out date, which can be gone inside a minute. The 60s tick
# was the latency floor for the one source that could afford to beat it.
# Shares `_tm_lock` with run_tm_check, so a watcher is never fetched twice at
# once and the loser of a race is a no-op against stored seat state.
TM_SNIPER_SECONDS = int(os.getenv("KARTIS_TM_SNIPER_SECONDS") or 20)
TM_SNIPER_ENABLED = (os.getenv("KARTIS_TM_SNIPER_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")
TM_SNIPER_SOURCES = {"ticketmaster", "tmdiscover"}
# Drop-checking can be disabled here when the watcher runs on another machine
# (e.g. the VPS), so the dashboard still serves inventory/sales without
# double-pinging Discord. Manual "check now" from the UI still works.
TM_CHECK_ENABLED = (os.getenv("TM_CHECK_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")

_last_intake = {"at": None, "fetched": 0, "saved": 0, "skipped_dupe": 0, "errors": 0, "error": None, "running": False}
_intake_lock = threading.Lock()
INTAKE_INTERVAL_MINUTES = int(os.getenv("KARTIS_INTAKE_INTERVAL_MINUTES") or 10)

_last_todo_remind = {"at": None, "sent_count": 0, "due_today": 0, "overdue": 0,
                     "error": None, "running": False, "paused": False,
                     "muted": False, "result": None}
_todo_remind_lock = threading.Lock()

# Pacha NYC new-event monitor — polls pacha-nyc.com/events and pings Discord
# on new events / waitlist→on-sale flips / GA price climbs / GA-or-VIP price
# DROPS (returns or a cheaper release re-opening — the "pacha shocks"
# channel). Pure HTTP (no Chrome), so it could run anywhere — but only ONE
# machine may have it enabled or Discord gets double pings. Prod = the VPS; a
# locally-run dashboard sets KARTIS_PACHA_MONITOR_ENABLED=0 (same idea as
# TM_CHECK_ENABLED).
_last_pacha_events = {"at": None, "events": 0, "new": 0, "onsale": 0,
                      "price_up": 0, "price_down": 0, "low_stock": 0,
                      "notified": 0, "baseline": False, "error": None,
                      "running": False}
_pacha_events_lock = threading.Lock()
# 1-minute default: a price dip (someone returns tickets) can sell out again
# in minutes, so the shock ping is only useful if the poll catches the dip.
PACHA_MONITOR_INTERVAL_MINUTES = int(os.getenv("KARTIS_PACHA_MONITOR_INTERVAL_MINUTES") or 1)
PACHA_MONITOR_ENABLED = (os.getenv("KARTIS_PACHA_MONITOR_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")
# One tick can't spam more pings than this — a parser anomaly that makes
# every event look "new" should hit the cap, not flood the channel.
PACHA_MAX_PINGS_PER_TICK = 12
# Low-stock alert: ping once when the GA release's tickets-left count
# crosses at/below this (re-armed by the next release).
PACHA_LOW_STOCK_THRESHOLD = int(os.getenv("KARTIS_PACHA_LOW_STOCK_THRESHOLD") or 20)
# The /market feed (market_events row + availability snapshot per event) keeps
# its old ~10-min cadence even though the tick now runs every minute —
# snapshots are inserts, and 1-min inserts would grow the series 10x for no
# extra velocity resolution.
PACHA_MARKET_SNAPSHOT_MINUTES = int(os.getenv("KARTIS_PACHA_MARKET_SNAPSHOT_MINUTES") or 10)
_pacha_last_market_at = None

# US EDM single-event trackers — posh.vip, events.leapevents.com and
# wl.eventim.us (edm_events.py + the three *_events fetchers). Same idea as
# the Pacha monitor and the same ping vocabulary, but pointed at NAMED
# events instead of a venue catalog: the poll list lives in
# edm_tracked_events and is managed with `python edm_events.py --add <url>`
# or POST /api/edm/add. Pure HTTP (no Chrome), so it can run on the VPS —
# and, like pacha, exactly ONE machine may enable it or Discord double-pings.
#
# Counts are asymmetric across the three: only posh publishes remaining
# inventory, so low_stock effectively only fires there. leap publishes none
# at all and eventim only leaks a count once a type drops under its
# per-order limit. Everywhere below, `None` means UNKNOWN and is never
# treated as zero.
_last_edm_events = {"at": None, "tracked": 0, "events": 0, "new": 0,
                    "onsale": 0, "restock": 0, "soldout": 0, "new_tier": 0,
                    "price_up": 0, "price_down": 0, "low_stock": 0,
                    "notified": 0, "baseline": False, "errors": {},
                    "error": None, "running": False}
_edm_events_lock = threading.Lock()
# 2 minutes: waves on these platforms sell out in minutes and a returned
# ticket re-lists without warning, so the shock ping is only worth having
# if the poll is tight. Still gentler than pacha's 1 min — there are three
# hosts here and each tracked event is its own request.
EDM_MONITOR_INTERVAL_MINUTES = int(os.getenv("KARTIS_EDM_MONITOR_INTERVAL_MINUTES") or 2)
EDM_MONITOR_ENABLED = (os.getenv("KARTIS_EDM_MONITOR_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")
EDM_MAX_PINGS_PER_TICK = 12
# Low-stock alert: ping once when the current release's remaining count
# crosses at/below this. Re-armed when the lead tier changes.
EDM_LOW_STOCK_THRESHOLD = int(os.getenv("KARTIS_EDM_LOW_STOCK_THRESHOLD") or 20)
# /market feed throttle, same reasoning as PACHA_MARKET_SNAPSHOT_MINUTES.
EDM_MARKET_SNAPSHOT_MINUTES = int(os.getenv("KARTIS_EDM_MARKET_SNAPSHOT_MINUTES") or 10)
_edm_last_market_at = None
# Venue-catalog sync (tao_events: the two Marquee NY calendars). This is the
# new-EVENT radar — it reconciles taogroup.com's listings into the poll list
# so a newly announced show starts being watched without anyone pasting a
# URL. It rides the EDM tick but on its own cadence: a show appears on the
# calendar hours-to-days before anything happens to its price, so polling
# the two calendars every couple of minutes would be pure waste.
EDM_CATALOG_SYNC_MINUTES = int(os.getenv("KARTIS_EDM_CATALOG_SYNC_MINUTES") or 10)
_edm_last_catalog_sync_at = None
_last_edm_catalogs = {}

# Israeli-sites new-event monitor — polls the kupat.co.il and
# ticketmaster.co.il listing feeds (kupat_events.py / tm_events.py, pure
# HTTP) and pings Discord on new events and, for TM, listed→on-sale flips.
# Same single-machine rule as the Pacha monitor: prod = the VPS; a locally
# run dashboard sets KARTIS_IL_EVENTS_ENABLED=0 or Discord double-pings.
IL_EVENT_SOURCES = {"kupat": kupat_events, "tm": tm_events, "barby": barby_events}
_last_il_events = {"at": None, "events": {}, "new": 0, "onsale": 0,
                   "newdate": 0, "notified": 0, "baseline": [], "errors": {},
                   "error": None, "running": False}
_il_events_lock = threading.Lock()
IL_EVENTS_INTERVAL_MINUTES = int(os.getenv("KARTIS_IL_EVENTS_INTERVAL_MINUTES") or 10)
IL_EVENTS_ENABLED = (os.getenv("KARTIS_IL_EVENTS_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")
IL_EVENTS_MAX_PINGS_PER_TICK = 12

# Market-wide tracker (/market) — hourly availability sweep of every event
# on kupat + TM-IL (+ manually-added tickchak) via market.py. Needs
# patchright Chromium for the kupat catalog, so it runs on the dashboard
# machine; the usual single-machine rule applies (KARTIS_MARKET_ENABLED=0
# elsewhere, or the snapshot series gets double-density from two boxes —
# harmless for the math, wasteful for the sites).
_last_market = {"at": None, "counts": {}, "errors": {}, "entities": 0,
                "duration_s": None, "error": None, "running": False}
_market_lock = threading.Lock()
MARKET_INTERVAL_MINUTES = int(os.getenv("KARTIS_MARKET_INTERVAL_MINUTES") or 60)
MARKET_ENABLED = (os.getenv("KARTIS_MARKET_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")

# Market-wide viagogo sales tracker (/vgsales) — polls the MarketDataV3
# magnifier popup for every event we're listed on + the watchlist and
# records the "past ten sales" grid (viagogo_market_sales.py docstring has
# the empirical facts). Needs the CDP Chrome, so dashboard machine only
# (KARTIS_VGSALES_ENABLED=0 elsewhere). No notifications in v1.
_last_vgsales = {"at": None, "targets": 0, "fetched": 0, "skipped_fresh": 0,
                 "deferred": 0, "new_sales": 0, "baselines": 0, "repriced": 0,
                 "overflows": 0, "errors": 0, "error": None, "running": False}
_vgsales_lock = threading.Lock()
VGSALES_ENABLED = (os.getenv("KARTIS_VGSALES_ENABLED") or "1").strip().lower() not in ("0", "false", "no", "off")


def run_mail_intake():
    if not _intake_lock.acquire(blocking=False):
        return
    _last_intake["running"] = True
    try:
        summary = mail_intake.run_intake()
        _last_intake.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=None,
            **summary,
        )
    except Exception as e:
        _last_intake.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_intake["running"] = False
        _intake_lock.release()


def _pacha_market_rows(events, now_iso):
    """Feed the /market page from the pacha tick: upsert one market_events
    row per event and snapshot availability so the velocity windows work.
    Availability = sum across the currently purchasable tiers; capacity is
    left NULL (tier quantities are per-release, not event capacity — same
    treatment as kupat GA rows)."""
    for ev in events:
        if ev.get("ga_sold_out") and not ev.get("total_available"):
            status = "soldout"
        elif ev.get("iswaitlist"):
            status = "waitlist"
        elif ev.get("on_sale"):
            status = "available"
        else:
            status = "unknown"
        first_ms = None
        if ev.get("start_date"):
            try:
                first_ms = int(datetime.fromisoformat(ev["start_date"]).timestamp() * 1000)
            except ValueError:
                pass
        db.market_upsert({
            "source": "pacha", "event_code": ev["event_id"], "perf_code": "0",
            "name": ev.get("name"), "venue": "Pacha NYC",
            "date_text": ev.get("date_text"), "first_date_ms": first_ms,
            "url": ev.get("page_url"), "status": status,
            "capacity": None, "available": ev.get("total_available"),
            "min_price": ev.get("ga_price"), "currency": "USD",
            "manual": False, "last_error": None,
        }, now_iso)
        if ev.get("total_available") is not None:
            db.sales_snapshot_insert("pacha", ev["event_id"], "0",
                                     None, ev["total_available"], None, now_iso)


def run_pacha_events():
    """One tick of the Pacha NYC new-event monitor. Fetch the current event
    list, diff against pacha_seen_events, ping Discord for: brand-new
    events, waitlist→on-sale flips, GA price climbs (both prices > 0), GA or
    VIP price DROPS (a return / a cheaper earlier release re-opening — the
    time-critical "shock" ping, routed to #pacha-shocks), and the GA
    release's tickets-left count crossing below PACHA_LOW_STOCK_THRESHOLD
    (early warning that the price will jump). The /market feed (market_events
    row + availability snapshot per event, source='pacha') rides the tick too
    but is throttled to PACHA_MARKET_SNAPSHOT_MINUTES so the 1-min poll
    doesn't inflate the snapshot series.

    Gating mirrors the drop checker: the very first tick ever (empty seen
    table) is a baseline — store everything, ping nothing. master_paused
    skips the tick entirely; master_muted keeps state current but skips the
    Discord sends. Low-stock fires only on a crossing within the SAME
    release (old count above threshold, new at/below), so it pings once per
    release, and a release flip re-arms it."""
    if not _pacha_events_lock.acquire(blocking=False):
        return
    _last_pacha_events["running"] = True
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        if db.setting_get_bool("master_paused", default=False):
            _last_pacha_events.update(at=now_iso, error=None)
            return

        events = pacha_events.fetch_events()
        seen = db.pacha_all_seen()
        baseline = not seen
        muted = db.setting_get_bool("master_muted", default=False)

        pings = []  # (kind, ev, old_row)
        if not baseline:
            for ev in events:
                old = seen.get(ev["event_id"])
                if old is None:
                    pings.append(("new", ev, None))
                    continue
                if not old["on_sale"] and ev["on_sale"]:
                    pings.append(("onsale", ev, old))
                old_ga, new_ga = old["ga_price"], ev["ga_price"]
                if old_ga and new_ga and new_ga > old_ga:
                    pings.append(("price_up", ev, old))
                elif (ev.get("ga_available") is not None
                        and old.get("ga_available") is not None
                        and old.get("ga_release") == ev.get("ga_release")
                        and old["ga_available"] > PACHA_LOW_STOCK_THRESHOLD
                        and ev["ga_available"] <= PACHA_LOW_STOCK_THRESHOLD):
                    pings.append(("low_stock", ev, old))
                # Shock: GA or VIP got CHEAPER — a return or an earlier/
                # cheaper release re-opened. Truthiness guards mirror the
                # climb check (None/0 on either side never fires), so the
                # first tick after the vip_price migration baselines VIP
                # silently. Independent of the chain above: a GA climb + VIP
                # drop in one tick correctly emits both pings.
                old_vip, new_vip = old.get("vip_price"), ev.get("vip_price")
                if ((old_ga and new_ga and new_ga < old_ga)
                        or (old_vip and new_vip and new_vip < old_vip)):
                    pings.append(("price_down", ev, old))

        # Shocks are the perishable ones (the cheap tickets are selling right
        # now) — send them first so a burst of "new" pings can't starve them
        # out of the per-tick cap.
        pings.sort(key=lambda p: p[0] != "price_down")

        notified = 0
        if not muted:
            for kind, ev, old in pings[:PACHA_MAX_PINGS_PER_TICK]:
                res = notify.notify_pacha_event(kind, ev, old)
                print(f"[pacha] {kind}: {ev['name']} ({ev.get('date_text')}) -> {res}")
                notified += 1
                time.sleep(0.5)
            if len(pings) > PACHA_MAX_PINGS_PER_TICK:
                print(f"[pacha] ping cap hit — {len(pings) - PACHA_MAX_PINGS_PER_TICK} suppressed")

        # Lifetime sold since tracking began: accumulate availability drops
        # between ticks. Release flips ADD inventory (available jumps up) —
        # clamped to 0 like _sales_windows, so they never subtract.
        for ev in events:
            old = seen.get(ev["event_id"]) or {}
            delta = 0
            if old.get("total_available") is not None and ev.get("total_available") is not None:
                delta = max(0, old["total_available"] - ev["total_available"])
            ev["sold_cum"] = (old.get("sold_cum") or 0) + delta
            ev["sold_cum_since"] = old.get("sold_cum_since") or now_iso

        for ev in events:
            db.pacha_upsert_seen(ev, now_iso)
            db.pacha_release_sync(ev["event_id"], ev.get("tiers"), now_iso)
        global _pacha_last_market_at
        now_dt = datetime.now(timezone.utc)
        if (_pacha_last_market_at is None
                or now_dt - _pacha_last_market_at
                >= timedelta(minutes=PACHA_MARKET_SNAPSHOT_MINUTES)):
            _pacha_market_rows(events, now_iso)
            _pacha_last_market_at = now_dt

        _last_pacha_events.update(
            at=now_iso, error=None, events=len(events), baseline=baseline,
            new=sum(1 for k, _, _ in pings if k == "new"),
            onsale=sum(1 for k, _, _ in pings if k == "onsale"),
            price_up=sum(1 for k, _, _ in pings if k == "price_up"),
            price_down=sum(1 for k, _, _ in pings if k == "price_down"),
            low_stock=sum(1 for k, _, _ in pings if k == "low_stock"),
            notified=notified,
        )
        if baseline:
            print(f"[pacha] baseline stored: {len(events)} events, no pings")
    except Exception as e:
        _last_pacha_events.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_pacha_events["running"] = False
        _pacha_events_lock.release()


def _edm_market_rows(events, now_iso):
    """Feed the /market page from the EDM tick: one market_events row per
    tracked event plus an availability snapshot so the velocity windows
    work. Only posh publishes counts, so leap/eventim rows carry
    available=None and simply have no velocity series — deliberately, since
    inventing a number would poison the sold-per-window math."""
    for ev in events:
        if ev.get("sold_out"):
            status = "soldout"
        elif ev.get("on_sale"):
            status = "available"
        else:
            status = "unknown"
        first_ms = None
        if ev.get("start_date"):
            try:
                first_ms = int(datetime.fromisoformat(
                    str(ev["start_date"]).replace("Z", "+00:00")).timestamp() * 1000)
            except ValueError:
                pass
        db.market_upsert({
            "source": ev["source"], "event_code": ev["event_key"], "perf_code": "0",
            "name": ev.get("name"), "venue": ev.get("venue"),
            "date_text": ev.get("date_text"), "first_date_ms": first_ms,
            "url": ev.get("page_url"), "status": status,
            "capacity": None, "available": ev.get("total_available"),
            "min_price": ev.get("min_price"), "currency": ev.get("currency") or "USD",
            "manual": False, "last_error": None,
        }, now_iso)
        if ev.get("total_available") is not None:
            db.sales_snapshot_insert(ev["source"], ev["event_key"], "0",
                                     None, ev["total_available"], None, now_iso)


def _edm_sync_catalogs(now_iso):
    """Reconcile the venue-catalog sources (tao) into the tracked list, at
    most every EDM_CATALOG_SYNC_MINUTES.

    A discovery failure is recorded in _last_edm_catalogs and otherwise
    ignored: the events already tracked keep polling, and the calendar gets
    another try on the next sync."""
    global _edm_last_catalog_sync_at
    if not edm_events.CATALOG_SOURCES:
        return
    now_dt = datetime.now(timezone.utc)
    if (_edm_last_catalog_sync_at is not None
            and now_dt - _edm_last_catalog_sync_at
            < timedelta(minutes=EDM_CATALOG_SYNC_MINUTES)):
        return
    _edm_last_catalog_sync_at = now_dt

    results = edm_events.sync_catalogs(now_iso=now_iso)
    _last_edm_catalogs.clear()
    _last_edm_catalogs.update({"at": now_iso, **results})
    for name, r in results.items():
        for eid in r["added"]:
            print(f"[edm] {name} catalog: now tracking {eid}")
        for eid in r["pruned"]:
            print(f"[edm] {name} catalog: dropped {eid} (no longer listed)")
        for where, msg in (r["errors"] or {}).items():
            print(f"[edm] {name} catalog error ({where}): {msg}")


def _edm_baseline_sources(seen):
    """Catalog sources with no observed state at all yet. Their whole
    listing lands in one tick, and "every show currently on sale" is
    backfill, not news — so that first tick stores state and pings nothing,
    exactly like the pacha monitor's. Every tick after it, an event with no
    stored row really is newly announced and gets its 'new' ping.

    Keyed off edm_seen_events rather than off the sync's own bookkeeping so
    it stays right however the poll list got populated (a manual
    `edm_events.py --sync-catalogs` before the first tick, a restored
    backup, a source added to CATALOG_SOURCES later)."""
    observed = {r.get("source") for r in seen.values()}
    return {s for s in edm_events.CATALOG_SOURCES if s not in observed}


def _edm_diff(ev, old):
    """Ping kinds for one event given its previous edm_seen_events row.
    Returns a list of kind strings (an event can legitimately emit several
    in one tick — e.g. a new wave opening both adds a tier and moves the
    price).

    Every comparison is guarded on BOTH sides being known: leap and eventim
    publish no counts, and a None on either side must never be read as a
    change. Same for prices, so the first tick after an event goes on sale
    can't fake a price move."""
    kinds = []
    was_on_sale = bool(old["on_sale"])
    was_sold_out = bool(old["sold_out"])

    if not was_on_sale and ev["on_sale"]:
        # A sold-out event coming back is the perishable case and gets its
        # own, louder kind.
        kinds.append("restock" if was_sold_out else "onsale")
    if not was_sold_out and ev["sold_out"]:
        kinds.append("soldout")

    old_names = set()
    if old.get("tiers_json"):
        try:
            old_names = {(t.get("name") or "").strip()
                         for t in json.loads(old["tiers_json"])}
        except (ValueError, TypeError):
            old_names = set()
    if old_names:
        # Only claim "new release" when we actually have a previous tier
        # list to compare against — an unparseable/missing one would make
        # every tier look brand new.
        fresh = [t for t in ev["tiers"]
                 if (t.get("name") or "").strip() not in old_names]
        if fresh:
            ev["_new_tiers"] = fresh
            kinds.append("new_tier")

    old_price, new_price = old.get("min_price"), ev.get("min_price")
    if old_price is not None and new_price is not None:
        if new_price > old_price:
            kinds.append("price_up")
        elif new_price < old_price:
            kinds.append("price_down")

    # Low stock: a crossing WITHIN the same release, so it pings once and a
    # new release re-arms it. Needs real counts on both sides — i.e. posh.
    old_left, new_left = old.get("lead_available"), ev.get("lead_available")
    if (old_left is not None and new_left is not None
            and old.get("lead_tier") == ev.get("lead_tier")
            and old_left > EDM_LOW_STOCK_THRESHOLD
            and new_left <= EDM_LOW_STOCK_THRESHOLD):
        kinds.append("low_stock")
    return kinds


# Perishable first, so a burst of routine pings can't push a price drop or
# a restock past the per-tick cap.
_EDM_KIND_PRIORITY = {"price_down": 0, "restock": 1, "low_stock": 2,
                      "new_tier": 3, "onsale": 4, "soldout": 5,
                      "price_up": 6, "new": 7}


def run_edm_events():
    """One tick of the US EDM event trackers. Fetch every tracked event
    (posh / leap / eventim), diff against edm_seen_events and ping Discord
    for price drops, restocks, low stock, new releases, on-sale flips,
    sell-outs and price climbs. The /market feed rides along, throttled to
    EDM_MARKET_SNAPSHOT_MINUTES.

    Gating mirrors the pacha monitor: the first tick for a given event
    (no stored row) is a baseline — store it, send at most the informational
    'new' ping. master_paused skips the tick entirely; master_muted keeps
    state current but sends nothing. A per-event fetch failure is recorded
    against that event and does NOT abort the others, and it leaves the
    last good state alone — a blocked fetch must never read as 'sold out'."""
    if not _edm_events_lock.acquire(blocking=False):
        return
    _last_edm_events["running"] = True
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        if db.setting_get_bool("master_paused", default=False):
            _last_edm_events.update(at=now_iso, error=None)
            return

        # Venue-catalog radar first, so a show announced since the last sync
        # is already in `tracked` and gets its state stored on this tick.
        _edm_sync_catalogs(now_iso)

        tracked = db.edm_tracked_all(include_paused=False)
        events, errors = edm_events.fetch_tracked(tracked)
        for event_id, msg in errors.items():
            db.edm_mark_error(event_id, msg, now_iso)
            print(f"[edm] fetch failed for {event_id}: {msg}")

        seen = db.edm_all_seen()
        muted = db.setting_get_bool("master_muted", default=False)
        baseline_sources = _edm_baseline_sources(seen)

        pings = []  # (kind, ev, old_row)
        for ev in events:
            old = seen.get(ev["event_id"])
            if old is None:
                # Newly tracked: baseline it. The 'new' ping is informational
                # and carries the current price, so adding an event confirms
                # in Discord that it's being watched — except on a catalog
                # source's very first tick, where the whole venue listing
                # arrives at once and is backfill, not news.
                if ev["source"] not in baseline_sources:
                    pings.append(("new", ev, None))
                continue
            for kind in _edm_diff(ev, old):
                pings.append((kind, ev, old))

        pings.sort(key=lambda p: _EDM_KIND_PRIORITY.get(p[0], 9))

        notified = 0
        if not muted:
            for kind, ev, old in pings[:EDM_MAX_PINGS_PER_TICK]:
                res = notify.notify_edm_event(kind, ev, old)
                print(f"[edm] {kind}: {ev['name']} ({ev['source']}) -> {res}")
                notified += 1
                time.sleep(0.5)
            if len(pings) > EDM_MAX_PINGS_PER_TICK:
                print(f"[edm] ping cap hit — {len(pings) - EDM_MAX_PINGS_PER_TICK} suppressed")

        # Lifetime sold since tracking began: accumulate availability drops
        # between ticks. A new release ADDS inventory, so increases are
        # clamped to 0 exactly like pacha's sold_cum and _sales_windows.
        for ev in events:
            old = seen.get(ev["event_id"]) or {}
            delta = 0
            if (old.get("total_available") is not None
                    and ev.get("total_available") is not None):
                delta = max(0, old["total_available"] - ev["total_available"])
            ev["sold_cum"] = (old.get("sold_cum") or 0) + delta
            ev["sold_cum_since"] = old.get("sold_cum_since") or now_iso

        for ev in events:
            db.edm_upsert_seen(ev, now_iso)
            db.edm_release_sync(ev["event_id"], ev.get("tiers"), now_iso)

        global _edm_last_market_at
        now_dt = datetime.now(timezone.utc)
        if events and (_edm_last_market_at is None
                       or now_dt - _edm_last_market_at
                       >= timedelta(minutes=EDM_MARKET_SNAPSHOT_MINUTES)):
            _edm_market_rows(events, now_iso)
            _edm_last_market_at = now_dt

        counts = {k: sum(1 for kk, _, _ in pings if kk == k)
                  for k in ("new", "onsale", "restock", "soldout", "new_tier",
                            "price_up", "price_down", "low_stock")}
        _last_edm_events.update(
            at=now_iso, error=None, tracked=len(tracked), events=len(events),
            # A catalog source's first tick stores state without a single
            # 'new' ping, so count the silent baselines too.
            baseline=(any(k == "new" for k, _, _ in pings)
                      or bool(baseline_sources & {e["source"] for e in events})),
            errors=errors, notified=notified, **counts)
    except Exception as e:
        _last_edm_events.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_edm_events["running"] = False
        _edm_events_lock.release()


def run_il_events():
    """One tick of the Israeli-sites new-event monitor. For each source in
    IL_EVENT_SOURCES: fetch the current listing, diff against
    site_seen_events, ping Discord for brand-new events and (TM only)
    listed→on-sale flips. Sources fail independently — a kupat outage
    doesn't skip the TM diff.

    TM's listing feeds don't say whether sales are open, so on_sale arrives
    as None and is resolved lazily via tm_events.check_on_sale — but only
    for events not already stored as on-sale, so steady-state ticks cost a
    couple of extra requests at most (the first tick resolves everything
    once for the baseline). If the resolver errors, the stored value is
    kept (or True for a brand-new event, so a bad first read can't fire a
    spurious on-sale ping later).

    Gating mirrors the Pacha monitor: first tick ever per source is a
    baseline — store everything, ping nothing. master_paused skips the tick
    entirely; master_muted keeps state current but skips the Discord sends."""
    if not _il_events_lock.acquire(blocking=False):
        return
    _last_il_events["running"] = True
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        if db.setting_get_bool("master_paused", default=False):
            _last_il_events.update(at=now_iso, error=None)
            return

        muted = db.setting_get_bool("master_muted", default=False)
        counts, baselines, errors = {}, [], {}
        pings = []  # (kind, ev, old_row)

        for source, mod in IL_EVENT_SOURCES.items():
            try:
                events = mod.fetch_events()
            except Exception as e:
                errors[source] = f"{type(e).__name__}: {e}"
                continue
            # Sources exposing fetch_presentations (kupat: one site-wide
            # catalog call; tm: one perf-list request per event) get per-date
            # visibility, so a new show added under an EXISTING event page
            # (same event key, so invisible to the listing-feed diff) pings
            # as 'newdate'. On fetch failure the perf diff is skipped this
            # tick — stored perfs_json state is never touched. An event_key
            # ABSENT from the map means "no data this tick" and is likewise
            # skipped (never treated as an empty date list).
            perf_map = None
            if hasattr(mod, "fetch_presentations"):
                try:
                    perf_map = mod.fetch_presentations(events)
                except Exception as e:
                    print(f"[il-events] {source} presentations fetch failed (perf diff skipped): {e}")
            seen = db.site_events_all_seen(source)
            baseline = not seen
            if baseline:
                baselines.append(source)
            counts[source] = len(events)

            for ev in events:
                old = seen.get(ev["event_key"])
                if ev.get("on_sale") is None:
                    if old is not None and old["on_sale"]:
                        ev["on_sale"] = True  # sale opened before; no re-check
                    else:
                        try:
                            ev["on_sale"] = mod.check_on_sale(ev)
                        except Exception as e:
                            print(f"[il-events] {source} check_on_sale({ev['event_key']}) failed: {e}")
                            ev["on_sale"] = True if old is None else bool(old["on_sale"])
                # Sale-opened is a one-way latch: a show dropping off the
                # kupat homepage (rotation, hiccup) or a TM status blip must
                # not re-arm the 'onsale' ping and fire again on its return.
                if old is not None and old["on_sale"] and not ev["on_sale"]:
                    ev["on_sale"] = True
                if perf_map is not None and ev["event_key"] in perf_map:
                    cur = perf_map[ev["event_key"]]
                    cur_keys = {p["perf_key"] for p in cur}
                    known_json = old.get("perfs_json") if old else None
                    if old is None or baseline or known_json is None:
                        # brand-new event / first tick with perf data: store
                        # a silent baseline (the 'new' ping already covers a
                        # brand-new event's dates).
                        ev["_perf_union"] = cur_keys
                    else:
                        known = set(json.loads(known_json))
                        added = [p for p in cur if p["perf_key"] not in known]
                        if added:
                            ev["new_perfs"] = added
                            pings.append(("newdate", ev, old))
                        ev["_perf_union"] = known | cur_keys
                if baseline:
                    continue
                if old is None:
                    pings.append(("new", ev, None))
                elif not old["on_sale"] and ev["on_sale"]:
                    pings.append(("onsale", ev, old))

            for ev in events:
                db.site_events_upsert_seen(source, ev, now_iso)
                if ev.get("_perf_union") is not None:
                    db.site_events_set_perfs(source, ev["event_key"], ev["_perf_union"])
            if baseline:
                print(f"[il-events] {source} baseline stored: {len(events)} events, no pings")

        notified = 0
        if not muted:
            for kind, ev, old in pings[:IL_EVENTS_MAX_PINGS_PER_TICK]:
                res = notify.notify_site_event(kind, ev, old)
                print(f"[il-events] {ev['source']} {kind}: {ev['name']} ({ev.get('date_text')}) -> {res}")
                notified += 1
                time.sleep(0.5)
            if len(pings) > IL_EVENTS_MAX_PINGS_PER_TICK:
                print(f"[il-events] ping cap hit — {len(pings) - IL_EVENTS_MAX_PINGS_PER_TICK} suppressed")

        _last_il_events.update(
            at=now_iso, events=counts, baseline=baselines, errors=errors,
            error=("; ".join(f"{s}: {e}" for s, e in errors.items())
                   if len(errors) == len(IL_EVENT_SOURCES) else None),
            new=sum(1 for k, _, _ in pings if k == "new"),
            onsale=sum(1 for k, _, _ in pings if k == "onsale"),
            newdate=sum(1 for k, _, _ in pings if k == "newdate"),
            notified=notified,
        )
    except Exception as e:
        _last_il_events.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_il_events["running"] = False
        _il_events_lock.release()


def run_market_sweep():
    """One hourly tick of the market-wide tracker: market.run_sweep()
    discovers + snapshots every trackable event (see market.py docstring).
    No notifications here — discovery pings belong to the 10-min il_events
    job; this job only feeds the /market page's availability history."""
    if not _market_lock.acquire(blocking=False):
        return
    _last_market["running"] = True
    try:
        if db.setting_get_bool("master_paused", default=False):
            _last_market.update(at=datetime.now(timezone.utc).isoformat(), error=None)
            return
        summary = market.run_sweep()
        _last_market.update(error=None, **summary)
        print(f"[market] sweep: {summary['entities']} entities in {summary['duration_s']}s "
              f"counts={summary['counts']} errors={summary['errors'] or '—'}")
    except Exception as e:
        _last_market.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_market["running"] = False
        _market_lock.release()


def run_vgsales(force=False):
    """One tick of the viagogo market-sales tracker. No notifications —
    this only feeds the /vgsales page. Chrome down surfaces as `error` in
    the status dict, same as the pricer/market jobs. force=True (the
    run-now button) bypasses the fetched-recently skip that normally lets
    the pricer piggyback halve requests."""
    if not _vgsales_lock.acquire(blocking=False):
        return
    _last_vgsales["running"] = True
    try:
        if db.setting_get_bool("master_paused", default=False):
            _last_vgsales.update(at=datetime.now(timezone.utc).isoformat(), error=None)
            return
        summary = viagogo_market_sales.run_sales_tick(force=force)
        _last_vgsales.update(
            at=datetime.now(timezone.utc).isoformat(), error=None, **summary)
        print(f"[vgsales] {summary['fetched']}/{summary['targets']} events, "
              f"{summary['new_sales']} new sales, {summary['baselines']} baselines, "
              f"{summary['overflows']} overflows, {summary['errors']} errors")
    except Exception as e:
        _last_vgsales.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_vgsales["running"] = False
        _vgsales_lock.release()


def run_todo_remind():
    """Daily digest of open to-dos that are due today or overdue. Mirrors the
    drop-checker gating: master_paused skips entirely, master_muted logs but
    doesn't send. Each task gets `notified_at` stamped so the same item
    isn't re-pinged later the same day."""
    if not _todo_remind_lock.acquire(blocking=False):
        return
    _last_todo_remind["running"] = True
    try:
        today_iso = date.today().strftime("%Y-%m-%d")
        if db.setting_get_bool("master_paused", default=False):
            _last_todo_remind.update(
                at=datetime.now(timezone.utc).isoformat(),
                sent_count=0, due_today=0, overdue=0,
                paused=True, muted=False, error=None, result=None,
            )
            return
        due = db.todo_due_open(today_iso)
        if not due:
            _last_todo_remind.update(
                at=datetime.now(timezone.utc).isoformat(),
                sent_count=0, due_today=0, overdue=0,
                paused=False, muted=False, error=None, result=None,
            )
            return
        due_today = [t for t in due if (t.get("due_date_iso") or "")[:10] == today_iso]
        overdue = [t for t in due if (t.get("due_date_iso") or "")[:10] < today_iso]
        muted = db.setting_get_bool("master_muted", default=False)
        if muted:
            _last_todo_remind.update(
                at=datetime.now(timezone.utc).isoformat(),
                sent_count=0,
                due_today=len(due_today), overdue=len(overdue),
                paused=False, muted=True, error=None,
                result={"discord": "skipped (master_muted)", "email": "skipped (master_muted)"},
            )
            return
        result = notify.send_todo_digest(due_today, overdue)
        # Only stamp notified_at if at least one channel actually sent — that
        # way a transient credential outage doesn't suppress the next day's
        # digest from re-trying these same items.
        if any(str(v).startswith("ok") for v in result.values()):
            db.todo_mark_notified([t["id"] for t in due], today_iso)
        _last_todo_remind.update(
            at=datetime.now(timezone.utc).isoformat(),
            sent_count=len(due),
            due_today=len(due_today), overdue=len(overdue),
            paused=False, muted=False, error=None, result=result,
        )
    except Exception as e:
        _last_todo_remind.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_todo_remind["running"] = False
        _todo_remind_lock.release()


def run_jerujam_import():
    if not _jerujam_lock.acquire(blocking=False):
        return
    _last_jerujam["running"] = True
    try:
        counts = import_jerujam.run()
        _last_jerujam.update(
            at=datetime.now(timezone.utc).isoformat(),
            count=counts,
            error=None,
        )
    except Exception as e:
        _last_jerujam.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_jerujam["running"] = False
        _jerujam_lock.release()


def run_backup():
    if not _backup_lock.acquire(blocking=False):
        return
    _last_backup["running"] = True
    try:
        stamp = datetime.now().strftime("%Y-%m-%d")
        dest = BACKUP_DIR / f"kartis-{stamp}.db"
        db.backup(dest)
        _prune_old_backups()
        _last_backup.update(
            at=datetime.now(timezone.utc).isoformat(),
            path=str(dest),
            error=None,
        )
    except Exception as e:
        _last_backup.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_backup["running"] = False
        _backup_lock.release()


def _prune_old_backups():
    if not BACKUP_DIR.exists():
        return
    cutoff = datetime.now().timestamp() - BACKUP_KEEP_DAYS * 86400
    for f in BACKUP_DIR.glob("kartis-*.db"):
        try:
            if f.stat().st_mtime < cutoff:
                f.unlink()
        except OSError:
            pass


def run_scraper():
    if not _run_lock.acquire(blocking=False):
        return
    _last_run["running"] = True
    try:
        counts = scraper.run_and_save()
        _last_run.update(
            at=datetime.now(timezone.utc).isoformat(),
            count=counts,
            error=None,
        )
    except Exception as e:
        _last_run.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_run["running"] = False
        _run_lock.release()


def run_pricer(force_full=False):
    global _last_pricer_full_at
    if not _pricer_lock.acquire(blocking=False):
        return
    started = datetime.now(timezone.utc)
    # Due for a full pass? Time-based rather than a tick counter so a pass
    # skipped by the lock (or a restart) doesn't push the whole book out by
    # another interval. The 30s grace absorbs scheduler jitter, which would
    # otherwise slip every third tick to the one after it.
    due_full = force_full or _last_pricer_full_at is None or \
        (started - _last_pricer_full_at) >= (
            timedelta(minutes=PRICER_INTERVAL_MINUTES) - timedelta(seconds=30))
    _last_pricer["running"] = True
    try:
        counters = viagogo_pricer.run_pricer_tick(only_fast=not due_full)
        if due_full:
            _last_pricer_full_at = started
        _last_pricer.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=None,
            **counters,
        )
    except Exception as e:
        _last_pricer.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
    finally:
        _last_pricer["running"] = False
        _pricer_lock.release()


# Consecutive-failure alerting for the CV pricer: the tick runs unattended
# every 15 min, and a persistent failure (Cloudflare challenge on the parked
# tab, expired CrowdVolt login) otherwise sits invisible until the user
# happens to open /cvpricer — seen 2026-08-13, down for a day+ behind a
# Turnstile. One ping when a streak takes hold, then a reminder every 6h
# while it lasts. Streak of 2 so a one-off blip (SPA closed our tab, Chrome
# restarting) stays silent.
_CV_PRICER_ALERT_STREAK = 2
_CV_PRICER_ALERT_REMIND_HOURS = 6
_cv_pricer_alert = {"streak": 0, "last_ping": None}


def _cv_pricer_maybe_alert(error_text):
    st = _cv_pricer_alert
    if not error_text:
        st["streak"] = 0
        st["last_ping"] = None
        return
    st["streak"] += 1
    if st["streak"] < _CV_PRICER_ALERT_STREAK:
        return
    now = datetime.now(timezone.utc)
    if st["last_ping"] is not None and \
            (now - st["last_ping"]) < timedelta(hours=_CV_PRICER_ALERT_REMIND_HOURS):
        return
    st["last_ping"] = now
    try:
        notify.notify_cv_pricer_down(error_text, st["streak"])
    except Exception:
        traceback.print_exc()


def run_cv_pricer():
    if not _cv_pricer_lock.acquire(blocking=False):
        return
    _last_cv_pricer["running"] = True
    try:
        counters = crowdvolt_pricer.run_cv_pricer_tick()
        _last_cv_pricer.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=None,
            **counters,
        )
        _cv_pricer_maybe_alert(None)
    except Exception as e:
        _last_cv_pricer.update(
            at=datetime.now(timezone.utc).isoformat(),
            error=f"{type(e).__name__}: {e}",
        )
        traceback.print_exc()
        _cv_pricer_maybe_alert(f"{type(e).__name__}: {e}")
    finally:
        _last_cv_pricer["running"] = False
        _cv_pricer_lock.release()


def _enrich(rows):
    for r in rows:
        cost = r.get("total_cost")
        lst = r.get("total_list")
        r["profit_loss"] = round(lst - cost, 2) if cost is not None and lst is not None else None
    return rows


def _enrich_viagogo(rows):
    for r in rows:
        avail = r.get("available") or 0
        sold = r.get("sold") or 0
        fv = r.get("face_value") or 0
        r["cost"] = round(avail * fv, 2)
        r["sold_cost"] = round(sold * fv, 2)
    return rows


def _norm(s):
    return (s or "").strip().lower()


_DATE_FORMATS = (
    "%b %d, %Y",
    "%B %d, %Y",
    "%m/%d/%Y",
    "%Y-%m-%d",
    "%d %b %Y",
    "%d %B %Y",
    # Comma-less variants — e.g. Viagogo's "Jun 14 2026" after stripping the
    # leading day-of-week and trailing time.
    "%b %d %Y",
    "%B %d %Y",
)

_DAYNAMES = {"mon", "tue", "tues", "wed", "thu", "thur", "thurs", "fri", "sat", "sun",
             "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}


def _parse_event_date(text):
    if not text:
        return None
    head = (text or "").split("•")[0].split("@")[0].strip().rstrip(",").strip()
    head = head.split(" ")
    # Strip leading day-of-week tokens like "Sun", "Sun,", "Sunday" — Viagogo
    # dates frequently start with these and they aren't part of any strptime
    # format we use.
    while head and head[0].rstrip(",").lower() in _DAYNAMES:
        head.pop(0)
    # Strip trailing time tokens like "08:00PM" if they slipped in
    while head and any(c.isdigit() for c in head[-1]) and (":" in head[-1] or head[-1].lower().endswith(("am", "pm"))):
        head.pop()
    candidate = " ".join(head).strip(", ")
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(candidate, fmt).date().isoformat()
        except ValueError:
            continue
    import re
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def _date_only(iso):
    if not iso:
        return iso
    return iso.split("T", 1)[0]


def _resolve_iso(row):
    return _date_only(row.get("event_date_iso") or _parse_event_date(row.get("event_date")))


def _norm_venue(s):
    """Normalize a venue string for cluster/dedupe matching.
    Drops the trailing ", City" / ", City, State" suffix that some sources
    (Lysted in particular) tack on, strips a leading "The ", and lowercases.
    Lets us treat "The Eastern" / "The Eastern, Atlanta" as the same room
    while keeping "The Eastern" and "District Atlanta" — same artist, same
    night, two distinct venues — apart.
    """
    s = (s or "").lower()
    # Drop city/state suffix after the first comma.
    s = s.split(",", 1)[0]
    s = re.sub(r"^the\s+", "", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _event_key(event_name, iso, venue=None):
    return (_norm(event_name), iso or "", _norm_venue(venue))


_INV_NUMERIC = {"qty_unsold", "cost", "cost_per_unit", "list_price"}
_SALE_NUMERIC = {"qty", "sale_price", "cost"}


def _norm_event_name(s):
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _event_groups():
    """Build a canonical map for (event_name, iso, venue) → group key.

    Cluster fuzzy duplicates ("A Boogie Wit Da Hoodie" vs "A Boogie With Da
    Hoodie") so inventory rows and sales rows share the same group key. The
    UI uses this for its event-grouped click-to-expand behaviour.

    Venue is part of the cluster identity so an artist playing two distinct
    rooms on the same night (e.g. Disclosure at The Eastern AND District
    Atlanta on 2026-05-01) gets two separate groups instead of one merged
    bucket. Venue is normalized via _norm_venue so trailing ", City" and
    leading "The " variations don't fragment the cluster.
    """
    from difflib import SequenceMatcher
    triples = set()
    for r in db.all_lysted_purchases():
        triples.add((r.get("event_name") or "", _date_only(_resolve_iso(r) or ""), _norm_venue(r.get("venue"))))
    for r in db.all_lysted_sales():
        triples.add((r.get("event_name") or "", _date_only(r.get("event_date_iso") or ""), _norm_venue(r.get("venue"))))
    for r in db.all_viagogo():
        triples.add((r.get("event_name") or "", _date_only(_resolve_iso(r) or ""), _norm_venue(r.get("venue"))))
    for r in db.all_viagogo_sales():
        triples.add((r.get("event_name") or "", _date_only(r.get("event_date_iso") or ""), _norm_venue(r.get("venue"))))
    for r in db.all_jerujam_tickets():
        triples.add((r.get("event_name") or "", _date_only(r.get("event_date_iso") or ""), _norm_venue(r.get("venue"))))
    j_tix = {t["id"]: t for t in db.all_jerujam_tickets()}
    for s in db.all_jerujam_sales():
        t = j_tix.get(s.get("ticket_id"), {})
        triples.add((t.get("event_name") or "", _date_only(t.get("event_date_iso") or ""), _norm_venue(t.get("venue"))))
    for r in db.all_crowdvolt_sales():
        triples.add((r.get("event_name") or "", _date_only(r.get("event_date_iso") or ""), _norm_venue(r.get("venue"))))
    # Inventory aggregate is a fallback source for the listing-rows path —
    # include its venues here too so the group key for an aggregate row
    # matches the corresponding sales rows.
    for r in db.all_inventory():
        triples.add((r.get("event_name") or "", _date_only(_resolve_iso(r) or ""), _norm_venue(r.get("venue"))))

    def _substr_match(short, long):
        # Whitespace-bounded substring match. Used for both event names and
        # venues so "Masonic Temple" and "Masonic Temple - Temple Theatre"
        # cluster together while "the eastern" and "district atlanta" stay
        # apart.
        if not short or not long:
            return False
        if short == long:
            return True
        if len(short) < 3:
            return False
        return (
            f" {short} " in f" {long} "
            or long.startswith(short + " ")
            or long.endswith(" " + short)
        )

    def _venues_match(a, b):
        if a == b:
            return True
        if not a or not b:
            # Treat blank venue as a wildcard so a sales row missing venue
            # doesn't fragment off into its own cluster.
            return True
        short, long = (a, b) if len(a) <= len(b) else (b, a)
        return _substr_match(short, long)

    canonicals = []  # list[(norm_name, iso, norm_venue)]
    mapping = {}
    # Sort shorter names first so the canonical for a cluster is the
    # cleanest representation ("jigitz" anchors the cluster, then
    # "jigitz rescheduled from 3 7 26" attaches to it via substring match).
    for name, iso, venue in sorted(triples, key=lambda p: (len(_norm_event_name(p[0])), p[0])):
        norm = _norm_event_name(name)
        if not norm:
            mapping[(name, iso, venue)] = f"|{iso}|{venue}"
            continue
        found = None
        for cn, ciso, cvenue in canonicals:
            if ciso != iso:
                continue
            if not _venues_match(cvenue, venue):
                continue
            # Substring match (with whitespace boundary) catches
            # "jigitz" vs "jigitz rescheduled from 3 7 26" which the
            # SequenceMatcher ratio would otherwise miss because the
            # extra suffix tanks the similarity score.
            short, long = (cn, norm) if len(cn) <= len(norm) else (norm, cn)
            if (
                cn == norm
                or _substr_match(short, long)
                or SequenceMatcher(None, cn, norm).ratio() >= 0.85
            ):
                found = (cn, ciso, cvenue)
                break
        if found:
            mapping[(name, iso, venue)] = f"{found[0]}|{found[1]}|{found[2]}"
        else:
            canonicals.append((norm, iso, venue))
            mapping[(name, iso, venue)] = f"{norm}|{iso}|{venue}"

    # User-driven merges: overwrite the auto-cluster mapping for any raw
    # group_key the user explicitly merged. Multiple raw keys can map to
    # the same canonical key — that's how two unrelated names ("Ishay Ribo
    # — early" + "Ishay Ribo — National Library of Israel") get unified
    # into one display group.
    user_merges = db.all_event_group_merges()
    if user_merges:
        for triple, auto_key in list(mapping.items()):
            if auto_key in user_merges:
                mapping[triple] = user_merges[auto_key]["canonical_group_key"]
    return mapping


def _event_group_displays():
    """Returns {canonical_group_key → {event_name, event_date,
    event_date_iso, venue}} for each user-merged group. Used by row
    builders to overwrite the per-row display fields after grouping so the
    chosen canonical name shows everywhere the merged group appears."""
    out = {}
    for m in db.all_event_group_merges().values():
        out[m["canonical_group_key"]] = {
            "event_name": m.get("canonical_event_name"),
            "event_date": m.get("canonical_event_date"),
            "event_date_iso": m.get("canonical_event_date_iso"),
            "venue": m.get("canonical_venue"),
        }
    return out


def _apply_group_displays(rows, displays):
    """Overwrite event_name/date/venue on rows whose group is in a merge."""
    if not displays:
        return
    for r in rows:
        d = displays.get(r.get("event_group"))
        if not d:
            continue
        for f in ("event_name", "event_date", "event_date_iso", "venue"):
            if d.get(f):
                r[f] = d[f]


def _bought_by_event(groups_map):
    """Total tickets purchased per event_group. Each source gives an
    independent "they bought at least N" signal; we take the max so
    JeruJam-tracked inventory that's *also* listed/sold via Lysted doesn't
    double-count.

    Sources:
      - lysted_purchases.qty (when the order is still on the purchases page)
      - lysted_active = inventory.tickets_count + lysted_sales.qty  (covers
        events where the purchase record is missing — common after a few
        weeks once Lysted clears stale orders — but the tickets are clearly
        still live or were sold)
      - jerujam_tickets.quantity
      - manual_inventory.qty (unmatched)
      - viagogo: available + sold
    """
    per_source = {}  # {group_key: {source: total_qty_bought}}
    # Honor inventory_hidden here too — otherwise a user hiding a listing
    # via × would still see the bought-count include it, leaving the sales
    # page to render a phantom "(no detail)" row that's itself undeletable.
    hidden = db.all_hidden_keys()

    def _add(src, name, iso, venue, qty):
        key = _row_group(groups_map, name, _date_only(iso or ""), venue)
        per_source.setdefault(key, {})
        per_source[key][src] = per_source[key].get(src, 0) + int(qty or 0)

    for r in db.all_lysted_purchases():
        if ("lysted", str(r.get("id"))) in hidden:
            continue
        _add("lysted_purchases", r.get("event_name"), _resolve_iso(r), r.get("venue"), r.get("qty"))
    # Lysted's currently-active scraped inventory + recent sales tells us
    # the same event existed even if the purchases page rolled it off.
    today_iso = datetime.now(timezone.utc).date().isoformat()
    for r in db.all_inventory():
        if ("lysted", str(r.get("id"))) in hidden:
            continue
        # Skip stale inventory for past events. A show that already happened
        # can't have tickets "still listed", but Lysted keeps reporting a
        # nonzero tickets_count on sold-out past listings for a while. Counting
        # it makes bought = stale_listed + sold overcount the true buy and
        # spawns an undeletable phantom "still listed" row on the sales page.
        # Sales still feed lysted_active below, so bought stays >= sold here.
        # (Blank-date rows fall through and are still counted.)
        iso_i = _date_only(_resolve_iso(r) or "")
        if iso_i and iso_i < today_iso:
            continue
        _add("lysted_active", r.get("event_name"), _resolve_iso(r), r.get("venue"), r.get("tickets_count"))
    for r in db.all_lysted_sales():
        _add("lysted_active", r.get("event_name"), r.get("event_date_iso"), r.get("venue"), r.get("qty"))
    for r in db.all_jerujam_tickets():
        if ("jerujam", str(r.get("id"))) in hidden:
            continue
        _add("jerujam", r.get("event_name"), r.get("event_date_iso"), r.get("venue"), r.get("quantity"))
    for r in db.all_manual_inventory():
        if r.get("matched_source"):
            continue  # already represented by the matched lysted/viagogo row
        if ("manual", str(r.get("id"))) in hidden:
            continue
        _add("manual", r.get("event_name"), r.get("event_date_iso"), r.get("venue"), r.get("qty"))
    # Skip stale viagogo rows: nothing prunes viagogo_listings, so a listing
    # deleted/relisted on viagogo keeps its last snapshot forever and its
    # available+sold double-counts against the replacement listing, inflating
    # bought and spawning a phantom "still listed" row. A row absent from the
    # last couple of scrapes (48h at the hourly cadence) is treated as gone.
    all_vg = db.all_viagogo()
    latest_vg_seen = max((r.get("last_seen_at") or "" for r in all_vg), default="")
    vg_stale_cutoff = ""
    if latest_vg_seen:
        try:
            vg_stale_cutoff = (
                datetime.fromisoformat(latest_vg_seen) - timedelta(hours=48)
            ).isoformat()
        except ValueError:
            pass
    for r in all_vg:
        if ("viagogo", str(r.get("id"))) in hidden:
            continue
        if vg_stale_cutoff and (r.get("last_seen_at") or "") < vg_stale_cutoff:
            continue
        _add("viagogo", r.get("event_name"), _resolve_iso(r), r.get("venue"), (r.get("available") or 0) + (r.get("sold") or 0))

    out = {}
    for k, sources in per_source.items():
        out[k] = max(sources.values()) if sources else 0
    return out


def _group_cost_per_unit(groups_map):
    """{event_group_key: cost_per_ticket} from hand-entered whole-event costs.

    The manual escape hatch for when a platform doesn't carry per-ticket cost
    across (CrowdVolt resales of Dice / Lysted stock, mainly): the user says
    "the block was $800 for 8 tickets" once on the event group header, and
    every ticket in that group -- sold rows AND still-listed inventory rows --
    prices at $100.

    Denominator is the ticket_count the user entered; if they left it blank we
    fall back to the auto "tickets bought" count that the group header shows.
    Returns an unrounded per-unit so an $800/3 split doesn't drift; callers
    round the per-row product to cents.
    """
    entries = db.all_event_group_costs()
    if not entries:
        return {}
    bought = None
    out = {}
    for key, e in entries.items():
        denom = e.get("ticket_count") or 0
        if denom <= 0:
            if bought is None:
                bought = _bought_by_event(groups_map)
            denom = bought.get(key) or 0
        if denom <= 0:
            continue
        out[key] = (e.get("total_cost") or 0) / denom
    return out


def _row_group(mapping, name, iso, venue=None):
    nv = _norm_venue(venue)
    iso_d = _date_only(iso or "")
    return mapping.get((name or "", iso_d, nv), f"{_norm_event_name(name)}|{iso_d}|{nv}")


def _apply_overrides(row, overrides_for_row, numeric_fields):
    if not overrides_for_row:
        return
    edited = []
    for field, value in overrides_for_row.items():
        if field in numeric_fields:
            try:
                row[field] = float(value) if value not in (None, "") else None
            except (ValueError, TypeError):
                row[field] = value
        else:
            row[field] = value
        edited.append(field)
    row["_edited"] = edited


def _ga_like(text):
    t = (text or "").strip().lower()
    if not t:
        return False
    if "general admission" in t:
        return True
    tokens = t.replace("/", " ").split()
    return "ga" in tokens or "pit" in tokens


def _build_unified_inventory():
    """Combine unsold tickets from Lysted + Viagogo + JeruJam.

    JeruJam tickets are deduped against Lysted by (event, section, row) and
    against Viagogo by (event, section). The Viagogo dedupe is intentionally
    coarser since Viagogo listings don't expose row-level data.

    Rows whose (source, source_id) is in inventory_hidden are excluded — that's
    the soft-delete the dashboard uses.
    """
    hidden = db.all_hidden_keys()
    lysted = db.all_lysted_purchases()
    viagogo = _enrich_viagogo(db.all_viagogo())
    jerujam = db.all_jerujam_tickets()
    j_sales = db.all_jerujam_sales()
    groups = _event_groups()

    # Exact StubHub event URLs captured per event by the Lysted inventory-grid
    # scrape (scraper.py:_row_stubhub_url). Keyed by the same canonical group
    # key the UI uses so the grid event and its purchase-order rows line up
    # despite event-date string differences.
    stubhub_by_group = {}
    for inv in db.all_inventory():
        url = inv.get("stubhub_url")
        if url:
            g = _row_group(groups, inv.get("event_name"), inv.get("event_date_iso"), inv.get("venue"))
            stubhub_by_group[g] = url

    sold_per_ticket = {}
    for s in j_sales:
        sold_per_ticket[s["ticket_id"]] = sold_per_ticket.get(s["ticket_id"], 0) + (s.get("quantity") or 0)
    # Cross-source / manual matches keyed by (inv_source, str(inv_source_id)).
    # Same dict feeds:
    #   - the JeruJam branch (was the only one honoring matches before — kept
    #     for backwards compatibility via ext_matched_per_ticket below)
    #   - Lysted / Viagogo / manual branches (NEW: previously ignored matches,
    #     so "Mark as Loss" on a Viagogo or Lysted row left the row sitting
    #     in active inventory because no qty was subtracted).
    matched_qty = {}
    ext_matched_per_ticket = {}
    for m in db.all_matches():
        src = m.get("inv_source")
        sid = m.get("inv_source_id")
        q = m.get("qty_matched") or 0
        matched_qty[(src, str(sid))] = matched_qty.get((src, str(sid)), 0) + q
        if src == "jerujam":
            ext_matched_per_ticket[sid] = ext_matched_per_ticket.get(sid, 0) + q

    # Pool of auto-detected Lysted sales keyed by (event_group, section_norm, row_norm).
    # Lysted's purchase-orders scrape returns the original purchase qty unchanged
    # after partial sales, so without this we'd show "12 unsold" on a listing
    # where Lysted has separately reported 4 sales. Allocated greedily across
    # matching listings as we iterate them below.
    lysted_sales_pool = {}
    for s in db.all_lysted_sales():
        ek_g = _row_group(groups, s.get("event_name"), s.get("event_date_iso"), s.get("venue"))
        bucket = (ek_g, _norm(s.get("section")), _norm(s.get("row_label")))
        lysted_sales_pool[bucket] = lysted_sales_pool.get(bucket, 0) + (s.get("qty") or 0)

    rows = []

    lysted_keys = set()
    lysted_event_groups = set()  # canonical group-level dedup for inventory-aggregate fallback below
    lysted_ga_event_rows = set()  # (event_key, row) for any GA-style Lysted listing
    for r in lysted:
        if (r.get("status") or "").strip().lower() == "sold":
            continue
        iso = _resolve_iso(r)
        ek = _event_key(r.get("event_name"), iso, r.get("venue"))
        sec_n = _norm(r.get("section"))
        row_n = _norm(r.get("row_label"))
        lysted_keys.add((ek, sec_n, row_n))
        lysted_event_groups.add(_row_group(groups, r.get("event_name"), iso, r.get("venue")))
        if _ga_like(sec_n):
            lysted_ga_event_rows.add((ek, row_n))
            lysted_ga_event_rows.add((ek, ""))
        if ("lysted", str(r.get("id"))) in hidden:
            continue
        qty_full = r.get("qty") or 0
        consumed = matched_qty.get(("lysted", str(r.get("id"))), 0)
        # Greedy draw from the auto-sales pool for this (event, section, row) bucket.
        bucket = (
            _row_group(groups, r.get("event_name"), iso, r.get("venue")),
            sec_n,
            row_n,
        )
        pool_left = lysted_sales_pool.get(bucket, 0)
        auto_consumed = min(max(0, qty_full - consumed), pool_left)
        if auto_consumed > 0:
            lysted_sales_pool[bucket] = pool_left - auto_consumed
        qty_remaining = max(0, qty_full - consumed - auto_consumed)
        if qty_remaining <= 0:
            continue
        # Prorate cost to the unsold qty — a partially-sold purchase order
        # shouldn't count its full purchase cost as cash still tied up.
        cost_full = r.get("total_cost") or 0
        if qty_remaining < qty_full and qty_full:
            cpu = r.get("cost_per_unit")
            cost = round(cpu * qty_remaining if cpu else cost_full * qty_remaining / qty_full, 2)
        else:
            cost = cost_full
        rows.append({
            "source": "lysted",
            "source_id": r.get("id"),
            "event_name": r.get("event_name"),
            "event_date": r.get("event_date"),
            "event_date_iso": iso,
            "venue": r.get("venue"),
            "section": r.get("section"),
            "row": r.get("row_label"),
            "seats": r.get("seats"),
            "qty_unsold": qty_remaining,
            "cost": cost,
            "cost_per_unit": r.get("cost_per_unit"),
            "delivery_type": r.get("delivery_type"),
            "list_price": None,
            "status": r.get("status") or "active",
            "stubhub_url": stubhub_by_group.get(
                _row_group(groups, r.get("event_name"), iso, r.get("venue"))
            ),
        })

    # Pre-scan Viagogo + JeruJam canonical group keys so the inventory-aggregate
    # fallback below can dedup against them too. Use _row_group (the same
    # canonical clustering the UI uses) so cross-source venue-string variants
    # like "Mortgage Matchup Center" vs "Mortgage Matchup Center • Phoenix, AZ"
    # collapse to the same group instead of slipping past the dedup.
    viagogo_event_groups = {
        _row_group(groups, r.get("event_name"), _resolve_iso(r), r.get("venue"))
        for r in viagogo if (r.get("available") or 0) > 0
    }
    jerujam_event_groups = {
        _row_group(groups, t.get("event_name"), _resolve_iso(t), t.get("venue"))
        for t in jerujam if (t.get("status") or "").strip().lower() != "sold"
    }

    # Fallback: include the inventory-table aggregate for events not covered
    # by lysted_purchases above. Tickets you've listed on Lysted that don't
    # have a matching purchase order (e.g. purchased outside Lysted, or the
    # purchase row rolled off) only show up here. Without this fallback,
    # those listings are invisible on the sales page and undeletable via × —
    # the renderer falls through to a synthesized phantom row with no
    # source_id. The composite inventory.id is used as source_id so
    # inventory_hidden works the same way.
    for inv in db.all_inventory():
        iso_i = _resolve_iso(inv)
        eg = _row_group(groups, inv.get("event_name"), iso_i, inv.get("venue"))
        if eg in lysted_event_groups or eg in viagogo_event_groups or eg in jerujam_event_groups:
            continue
        qty = inv.get("tickets_count") or 0
        if qty <= 0:
            continue
        if ("lysted", str(inv.get("id"))) in hidden:
            continue
        consumed = matched_qty.get(("lysted", str(inv.get("id"))), 0)
        qty_remaining = max(0, qty - consumed)
        if qty_remaining <= 0:
            continue
        total_cost = inv.get("total_cost") or 0
        cost_per = round(total_cost / qty, 2) if qty else None
        rows.append({
            "source": "lysted",
            "source_id": inv.get("id"),
            "event_name": inv.get("event_name"),
            "event_date": inv.get("event_date"),
            "event_date_iso": iso_i,
            "venue": inv.get("venue"),
            "section": "",
            "row": "",
            "seats": "",
            "qty_unsold": qty_remaining,
            "cost": total_cost,
            "cost_per_unit": cost_per,
            "delivery_type": "",
            "list_price": inv.get("total_list"),
            "status": "active",
            "stubhub_url": inv.get("stubhub_url"),
        })

    viagogo_keys = set()
    viagogo_ga_events = set()  # event_keys for any GA-style Viagogo listing
    for r in viagogo:
        avail = r.get("available") or 0
        if avail <= 0:
            continue
        iso = _resolve_iso(r)
        ek = _event_key(r.get("event_name"), iso, r.get("venue"))
        sec_n = _norm(r.get("section"))
        viagogo_keys.add((ek, sec_n))
        if _ga_like(sec_n):
            viagogo_ga_events.add(ek)
        if ("viagogo", str(r.get("id"))) in hidden:
            continue
        consumed = matched_qty.get(("viagogo", str(r.get("id"))), 0)
        avail_remaining = max(0, avail - consumed)
        if avail_remaining <= 0:
            continue
        rows.append({
            "source": "viagogo",
            "source_id": r.get("id"),
            "event_name": r.get("event_name"),
            "event_date": r.get("event_date"),
            "event_date_iso": iso,
            "venue": r.get("venue"),
            "section": r.get("section"),
            "row": "",
            "seats": r.get("ticket_type"),
            "qty_unsold": avail_remaining,
            # face_value × remaining, so manual matches ("Mark as Loss" etc.)
            # release their share of cash tied — viagogo's own sales are
            # already excluded because the scrape's `available` drops on sale.
            "cost": round((r.get("face_value") or 0) * avail_remaining, 2),
            "cost_per_unit": r.get("face_value"),
            "delivery_type": r.get("ticket_type"),
            "list_price": (r.get("price") or 0) * avail_remaining,
            "status": r.get("visibility") or "active",
        })

    skipped_jerujam = 0
    for t in jerujam:
        if (t.get("status") or "").strip().lower() == "sold":
            continue
        qty = t.get("quantity") or 0
        sold = sold_per_ticket.get(t.get("id"), 0)
        ext_sold = ext_matched_per_ticket.get(t.get("id"), 0)
        remaining = max(0, qty - sold - ext_sold)
        if remaining <= 0:
            continue
        iso_j = _resolve_iso(t)
        ek = _event_key(t.get("event_name"), iso_j, t.get("venue"))
        sec = _norm(t.get("section"))
        row = _norm(t.get("row_label"))
        if (ek, sec, row) in lysted_keys:
            skipped_jerujam += 1
            continue
        # GA/PIT label variants — Lysted is authoritative when both sides are
        # any flavor of GA. JeruJam GA PIT (row blank) collapses into Lysted PIT G5.
        if _ga_like(sec) and ((ek, row) in lysted_ga_event_rows or (ek, "") in lysted_ga_event_rows):
            skipped_jerujam += 1
            continue
        if (ek, sec) in viagogo_keys:
            skipped_jerujam += 1
            continue
        if _ga_like(sec) and ek in viagogo_ga_events:
            skipped_jerujam += 1
            continue
        if ("jerujam", str(t.get("id"))) in hidden:
            continue
        cost_per = t.get("cost_per_ticket") or 0
        rows.append({
            "source": "jerujam",
            "source_id": t.get("id"),
            "event_name": t.get("event_name"),
            "event_date": t.get("event_date"),
            "event_date_iso": iso_j,
            "venue": t.get("venue"),
            "section": t.get("section"),
            "row": t.get("row_label"),
            "seats": t.get("seat_numbers"),
            "qty_unsold": remaining,
            "cost": round(cost_per * remaining, 2),
            "cost_per_unit": cost_per or None,
            "delivery_type": t.get("listing_platform") or "",
            "list_price": (t.get("listing_price") or None) and round((t.get("listing_price") or 0) * remaining, 2),
            "status": t.get("status") or "(none)",
        })

    # Pending (manually-added) tickets the user has but hasn't listed yet
    for m in db.all_manual_inventory():
        if m.get("matched_source"):
            continue  # already showed up on Lysted/Viagogo
        if ("manual", str(m.get("id"))) in hidden:
            continue
        qty = m.get("qty") or 0
        if qty <= 0:
            continue
        consumed = matched_qty.get(("manual", str(m.get("id"))), 0)
        qty = max(0, qty - consumed)
        if qty <= 0:
            continue
        cost_per = m.get("cost_per_unit")
        rows.append({
            "source": "manual",
            "source_id": str(m.get("id")),
            "event_name": m.get("event_name"),
            "event_date": m.get("event_date"),
            "event_date_iso": m.get("event_date_iso"),
            "venue": m.get("venue"),
            "section": m.get("section"),
            "row": m.get("row_label"),
            "seats": m.get("seats"),
            "qty_unsold": qty,
            "cost": round((cost_per or 0) * qty, 2) if cost_per else 0,
            "cost_per_unit": cost_per,
            "delivery_type": m.get("note") or "",
            "list_price": None,
            "status": "pending",
        })

    # DICE holdings — auto-recorded per account from forwarded DICE purchase
    # emails (dice_purchases). "Sold" here = qty linked to a Viagogo/CrowdVolt/
    # Lysted sale on the /dice page; transfers are deliberately excluded (a
    # resale-platform sale is the sale, even before the DICE transfer goes
    # out). Unsold = qty - linked-sold - any cross-source match. Costs are USD,
    # which the page already renders with '$'.
    dice_linked = db.dice_linked_qty_by_purchase()
    for p in db.dice_purchases_all():
        if ("dice", str(p.get("id"))) in hidden:
            continue
        qty = p.get("qty") or 0
        if qty <= 0:
            continue
        sold = dice_linked.get(p.get("id"), 0)
        consumed = matched_qty.get(("dice", str(p.get("id"))), 0)
        remaining = max(0, qty - sold - consumed)
        if remaining <= 0:
            continue  # fully sold via linked resale-platform sales
        cost_per = p.get("price_per_unit")
        rows.append({
            "source": "dice",
            "source_id": str(p.get("id")),
            "event_name": p.get("event_name"),
            "event_date": p.get("event_date_iso"),
            "event_date_iso": p.get("event_date_iso"),
            "venue": p.get("venue"),
            "section": p.get("ticket_type") or "",
            "row": "",
            "seats": "",
            "qty_unsold": remaining,
            "cost": round((cost_per or 0) * remaining, 2) if cost_per else 0,
            "cost_per_unit": cost_per,
            "delivery_type": p.get("account_email") or "DICE",
            "list_price": None,
            "status": "dice",
        })

    inv_overrides = db.all_inv_overrides()
    seats_sold_map = db.seats_sold_by_inv()
    group_cpu = _group_cost_per_unit(groups)
    for r in rows:
        ov = inv_overrides.get((r.get("source"), str(r.get("source_id"))))
        if ov:
            _apply_overrides(r, ov, _INV_NUMERIC)
        r["event_group"] = _row_group(groups, r.get("event_name"), r.get("event_date_iso"), r.get("venue"))
        # Same whole-event split the sales page applies -- unsold tickets in
        # the block are carried at the same per-ticket cost as the sold ones.
        cpu = group_cpu.get(r["event_group"])
        if cpu is not None and not (ov and ("cost" in ov or "cost_per_unit" in ov)):
            r["cost_per_unit"] = round(cpu, 2)
            r["cost"] = round(cpu * (r.get("qty_unsold") or 0), 2)
            r["cost_source"] = "event_split"
        r["seats_sold_already"] = seats_sold_map.get((r.get("source"), str(r.get("source_id"))), "")
    # Apply user-merged-group display overrides AFTER per-row overrides so a
    # merged event uses the canonical name even if a single row had its own
    # event_name override pre-merge.
    _apply_group_displays(rows, _event_group_displays())

    # "Didn't Sell" archive — filter rows whose content fingerprint is in
    # inventory_unsold. Done as a final pass so overrides are already applied
    # (so the fingerprint matches what was captured at archive time).
    unsold_fps = db.all_unsold_fingerprints()
    if unsold_fps:
        kept = []
        for r in rows:
            fp = db.unsold_fingerprint(
                r.get("source"), r.get("event_name"), r.get("event_date_iso"),
                r.get("section"), r.get("row"), r.get("seats"), r.get("qty_unsold"),
            )
            if fp in unsold_fps:
                continue
            kept.append(r)
        rows = kept
    return rows, skipped_jerujam


def _migrate_legacy_unsold_overrides():
    """One-shot migration: convert pre-existing inventory_overrides rows
    whose status field matches the "not sold" pattern into proper
    inventory_unsold archive entries.

    Idempotent — once an override is migrated and deleted, subsequent runs
    find no candidates. Safe to call on every startup.

    Returns the number of rows archived.
    """
    candidates = []
    with db.connect() as conn:
        for r in conn.execute(
            "SELECT source, source_id, value FROM inventory_overrides "
            "WHERE field = 'status'"
        ).fetchall():
            if r["value"] and _UNSOLD_RE.match(str(r["value"])):
                candidates.append((r["source"], r["source_id"]))
    if not candidates:
        return 0
    rows, _ = _build_unified_inventory()
    by_key = {(r.get("source"), str(r.get("source_id"))): r for r in rows}
    now_iso = datetime.now(timezone.utc).isoformat()
    archived = 0
    for src, sid in candidates:
        r = by_key.get((src, sid))
        if r:
            fp = db.unsold_fingerprint(
                src, r.get("event_name"), r.get("event_date_iso"),
                r.get("section"), r.get("row"), r.get("seats"), r.get("qty_unsold"),
            )
            snap = {
                "fingerprint": fp, "source": src, "source_id": sid,
                "event_name": r.get("event_name"),
                "event_date": r.get("event_date"),
                "event_date_iso": r.get("event_date_iso"),
                "venue": r.get("venue"),
                "section": r.get("section"),
                "row_label": r.get("row"),
                "seats": r.get("seats"),
                "qty": r.get("qty_unsold"),
                "cost": r.get("cost"),
                "cost_per_unit": r.get("cost_per_unit"),
                "list_price": r.get("list_price"),
                "delivery_type": r.get("delivery_type"),
                "note": "auto-migrated from legacy status='not sold' override",
            }
            db.mark_inventory_unsold(snap, now_iso)
            archived += 1
    # Drop the legacy overrides whether or not the row still exists — if the
    # source row is gone, the override is dead weight; if it was migrated,
    # the override would conflict with the displayed status next render.
    with db.connect() as conn:
        for src, sid in candidates:
            conn.execute(
                "DELETE FROM inventory_overrides "
                "WHERE source = ? AND source_id = ? AND field = 'status'",
                (src, sid),
            )
    return archived


def _matched_cost(matches_idx, jerujam_idx, sale_source, sale_id, qty):
    """If a sale is paired to a JeruJam inventory row, use that ticket's
    cost_per_ticket × qty as the sale's cost."""
    m = matches_idx.get((sale_source, str(sale_id)))
    if not m or m["inv_source"] != "jerujam":
        return None
    j = jerujam_idx.get(m["inv_source_id"])
    if not j:
        return None
    cpt = j.get("cost_per_ticket")
    if cpt is None:
        return None
    return round(cpt * (qty or 0), 2)


# Every source _build_combined_sales can emit, in the order the by-source
# rollups should present them. Keep in sync with the tail of that function —
# `manual` was missing from /profit's rollup while still counting toward the
# day and month rows, so the By Source table never summed to the monthly total.
_SALE_SOURCES = ("lysted", "viagogo", "jerujam", "crowdvolt", "manual")


def _sale_payout(s):
    """What actually lands in the bank for one combined-sale row: the
    platform's own payout where it reports one (Lysted), else the sale price.
    Mirrors _payout() in sales.html / profit.html so every view agrees."""
    p = s.get("payout")
    return (p if p is not None else s.get("sale_price")) or 0


def _build_combined_sales(only_canceled=False):
    """Combine sale events across sources. Sources differ in fidelity:
    JeruJam has per-sale rows; Lysted/Viagogo only expose aggregates so
    each sold-row contributes a single coarse entry.

    By default returns the active sales list (excludes both hidden and
    canceled). Pass only_canceled=True to invert the filter and return only
    rows in the canceled archive — used to power the "// CANCELED" section
    on the sales page.
    """
    out = []
    hidden_sales = db.all_hidden_sale_keys()
    canceled_sales = db.all_canceled_sale_keys()
    if only_canceled:
        sale_skip = hidden_sales  # keep canceled, drop hidden
    else:
        sale_skip = hidden_sales | canceled_sales

    def _skip_sale(key):
        if key in sale_skip:
            return True
        if only_canceled and key not in canceled_sales:
            return True
        return False

    j_tickets = {t["id"]: t for t in db.all_jerujam_tickets()}
    matches_idx = {(m["sale_source"], m["sale_id"]): m for m in db.all_matches()}
    # Sales linked to a DICE purchase on the /dice page carry their cost from
    # the purchase itself — no manual cost entry needed on this page.
    dice_costs = db.dice_cost_by_sale()
    jerujam_keys = set()
    for s in db.all_jerujam_sales():
        t = j_tickets.get(s.get("ticket_id"), {})
        jerujam_keys.add((
            _norm(t.get("event_name")),
            (s.get("sale_date") or "")[:10],
            s.get("quantity") or 0,
            round(s.get("sale_price") or 0, 0),
        ))
        if _skip_sale(("jerujam", str(s.get("id")))):
            continue
        sale_price = s.get("sale_price") or 0
        out.append({
            "source": "jerujam",
            "sale_id": str(s.get("id")),
            "order_id": "",
            "payout": sale_price,
            "sale_date": s.get("sale_date") or "",
            "sale_date_iso": (s.get("sale_date") or "")[:10],
            "event_name": t.get("event_name") or "",
            "event_date": t.get("event_date") or "",
            "event_date_iso": _date_only(t.get("event_date_iso") or _parse_event_date(t.get("event_date"))) or "",
            "venue": t.get("venue") or "",
            "section": t.get("section") or "",
            "row": t.get("row_label") or "",
            "platform": s.get("platform") or "",
            "qty": s.get("quantity") or 0,
            "sale_price": s.get("sale_price") or 0,
            "cost": round((t.get("cost_per_ticket") or 0) * (s.get("quantity") or 0), 2),
            "is_new": False,
        })

    purchases_by_id = {r.get("id"): r for r in db.all_lysted_purchases()}
    purchases_cost_by_event = {}
    for r in db.all_lysted_purchases():
        key = (r.get("event_name") or "", r.get("section") or "", r.get("row_label") or "")
        purchases_cost_by_event[key] = r.get("cost_per_unit")
    for r in db.all_lysted_sales():
        qty = r.get("qty") or 0
        # Lysted's API returns cost directly — that's the source of truth.
        cost = r.get("cost")
        if not cost:
            cost = dice_costs.get(("lysted", str(r.get("id")))) or cost
        if cost is None:
            # Fallback chain for older rows that haven't been re-scraped yet.
            cost_per = None
            match = purchases_by_id.get(r.get("id"))
            if match:
                cost_per = match.get("cost_per_unit")
            if cost_per is None:
                cost_per = purchases_cost_by_event.get(
                    (r.get("event_name") or "", r.get("section") or "", r.get("row_label") or "")
                )
            cost = round((cost_per or 0) * qty, 2) if cost_per is not None else 0
            if cost == 0:
                mc = _matched_cost(matches_idx, j_tickets, "lysted", r.get("id"), qty)
                if mc is not None:
                    cost = mc
        sale_iso_short = (r.get("sale_date_iso") or "")[:10]
        key = (_norm(r.get("event_name")), sale_iso_short, qty, round(r.get("sale_price") or 0, 0))
        if _skip_sale(("lysted", str(r.get("id")))):
            continue
        payout = r.get("payout") if r.get("payout") is not None else r.get("sale_price")
        out.append({
            "source": "lysted",
            "sale_id": str(r.get("id")),
            "order_id": str(r.get("order_id") or "").strip(),
            "payout": payout,
            "sale_date": r.get("sale_date") or "",
            "sale_date_iso": (r.get("sale_date_iso") or "")[:10],
            "event_name": r.get("event_name") or "",
            "event_date": r.get("event_date") or "",
            "event_date_iso": _date_only(r.get("event_date_iso") or _parse_event_date(r.get("event_date"))) or "",
            "venue": r.get("venue") or "",
            "section": r.get("section") or "",
            "row": r.get("row_label") or "",
            "platform": "lysted",
            "qty": qty,
            "sale_price": r.get("sale_price"),
            "cost": cost,
            "is_new": key not in jerujam_keys,
        })

    viagogo_listings = _enrich_viagogo(db.all_viagogo())
    listing_lookup = {}

    def _split_listing_section(text):
        """Viagogo listings cram "Section\\nRow X , Seat Y" into one cell.
        Pull out the section name and row number separately."""
        if not text:
            return ("", "")
        first_line = text.splitlines()[0].strip() if "\n" in text else text.strip()
        m = re.search(r"Row\s+(\S+)", text or "")
        row = m.group(1) if m else ""
        return (first_line, row)

    def _clean_row(text):
        if not text:
            return ""
        m = re.match(r"\s*(\S+)", text)
        return (m.group(1) if m else "").strip()

    for r in viagogo_listings:
        sec, row = _split_listing_section(r.get("section") or "")
        ev_n = _norm(r.get("event_name"))
        # Index by (event, section, row) — most specific
        listing_lookup.setdefault((ev_n, _norm(sec), _norm(row)), r)
        # Also keep an event+section fallback for sales that don't have a row
        listing_lookup.setdefault((ev_n, _norm(sec), ""), r)

    for r in db.all_viagogo_sales():
        ev_n = _norm(r.get("event_name"))
        sec_n = _norm(r.get("section"))
        row_n = _norm(_clean_row(r.get("row_label") or ""))
        listing = (listing_lookup.get((ev_n, sec_n, row_n))
                   or listing_lookup.get((ev_n, sec_n, "")))
        cost_per = listing.get("face_value") if listing else None
        qty = r.get("qty") or 0
        cost = round((cost_per or 0) * qty, 2) if cost_per is not None else 0
        dice_cost = dice_costs.get(("viagogo", str(r.get("id"))))
        if dice_cost:
            cost = dice_cost
        if cost == 0:
            mc = _matched_cost(matches_idx, j_tickets, "viagogo", r.get("id"), qty)
            if mc is not None:
                cost = mc
        key = (_norm(r.get("event_name")), (r.get("sale_date_iso") or "")[:10], qty, round(r.get("sale_price") or 0, 0))
        if _skip_sale(("viagogo", str(r.get("id")))):
            continue
        out.append({
            "source": "viagogo",
            "sale_id": str(r.get("id")),
            "order_id": str(r.get("order_id") or "").strip(),
            "payout": r.get("sale_price"),
            "sale_date": r.get("sale_date") or "",
            "sale_date_iso": (r.get("sale_date_iso") or "")[:10],
            "event_name": r.get("event_name") or "",
            "event_date": r.get("event_date") or "",
            "event_date_iso": _date_only(r.get("event_date_iso") or _parse_event_date(r.get("event_date"))) or "",
            "venue": r.get("venue") or "",
            "section": r.get("section") or "",
            "row": r.get("row_label") or "",
            "platform": "viagogo",
            "qty": qty,
            "sale_price": r.get("sale_price"),
            "cost": cost,
            "is_new": key not in jerujam_keys,
        })

    for r in db.all_crowdvolt_sales():
        qty = r.get("qty") or 0
        sale_price = r.get("sale_price") or 0
        key = (_norm(r.get("event_name")), (r.get("sale_date_iso") or "")[:10], qty, round(sale_price, 0))
        cost = dice_costs.get(("crowdvolt", str(r.get("id")))) or 0
        # CrowdVolt is a last-minute dump for tickets bought via Lysted or
        # Viagogo. Pull cost from a matching purchase/listing in those tables.
        cv_event = r.get("event_name") or ""
        cv_section = r.get("ticket_type") or ""
        cv_qty = qty
        if not cost:
            for lp in db.all_lysted_purchases():
                if not matcher._events_match(lp.get("event_name"), cv_event):
                    continue
                if (lp.get("qty") or 0) != cv_qty:
                    continue
                cost = lp.get("total_cost") or ((lp.get("cost_per_unit") or 0) * cv_qty)
                if cost:
                    break
        if not cost:
            for vl in viagogo_listings:
                if not matcher._events_match(vl.get("event_name"), cv_event):
                    continue
                v_sec = (vl.get("section") or "").splitlines()[0] if vl.get("section") else ""
                if cv_section and v_sec and not (
                    _norm(cv_section) == _norm(v_sec)
                    or (_ga_like(cv_section) and _ga_like(v_sec))
                ):
                    continue
                face = vl.get("face_value") or 0
                if face > 0:
                    cost = round(face * cv_qty, 2)
                    break
        if not cost:
            mc = _matched_cost(matches_idx, j_tickets, "crowdvolt", r.get("id"), qty)
            if mc is not None:
                cost = mc
        if _skip_sale(("crowdvolt", str(r.get("id")))):
            continue
        out.append({
            "source": "crowdvolt",
            "sale_id": str(r.get("id")),
            "order_id": str(r.get("order_id") or "").strip(),
            "payout": sale_price,
            "sale_date": r.get("sale_date") or "",
            "sale_date_iso": (r.get("sale_date_iso") or "")[:10],
            "event_name": r.get("event_name") or "",
            "event_date": r.get("event_date") or "",
            "event_date_iso": _date_only(r.get("event_date_iso")) or "",
            "venue": r.get("venue") or "",
            "section": r.get("ticket_type") or "",
            "row": "",
            "platform": "crowdvolt",
            "qty": qty,
            "sale_price": sale_price,
            "cost": cost,
            "is_new": key not in jerujam_keys,
        })

    for m in db.all_manual_sales():
        if _skip_sale(("manual", str(m.get("id")))):
            continue
        qty = m.get("qty") or 0
        sale_price = m.get("sale_price") or 0
        out.append({
            "source": "manual",
            "sale_id": str(m.get("id")),
            "order_id": "",
            "payout": sale_price,
            "sale_date": m.get("sale_date") or "",
            "sale_date_iso": (m.get("sale_date_iso") or "")[:10],
            "event_name": m.get("event_name") or "",
            "event_date": m.get("event_date") or "",
            "event_date_iso": _date_only(m.get("event_date_iso") or _parse_event_date(m.get("event_date"))) or "",
            "venue": m.get("venue") or "",
            "section": m.get("section") or "",
            "row": m.get("row_label") or "",
            "platform": m.get("platform") or "manual",
            "qty": qty,
            "sale_price": sale_price,
            "cost": m.get("cost") or 0,
            "is_loss": bool(m.get("is_loss")),
            "is_new": False,
        })

    sale_overrides = db.all_sale_overrides()
    groups = _event_groups()
    group_cpu = _group_cost_per_unit(groups)
    for r in out:
        ov = sale_overrides.get((r.get("source"), str(r.get("sale_id"))))
        if ov:
            _apply_overrides(r, ov, _SALE_NUMERIC)
        r["event_group"] = _row_group(groups, r.get("event_name"), r.get("event_date_iso"), r.get("venue"))
        # Whole-event cost split beats whatever cost the scrapers guessed --
        # but a cost the user typed on this specific row beats both.
        cpu = group_cpu.get(r["event_group"])
        if cpu is not None and not (ov and "cost" in ov):
            r["cost"] = round(cpu * (r.get("qty") or 0), 2)
            r["cost_per_unit"] = round(cpu, 2)
            r["cost_source"] = "event_split"
    _apply_group_displays(out, _event_group_displays())
    return out


@app.route("/")
def home():
    return render_template("inventory.html")


@app.route("/sw.js")
def service_worker():
    # Served from the root (not /static/) so its scope covers the whole app.
    return send_from_directory(app.static_folder, "sw.js", mimetype="application/javascript")


@app.route("/sources")
def dashboard():
    return render_template("dashboard.html")


@app.route("/pricer")
def pricer_page():
    return render_template("pricer.html")


@app.route("/cvpricer")
def cvpricer_page():
    return render_template("cvpricer.html")


@app.route("/cvfees")
def cvfees_page():
    return render_template("cvfees.html")


@app.route("/api/cvfees")
def api_cvfees():
    """Fee rows plus the groupings and rule-fit /cvfees renders.

    CrowdVolt publishes no seller fee schedule and the observed rate is not
    flat (3.3%-8.8%), so the point of this endpoint is to expose the raw
    numbers next to every grouping that might explain them, rather than to
    assert a formula.
    """
    import math

    rows = db.crowdvolt_fee_rows()
    total_gross = sum(r["gross"] for r in rows)
    total_fee = sum(r["fee"] or 0 for r in rows)
    total_payout = sum(r["payout"] or 0 for r in rows)

    def group(key):
        acc = {}
        for r in rows:
            k = r.get(key) or "(unknown)"
            g = acc.setdefault(k, {"key": k, "n": 0, "gross": 0.0,
                                   "fee": 0.0, "rates": []})
            g["n"] += 1
            g["gross"] += r["gross"]
            g["fee"] += r["fee"] or 0
            if r["rate"] is not None:
                g["rates"].append(r["rate"])
        out = []
        for g in acc.values():
            rates = g.pop("rates")
            g["rate"] = (g["fee"] / g["gross"]) if g["gross"] else None
            g["min_rate"] = min(rates) if rates else None
            g["max_rate"] = max(rates) if rates else None
            g["gross"] = round(g["gross"], 2)
            g["fee"] = round(g["fee"], 2)
            out.append(g)
        return sorted(out, key=lambda g: -g["n"])

    # Best-fitting simple rule. Reported with its miss count so a poor fit
    # reads as "no simple rule", which is the honest answer today.
    rounders = {"round": lambda x: float(round(x)),
                "ceil": lambda x: float(math.ceil(x - 1e-9)),
                "floor": lambda x: float(math.floor(x + 1e-9))}
    best = None
    for basis in ("per_ticket", "total"):
        for name, fn in rounders.items():
            for i in range(200, 1401):
                k = i / 20000.0
                hits = 0
                for r in rows:
                    price, qty = r["price_per_ticket"], r["qty"]
                    pred = (fn(price * k) * qty if basis == "per_ticket"
                            else fn(price * qty * k))
                    if abs(pred - (r["fee"] or 0)) < 0.001:
                        hits += 1
                if best is None or hits > best["hits"]:
                    best = {"hits": hits, "basis": basis, "rounding": name,
                            "rate": k}
    if best:
        best["pct"] = round(best["hits"] / len(rows) * 100, 1) if rows else 0

    hist = {}
    for r in rows:
        if r["rate"] is None:
            continue
        b = round(r["rate"] * 200) / 2.0
        hist[b] = hist.get(b, 0) + 1

    return jsonify({
        "rows": rows,
        "totals": {
            "orders": len(rows),
            "tickets": sum(r["qty"] or 0 for r in rows),
            "gross": round(total_gross, 2),
            "fee": round(total_fee, 2),
            "payout": round(total_payout, 2),
            "rate": (total_fee / total_gross) if total_gross else None,
            "min_rate": min((r["rate"] for r in rows if r["rate"] is not None),
                            default=None),
            "max_rate": max((r["rate"] for r in rows if r["rate"] is not None),
                            default=None),
        },
        "by_source": group("ticket_source"),
        "by_event": group("event_name"),
        "by_ticket_type": group("ticket_type"),
        "by_month": group("sale_date_iso") if False else _cvfees_by_month(rows),
        "fit": best,
        "histogram": [{"rate": k, "n": hist[k]} for k in sorted(hist)],
    })


def _cvfees_by_month(rows):
    acc = {}
    for r in rows:
        d = (r.get("sale_date_iso") or "")[:7] or "(unknown)"
        g = acc.setdefault(d, {"key": d, "n": 0, "gross": 0.0, "fee": 0.0})
        g["n"] += 1
        g["gross"] += r["gross"]
        g["fee"] += r["fee"] or 0
    out = []
    for g in acc.values():
        g["rate"] = (g["fee"] / g["gross"]) if g["gross"] else None
        g["gross"] = round(g["gross"], 2)
        g["fee"] = round(g["fee"], 2)
        out.append(g)
    return sorted(out, key=lambda g: g["key"])


@app.route("/sales")
def sales_page():
    return render_template("sales.html")


@app.route("/mockups")
def mockups_index():
    return render_template("mockups/index.html")


@app.route("/mockups/<name>")
def mockup(name):
    if name not in {"slate", "cocoa", "studio"}:
        return "Unknown mockup", 404
    return render_template(f"mockups/{name}.html")


@app.route("/api/inventory-all")
def api_inventory_all():
    rows, skipped = _build_unified_inventory()
    by_source = {"lysted": 0, "viagogo": 0, "jerujam": 0}
    manual_ids = [str(r.get("source_id")) for r in rows if r.get("source") == "manual"]
    atts_by_owner = db.list_attachments_for_owners("manual_inventory", manual_ids)
    for r in rows:
        if r.get("source") == "manual":
            r["attachments"] = atts_by_owner.get(str(r.get("source_id")), [])
        by_source[r["source"]] = by_source.get(r["source"], 0) + 1
    cost_by_source = {}
    for r in rows:
        cost_by_source[r["source"]] = cost_by_source.get(r["source"], 0) + (r["cost"] or 0)
    # DICE holdings aren't unified-inventory rows (they live on /dice), but
    # they're still cash tied up: avail = qty - user-matched resale sales
    # (transfers deliberately ignored, same as api_dice_purchases).
    dice_links = db.dice_sale_links_by_purchase()
    dice_cost = 0.0
    dice_tickets = 0
    for p in db.dice_purchases_all():
        qty = p.get("qty") or 0
        sold = sum(l.get("qty") or 0 for l in dice_links.get(p["id"], []))
        avail = max(0, qty - sold)
        if avail <= 0:
            continue
        ppu = p.get("price_per_unit")
        dice_cost += ppu * avail if ppu else (p.get("price_total") or 0) * (avail / qty if qty else 0)
        dice_tickets += avail
    if dice_tickets:
        cost_by_source["dice"] = dice_cost
    cost_by_source = {k: round(v, 2) for k, v in cost_by_source.items()}
    # Split lysted into listed (status Active — this is the number Lysted's own
    # dashboard shows) vs unlisted (Ready/Hold purchase orders: cash already
    # spent that Lysted doesn't count as inventory value).
    lysted_listed = sum(
        r["cost"] or 0 for r in rows
        if r["source"] == "lysted" and (r.get("status") or "").strip().lower() == "active"
    )
    lysted_split = {
        "listed": round(lysted_listed, 2),
        "unlisted": round(cost_by_source.get("lysted", 0) - lysted_listed, 2),
    }
    totals = {
        "rows": len(rows),
        "tickets": sum(r["qty_unsold"] for r in rows) + dice_tickets,
        "total_cost": round(sum(r["cost"] or 0 for r in rows) + dice_cost, 2),
        "cost_by_source": cost_by_source,
        "lysted_split": lysted_split,
        "dice_tickets": dice_tickets,
        "by_source": by_source,
        "jerujam_skipped_dedupe": skipped,
        "auto_matched": len(db.all_matches()),
    }
    return jsonify({"rows": rows, "totals": totals, "last_run": _last_run, "last_backup": _last_backup})


@app.route("/api/events/suggest")
def api_events_suggest():
    """Slim event suggestions for the Pending modal autocomplete. Aggregates
    raw (event_name, iso, venue) triples from every source so a typed prefix
    finds events the user has touched anywhere in the app.

    Returns up to 8 suggestions of {event_name, event_date_iso, venue}, sorted
    by date descending so upcoming/recent events surface first.
    """
    from flask import request
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"suggestions": []})
    qn = _norm_event_name(q)
    if not qn:
        return jsonify({"suggestions": []})

    raw = []
    for r in db.all_lysted_purchases():
        raw.append((r.get("event_name") or "", _date_only(_resolve_iso(r) or ""), r.get("venue") or ""))
    for r in db.all_viagogo():
        raw.append((r.get("event_name") or "", _date_only(_resolve_iso(r) or ""), r.get("venue") or ""))
    for r in db.all_jerujam_tickets():
        raw.append((r.get("event_name") or "", _date_only(r.get("event_date_iso") or ""), r.get("venue") or ""))
    for r in db.all_manual_inventory():
        raw.append((r.get("event_name") or "", _date_only(r.get("event_date_iso") or ""), r.get("venue") or ""))
    for r in db.all_inventory():
        raw.append((r.get("event_name") or "", _date_only(_resolve_iso(r) or ""), r.get("venue") or ""))

    # Dedupe by (norm_name, iso, norm_venue) and keep the longest raw name +
    # longest raw venue per cluster — that's our display representative.
    by_key = {}
    for name, iso, venue in raw:
        if not name:
            continue
        nn = _norm_event_name(name)
        if qn not in nn:
            continue
        key = (nn, iso, _norm_venue(venue))
        cur = by_key.get(key)
        if cur is None:
            by_key[key] = (name, iso, venue)
        else:
            cn, ci, cv = cur
            new_name = name if len(name) > len(cn) else cn
            new_venue = venue if len(venue or "") > len(cv or "") else cv
            by_key[key] = (new_name, ci, new_venue)

    items = list(by_key.values())
    items.sort(key=lambda t: t[1] or "", reverse=True)
    items = items[:8]
    return jsonify({
        "suggestions": [
            {"event_name": name, "event_date_iso": iso, "venue": venue}
            for (name, iso, venue) in items
        ]
    })


@app.route("/series")
def series_page():
    return render_template("series.html")


@app.route("/api/series")
def api_series():
    from flask import request
    name = (request.args.get("series") or "NEXT").strip()
    return jsonify(series.build(name))


@app.route("/api/series/purchase", methods=["POST"])
def api_series_purchase_add():
    """Add a block by hand. Email intake uses db.series_purchase_insert directly."""
    from flask import request
    body = request.get_json(force=True, silent=True) or {}
    required = ("event_date_iso", "qty", "account")
    missing = [k for k in required if not body.get(k)]
    if missing:
        return jsonify({"ok": False, "error": f"missing: {', '.join(missing)}"}), 400
    now = datetime.now(timezone.utc).isoformat()
    aliases = db.series_alias_map()
    qty = int(body.get("qty") or 0)
    unit = body.get("unit_cost")
    total = body.get("total_cost")
    if total in (None, "") and unit not in (None, ""):
        total = float(unit) * qty
    rec = {
        "series": (body.get("series") or "NEXT").strip(),
        "event_date_iso": body["event_date_iso"].strip(),
        "venue": (body.get("venue") or "").strip(),
        "section": (body.get("section") or "").strip(),
        "row_label": (body.get("row_label") or "").strip(),
        "seats": (body.get("seats") or "").strip(),
        "qty": qty,
        "unit_cost": float(unit) if unit not in (None, "") else None,
        "total_cost": float(total) if total not in (None, "") else None,
        "account": db.series_canonical_account(body["account"], aliases),
        "marketplace": (body.get("marketplace") or "viagogo").strip(),
        "listed": 1 if body.get("listed") else 0,
        "etickets": body.get("etickets"),
        "source": (body.get("source") or "manual").strip(),
        "intake_id": body.get("intake_id"),
        "note": (body.get("note") or "").strip(),
    }
    pid = db.series_purchase_insert(rec, now)
    return jsonify({"ok": True, "id": pid})


@app.route("/api/series/purchase/<int:pid>", methods=["POST"])
def api_series_purchase_update(pid):
    from flask import request
    body = request.get_json(force=True, silent=True) or {}
    allowed = ("event_date_iso", "venue", "section", "row_label", "seats", "qty",
               "unit_cost", "total_cost", "account", "marketplace", "listed",
               "etickets", "note")
    fields = {k: body[k] for k in allowed if k in body}
    if "account" in fields:
        fields["account"] = db.series_canonical_account(fields["account"])
    if not fields:
        return jsonify({"ok": False, "error": "nothing to update"}), 400
    db.series_purchase_update(pid, fields, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True})


@app.route("/api/series/purchase/<int:pid>/delete", methods=["POST"])
def api_series_purchase_delete(pid):
    db.series_purchase_delete(pid)
    return jsonify({"ok": True})


@app.route("/api/series/alias", methods=["POST"])
def api_series_alias():
    """Teach the page that two nicknames are one account, so the 9-cap counts
    them together."""
    from flask import request
    body = request.get_json(force=True, silent=True) or {}
    alias, canonical = (body.get("alias") or "").strip(), (body.get("canonical") or "").strip()
    if not alias or not canonical:
        return jsonify({"ok": False, "error": "alias and canonical required"}), 400
    now = datetime.now(timezone.utc).isoformat()
    db.series_alias_set(alias, canonical, now)
    # Re-point existing rows so the cap math updates immediately.
    for row in db.series_purchases_all((body.get("series") or "NEXT").strip()):
        if (row.get("account") or "").lower() == alias.lower():
            db.series_purchase_update(row["id"], {"account": canonical.lower()}, now)
    return jsonify({"ok": True})


@app.route("/profit")
def profit_page():
    return render_template("profit.html")


def _profit_response():
    """Build the same payload returned by /api/profit/daily. Reused by the
    Maaser summary so both pages agree on what 'profit' means.

    `profit` is PAYOUT − cost, not sale_price − cost: the platform's fee is
    money that never arrives, so counting it as profit overstated every
    rollup here (and the Maaser base) by the Lysted fee — roughly 8% of
    Lysted volume. `revenue` stays gross, so profit_pct reads as a net
    margin on gross sales. Matches _payout() in sales.html / profit.html.

    Adds month["expenses"] / month["net_profit"] derived from the expenses
    table (operating costs reduce the Maaser base per the user's policy)."""
    sales = _build_combined_sales()
    by_day = {}
    by_month = {}

    def _blank_bucket(**extra):
        b = {"count": 0, "qty": 0, "revenue": 0, "cost": 0, "profit": 0, "payout": 0}
        b.update(extra)
        return b

    # Known sources keep a fixed order (and a zero row) so the table doesn't
    # reshuffle as volume moves; anything unexpected still gets a row rather
    # than being dropped on the floor.
    by_source = {name: _blank_bucket() for name in _SALE_SOURCES}
    for s in sales:
        date = (s.get("sale_date_iso") or "")[:10]
        if not date or not date.startswith("20"):
            continue
        rev = s.get("sale_price") or 0
        cost = s.get("cost") or 0
        qty = s.get("qty") or 0
        # Payout comes off the sale row itself, NOT from a second pass over
        # db.all_lysted_sales() — that pass skipped the hidden/canceled filter
        # _build_combined_sales applies, so hidden and canceled Lysted sales
        # were landing in the payout column while contributing to no other
        # one. Reading it here also covers every source, matching the PAYOUT
        # metric box above these tables and the one on /sales.
        pay = _sale_payout(s)
        src = s.get("source") or "?"
        bucket = by_day.setdefault(date, _blank_bucket(
            date=date, by_source={name: 0 for name in _SALE_SOURCES}))
        bucket["count"] += 1
        bucket["qty"] += qty
        bucket["revenue"] += rev
        bucket["cost"] += cost
        bucket["profit"] += (pay - cost)
        bucket["payout"] += pay
        bucket["by_source"][src] = bucket["by_source"].get(src, 0) + 1
        # Per-source totals
        sb = by_source.setdefault(src, _blank_bucket())
        sb["count"] += 1
        sb["qty"] += qty
        sb["revenue"] += rev
        sb["cost"] += cost
        sb["profit"] += (pay - cost)
        sb["payout"] += pay

    # Build month rollup from days (sale-month: bucketed by when the
    # ticket was sold). This is the cash-flow-style view.
    for d, row in by_day.items():
        m = d[:7]  # YYYY-MM
        mb = by_month.setdefault(m, _blank_bucket(month=m))
        mb["count"] += row["count"]; mb["qty"] += row["qty"]
        mb["revenue"] += row["revenue"]; mb["cost"] += row["cost"]
        mb["profit"] += row["profit"]; mb["payout"] += row["payout"]

    # Build a parallel month rollup keyed by EVENT date — "May events"
    # means sales for shows that take place in May, regardless of when
    # the sale itself happened. This is the view the dashboard treats as
    # canonical for monthly profit (and what the Maaser obligation reads
    # from). Sales whose event_date is unknown fall back to sale_date so
    # totals still tie out across both views.
    by_month_event = {}
    for s in sales:
        ev_date = (s.get("event_date_iso") or "")[:10]
        if not ev_date or not ev_date.startswith("20"):
            ev_date = (s.get("sale_date_iso") or "")[:10]
        if not ev_date or not ev_date.startswith("20"):
            continue
        m_key = ev_date[:7]
        rev = s.get("sale_price") or 0
        cost = s.get("cost") or 0
        qty = s.get("qty") or 0
        # Same per-row payout as the sale-date view, bucketed by event month
        # so the two views stay comparable.
        pay = _sale_payout(s)
        mb = by_month_event.setdefault(m_key, _blank_bucket(month=m_key))
        mb["count"] += 1; mb["qty"] += qty
        mb["revenue"] += rev; mb["cost"] += cost
        mb["profit"] += (pay - cost)
        mb["payout"] += pay

    days = sorted(by_day.values(), key=lambda r: r["date"], reverse=True)
    months = sorted(by_month.values(), key=lambda r: r["month"], reverse=True)
    months_by_event = sorted(by_month_event.values(), key=lambda r: r["month"], reverse=True)

    # Round and add profit_pct
    def _finish(row):
        for k in ("revenue","cost","profit","payout"):
            row[k] = round(row[k] or 0, 2)
        rev = row["revenue"] or 0
        row["profit_pct"] = round((row["profit"] / rev) * 100, 1) if rev else None
        return row
    days = [_finish(r) for r in days]
    months = [_finish(r) for r in months]
    months_by_event = [_finish(r) for r in months_by_event]
    for k, v in by_source.items():
        _finish(v)

    totals = _finish({
        "count": sum(r["count"] for r in days),
        "qty": sum(r["qty"] for r in days),
        "revenue": sum(r["revenue"] for r in days),
        "cost": sum(r["cost"] for r in days),
        "profit": sum(r["profit"] for r in days),
        "payout": sum(r["payout"] for r in days),
    })

    # Layer in operating expenses at month-level (subscriptions + one-offs).
    # Expenses are tied to when they were paid (calendar month), not to any
    # particular event — so the same expense number applies to BOTH views.
    expenses_by_month = {}
    for e in db.all_expenses():
        mk = (e.get("date_iso") or "")[:7]
        if not mk:
            continue
        expenses_by_month[mk] = expenses_by_month.get(mk, 0) + (e.get("amount") or 0)
    # Cost of tickets that didn't sell, bucketed by the month the event
    # took place. For the by-sale-date monthly view this is treated as a
    # realized loss for that month — the show is over, the cost is sunk.
    unsold_cost_by_month = {}
    for u in db.all_unsold():
        mk = (u.get("event_date_iso") or "")[:7]
        if not mk:
            continue
        unsold_cost_by_month[mk] = unsold_cost_by_month.get(mk, 0) + (u.get("cost") or 0)
    for m in months:
        m["expenses"] = round(expenses_by_month.get(m["month"], 0), 2)
        m["unsold_cost"] = round(unsold_cost_by_month.get(m["month"], 0), 2)
        m["net_profit"] = round((m.get("profit") or 0) - m["expenses"] - m["unsold_cost"], 2)
    for m in months_by_event:
        m["expenses"] = round(expenses_by_month.get(m["month"], 0), 2)
        m["net_profit"] = round((m.get("profit") or 0) - m["expenses"], 2)
    totals_expenses = round(sum(expenses_by_month.values()), 2)
    totals_unsold = round(sum(unsold_cost_by_month.values()), 2)
    totals["expenses"] = totals_expenses
    totals["unsold_cost"] = totals_unsold
    totals["net_profit"] = round((totals.get("profit") or 0) - totals_expenses - totals_unsold, 2)

    return {
        "days": days,
        "months": months,                     # by sale date
        "months_by_event": months_by_event,   # by event date — canonical view
        "totals": totals,
        "by_source": by_source,
    }


@app.route("/api/profit/daily")
def api_profit_daily():
    return jsonify(_profit_response())


@app.route("/api/sales-all")
def api_sales_all():
    rows = _build_combined_sales()
    by_source = {name: 0 for name in _SALE_SOURCES}
    new_count = 0
    for r in rows:
        by_source[r["source"]] = by_source.get(r["source"], 0) + 1
        if r.get("is_new"):
            new_count += 1
    totals = {
        "rows": len(rows),
        "qty": sum(r["qty"] or 0 for r in rows),
        "revenue": round(sum((r["sale_price"] or 0) for r in rows), 2),
        "cost": round(sum((r["cost"] or 0) for r in rows), 2),
        "by_source": by_source,
        "new_since_jerujam": new_count,
    }
    # "Didn't Sell" archive — separate top-level array so existing rollups
    # (revenue, profit, maaser) stay clean. Frontend renders these in their
    # own section below the sold table.
    unsold_rows = db.all_unsold()
    unsold_totals = {
        "rows": len(unsold_rows),
        "qty": sum((u.get("qty") or 0) for u in unsold_rows),
        "cost": round(sum((u.get("cost") or 0) for u in unsold_rows), 2),
    }
    # Canceled-sale archive — same pattern. Each row carries the platform's
    # canceled_at + optional reason on top of the regular sale fields.
    canceled_rows = _build_combined_sales(only_canceled=True)
    canceled_meta = {(c["source"], c["sale_id"]): c for c in db.all_canceled_sales()}
    for cr in canceled_rows:
        meta = canceled_meta.get((cr.get("source"), str(cr.get("sale_id"))), {})
        cr["canceled_at"] = meta.get("canceled_at")
        cr["cancel_reason"] = meta.get("reason")
    canceled_totals = {
        "rows": len(canceled_rows),
        "qty": sum((c.get("qty") or 0) for c in canceled_rows),
        "lost_revenue": round(sum((c.get("sale_price") or 0) for c in canceled_rows), 2),
    }
    # Whole-event cost splits, keyed by event_group. The page uses these to
    # badge the group header and to price the synthesized "(no detail)"
    # still-listed rows that have no inventory row of their own.
    groups = _event_groups()
    bought = _bought_by_event(groups)
    group_costs = {}
    for key, e in db.all_event_group_costs().items():
        denom = e.get("ticket_count") or bought.get(key) or 0
        group_costs[key] = {
            "total_cost": e.get("total_cost"),
            "ticket_count": e.get("ticket_count"),
            "effective_count": denom,
            "per_unit": round((e.get("total_cost") or 0) / denom, 2) if denom else None,
            "note": e.get("note") or "",
        }
    return jsonify({
        "rows": rows,
        "totals": totals,
        "bought_by_event": bought,
        "group_costs": group_costs,
        "last_run": _last_run,
        "unsold": unsold_rows,
        "unsold_totals": unsold_totals,
        "canceled": canceled_rows,
        "canceled_totals": canceled_totals,
    })


@app.route("/api/inventory")
def api_inventory():
    rows = _enrich(db.all_inventory())
    totals = {
        "events": len(rows),
        "listings": sum(r.get("listings_count") or 0 for r in rows),
        "tickets": sum(r.get("tickets_count") or 0 for r in rows),
        "total_cost": round(sum(r.get("total_cost") or 0 for r in rows), 2),
        "total_list": round(sum(r.get("total_list") or 0 for r in rows), 2),
    }
    totals["total_pl"] = round(totals["total_list"] - totals["total_cost"], 2)
    return jsonify({"rows": rows, "totals": totals, "last_run": _last_run, "last_backup": _last_backup})


@app.route("/api/lysted-purchases")
def api_lysted_purchases():
    rows = db.all_lysted_purchases()
    def _live(r):
        s = (r.get("status") or "").strip().lower()
        return s != "sold"
    active = [r for r in rows if _live(r)]
    sold = [r for r in rows if not _live(r)]
    totals = {
        "tickets": sum(r.get("qty") or 0 for r in rows),
        "active_tickets": sum(r.get("qty") or 0 for r in active),
        "sold_tickets": sum(r.get("qty") or 0 for r in sold),
        "active_cost": round(sum(r.get("total_cost") or 0 for r in active), 2),
        "sold_cost": round(sum(r.get("total_cost") or 0 for r in sold), 2),
        "total_cost": round(sum(r.get("total_cost") or 0 for r in rows), 2),
    }
    return jsonify({"rows": rows, "totals": totals, "last_run": _last_run, "last_backup": _last_backup})


@app.route("/api/viagogo")
def api_viagogo():
    rows = _enrich_viagogo(db.all_viagogo())
    # Merge each row with its auto-pricer config so the dashboard renders
    # the AUTO/FLOOR/STATE controls without a second fetch.
    pricer_cfgs = db.pricer_config_all()
    for r in rows:
        cfg = pricer_cfgs.get(str(r.get("id"))) or {}
        r["pricer_enabled"] = bool(cfg.get("enabled"))
        r["pricer_floor"] = cfg.get("floor_price")
        r["pricer_paused"] = bool(cfg.get("paused"))
        r["pricer_paused_reason"] = cfg.get("paused_reason")
        r["pricer_last_set_price"] = cfg.get("last_set_price")
        r["pricer_no_drop_cap"] = bool(cfg.get("no_drop_cap"))
        for src_key, out_key in (("compete_sections", "pricer_compete_sections"),
                                 ("compete_include", "pricer_compete_include"),
                                 ("compete_exclude", "pricer_compete_exclude")):
            try:
                r[out_key] = json.loads(cfg.get(src_key) or "null")
            except (TypeError, ValueError):
                r[out_key] = None
    totals = {
        "listings": len(rows),
        "tickets_available": sum(r.get("available") or 0 for r in rows),
        "tickets_sold": sum(r.get("sold") or 0 for r in rows),
        "total_cost": round(sum(r.get("cost") or 0 for r in rows), 2),
        "sold_cost": round(sum(r.get("sold_cost") or 0 for r in rows), 2),
        "total_price": round(sum((r.get("price") or 0) * (r.get("available") or 0) for r in rows), 2),
        "total_proceeds": round(sum((r.get("proceeds") or 0) * (r.get("available") or 0) for r in rows), 2),
    }
    return jsonify({"rows": rows, "totals": totals, "last_run": _last_run, "last_backup": _last_backup})


@app.route("/api/sales/hide", methods=["POST"])
def api_sales_hide():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    if not source or not sale_id:
        return jsonify({"error": "source and sale_id required"}), 400
    db.hide_sale(source, sale_id, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True})


@app.route("/api/sales/unhide", methods=["POST"])
def api_sales_unhide():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    if not source or not sale_id:
        return jsonify({"error": "source and sale_id required"}), 400
    db.unhide_sale(source, sale_id)
    return jsonify({"ok": True})


@app.route("/api/sales/cancel", methods=["POST"])
def api_sales_cancel():
    """Mark a sale as canceled. Distinct from hide (× delete):
    - Sale moves out of active sales rollups (revenue / profit / maaser)
    - Surfaced separately on the sales page under "// CANCELED"
    - Any matcher pairing for this sale is cleared, and if the inventory
      row was auto-hidden because of the match, it is restored
    - Matcher blocklist gets the sale_id so the next pass doesn't re-pair
    The platform (Viagogo etc.) typically auto-relists the underlying
    ticket; the regular scrape picks that up — Kartis doesn't synthesize
    a new inventory row.
    """
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    reason = (body.get("reason") or "").strip() or None
    if not source or not sale_id:
        return jsonify({"error": "source and sale_id required"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    db.cancel_sale(source, sale_id, now_iso, reason=reason)
    # Match cleanup — find any inventory_matches row for this sale, drop it,
    # and unhide the matched inventory row (matcher.py:247 hides whole-qty
    # matches). Also blocklist the sale so the matcher doesn't immediately
    # re-pair on next run.
    matched_inv = []
    for m in db.all_matches():
        if m.get("sale_source") == source and m.get("sale_id") == sale_id:
            matched_inv.append((m.get("inv_source"), m.get("inv_source_id")))
    for inv_src, inv_sid in matched_inv:
        db.unhide_inventory(inv_src, inv_sid)
    db.delete_match(source, sale_id)
    db.add_blocklist(source, sale_id, now_iso)
    return jsonify({
        "ok": True,
        "source": source, "sale_id": sale_id,
        "released_inventory": [{"source": s, "source_id": i} for s, i in matched_inv],
    })


@app.route("/api/sales/uncancel", methods=["POST"])
def api_sales_uncancel():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    if not source or not sale_id:
        return jsonify({"error": "source and sale_id required"}), 400
    db.uncancel_sale(source, sale_id)
    return jsonify({"ok": True})


@app.route("/api/event-groups/cost", methods=["POST"])
def api_event_group_cost_set():
    """Set one whole-event cost and let Kartis split it per ticket.

    The manual path for stock whose per-ticket cost doesn't survive the hop
    between platforms -- e.g. a Dice or Lysted block relisted on CrowdVolt,
    where CrowdVolt has no idea what you paid. You enter what the block cost
    in total and how many tickets it covers; every sold row and every
    still-listed row in the group then carries total / count each.

    Body: {"group_key": "...", "total_cost": 800, "ticket_count": 8,
           "note": "optional"}
    ticket_count may be omitted/0 -- then the group's auto "tickets bought"
    count is used as the denominator.
    """
    from flask import request
    body = request.get_json(silent=True) or {}
    group_key = (body.get("group_key") or "").strip()
    if not group_key:
        return jsonify({"error": "group_key required"}), 400
    try:
        total_cost = float(body.get("total_cost"))
    except (TypeError, ValueError):
        return jsonify({"error": "total_cost must be a number"}), 400
    if total_cost < 0:
        return jsonify({"error": "total_cost must be >= 0"}), 400
    raw_count = body.get("ticket_count")
    ticket_count = None
    if raw_count not in (None, "", 0, "0"):
        try:
            ticket_count = int(raw_count)
        except (TypeError, ValueError):
            return jsonify({"error": "ticket_count must be a whole number"}), 400
        if ticket_count <= 0:
            return jsonify({"error": "ticket_count must be > 0"}), 400
    note = (body.get("note") or "").strip() or None
    db.set_event_group_cost(group_key, total_cost, ticket_count,
                            note, datetime.now(timezone.utc).isoformat())
    denom = ticket_count or _bought_by_event(_event_groups()).get(group_key) or 0
    return jsonify({
        "ok": True,
        "group_key": group_key,
        "total_cost": total_cost,
        "ticket_count": ticket_count,
        "effective_count": denom,
        "per_unit": round(total_cost / denom, 2) if denom else None,
    })


@app.route("/api/event-groups/cost/clear", methods=["POST"])
def api_event_group_cost_clear():
    """Drop the whole-event cost split; rows fall back to scraped cost."""
    from flask import request
    body = request.get_json(silent=True) or {}
    group_key = (body.get("group_key") or "").strip()
    if not group_key:
        return jsonify({"error": "group_key required"}), 400
    db.delete_event_group_cost(group_key)
    return jsonify({"ok": True, "group_key": group_key})


@app.route("/api/event-groups/merge", methods=["POST"])
def api_event_groups_merge():
    """Merge 2+ raw event groups into one canonical group with a chosen
    display name + date + venue. Subsequent inventory and sales rows that
    fall in any of the merged raw groups will display under the canonical
    name. Body:
        {
          "group_keys": ["<raw_key_1>", "<raw_key_2>", ...],
          "event_name": "Ishay Ribo at the National Library",
          "event_date_iso": "2026-06-12",       # optional
          "event_date": "Jun 12, 2026",         # optional display text
          "venue": "National Library of Israel" # optional
        }
    """
    from flask import request
    body = request.get_json(silent=True) or {}
    group_keys = body.get("group_keys") or []
    event_name = (body.get("event_name") or "").strip()
    if not isinstance(group_keys, list) or len(group_keys) < 2:
        return jsonify({"error": "group_keys must be a list of 2+ raw group keys"}), 400
    if not event_name:
        return jsonify({"error": "event_name required"}), 400
    iso = (body.get("event_date_iso") or "").strip()[:10]
    raw_date = (body.get("event_date") or "").strip()
    venue = (body.get("venue") or "").strip()
    canonical_group_key = f"{_norm_event_name(event_name)}|{iso}|{_norm_venue(venue)}"
    # Resolve already-merged group_keys: if the user selected a group that
    # was itself a canonical from a prior merge, replace it with all its raw
    # keys so we don't strand old mappings pointing at a now-orphan canonical.
    existing = db.all_event_group_merges()
    canonical_to_raws = {}
    for raw, m in existing.items():
        canonical_to_raws.setdefault(m["canonical_group_key"], []).append(raw)
    expanded = set()
    for k in group_keys:
        expanded.add(k)
        if k in canonical_to_raws:
            for raw in canonical_to_raws[k]:
                expanded.add(raw)
    now_iso = datetime.now(timezone.utc).isoformat()
    db.merge_event_groups(
        sorted(expanded), canonical_group_key,
        event_name, raw_date, iso, venue,
        now_iso,
    )
    # Carry any whole-event cost splits onto the new canonical key -- the old
    # keys stop being reachable after the merge, so a split left behind would
    # silently stop applying. Two merged blocks add up: $800/8 + $300/3 on the
    # same event becomes $1100/11.
    all_costs = db.all_event_group_costs()
    involved = [k for k in (expanded | {canonical_group_key} | set(group_keys)) if k in all_costs]
    merged_cost = None
    if involved and not (len(involved) == 1 and involved[0] == canonical_group_key):
        total = sum(all_costs[k].get("total_cost") or 0 for k in involved)
        counts = [all_costs[k].get("ticket_count") or 0 for k in involved]
        # Only keep an explicit denominator if every part had one; otherwise
        # fall back to the auto bought-count.
        count = sum(counts) if all(c > 0 for c in counts) else None
        notes = [all_costs[k].get("note") for k in involved if all_costs[k].get("note")]
        for k in involved:
            db.delete_event_group_cost(k)
        db.set_event_group_cost(canonical_group_key, total, count,
                                "; ".join(notes) or None, now_iso)
        merged_cost = {"total_cost": total, "ticket_count": count}
    return jsonify({
        "ok": True,
        "canonical_group_key": canonical_group_key,
        "merged_raw_keys": sorted(expanded),
        "carried_cost": merged_cost,
    })


@app.route("/api/event-groups/unmerge", methods=["POST"])
def api_event_groups_unmerge():
    """Drop a merge. Body either:
      {"canonical_group_key": "..."}  → splits the entire merged group
      {"raw_group_key": "..."}        → removes just one raw key from its merge
    """
    from flask import request
    body = request.get_json(silent=True) or {}
    canonical = (body.get("canonical_group_key") or "").strip()
    raw = (body.get("raw_group_key") or "").strip()
    if canonical:
        db.unmerge_canonical(canonical)
    elif raw:
        db.unmerge_event_group(raw)
    else:
        return jsonify({"error": "canonical_group_key or raw_group_key required"}), 400
    return jsonify({"ok": True})


@app.route("/api/event-groups/merges")
def api_event_groups_merges():
    """List all current merges grouped by canonical key."""
    merges = db.all_event_group_merges()
    by_canonical = {}
    for raw, m in merges.items():
        by_canonical.setdefault(m["canonical_group_key"], {
            "canonical_group_key": m["canonical_group_key"],
            "event_name": m.get("canonical_event_name"),
            "event_date": m.get("canonical_event_date"),
            "event_date_iso": m.get("canonical_event_date_iso"),
            "venue": m.get("canonical_venue"),
            "raw_keys": [],
        })["raw_keys"].append(raw)
    return jsonify({"merges": list(by_canonical.values())})


@app.route("/pending")
def pending_page():
    return render_template("pending.html")


@app.route("/listings")
def listings_page():
    return render_template("listings.html")


@app.route("/api/pending-intake")
def api_pending_intake():
    rows = db.all_pending_intake(status="new")
    ids = [r["id"] for r in rows]
    atts = db.list_attachments_for_owners("manual_intake", ids)
    for r in rows:
        r["attachments"] = atts.get(r["id"], [])
    return jsonify({
        "rows": rows,
        "last_intake": _last_intake,
    })


@app.route("/api/pending-intake/poll", methods=["POST"])
def api_pending_intake_poll():
    if _intake_lock.locked():
        return jsonify({"ok": False, "error": "already running"}), 429
    threading.Thread(target=run_mail_intake, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/pending-intake/confirm", methods=["POST"])
def api_pending_intake_confirm():
    """Promote a pending_intake row + its attachments to a manual_inventory
    row. The user can override any of the parsed fields in the request body
    (the UI sends the edited values straight from the modal). Files stay on
    disk where they were saved during intake — we just rename the on-disk
    folder and re-point the DB attachment rows at the new owner_id."""
    from flask import request
    body = request.get_json(silent=True) or {}
    intake_id = (body.get("id") or "").strip()
    if not intake_id:
        return jsonify({"error": "id required"}), 400
    intake = db.get_pending_intake(intake_id)
    if not intake:
        return jsonify({"error": "not found"}), 404
    event_name = (body.get("event_name") or intake.get("event_name") or "").strip()
    if not event_name:
        return jsonify({"error": "event_name required"}), 400
    try:
        qty = int(body.get("qty") or intake.get("qty") or 0)
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0:
        return jsonify({"error": "qty must be > 0"}), 400
    cost_per = body.get("cost_per_unit") if "cost_per_unit" in body else intake.get("cost_per_unit")
    try:
        cost_per = float(cost_per) if cost_per not in (None, "") else None
    except (TypeError, ValueError):
        cost_per = None
    iso = (body.get("event_date_iso") or intake.get("event_date_iso") or "").strip()
    new_id = "pending-" + uuid.uuid4().hex[:12]
    row = {
        "id": new_id,
        "event_name": event_name,
        "event_date": iso,
        "event_date_iso": iso,
        "venue": (body.get("venue") or intake.get("venue") or "").strip(),
        "section": (body.get("section") or intake.get("section") or "").strip(),
        "row_label": (body.get("row") or body.get("row_label") or intake.get("row_label") or "").strip(),
        "seats": (body.get("seats") or intake.get("seats") or "").strip(),
        "qty": qty,
        "cost_per_unit": cost_per,
        "note": (body.get("note") or "").strip(),
        "email": (body.get("email") or intake.get("email_from") or "").strip(),
    }
    now_iso = datetime.now(timezone.utc).isoformat()
    db.insert_manual_inventory(row, now_iso)
    # Move the on-disk folder so the new owner_id matches the file path.
    old_dir = attachments_mod.ATTACH_DIR / intake_id
    new_dir = attachments_mod.ATTACH_DIR / new_id
    if old_dir.exists():
        try:
            old_dir.rename(new_dir)
        except OSError:
            pass
    # Update each attachment row: change owner_type/owner_id + stored_path prefix.
    for a in db.list_attachments("manual_intake", intake_id):
        new_stored = a["stored_path"].replace(intake_id, new_id, 1)
        with db.connect() as conn:
            conn.execute(
                "UPDATE attachments SET owner_type=?, owner_id=?, stored_path=? WHERE id=?",
                ("manual_inventory", new_id, new_stored, a["id"]),
            )
    db.update_pending_intake(intake_id, {"status": "confirmed"})
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/pending-intake/reject", methods=["POST"])
def api_pending_intake_reject():
    """Drop the intake row + its attachments. Disk files are deleted too —
    rejection means we don't want them; if the user changes their mind they
    can re-forward the email."""
    from flask import request
    body = request.get_json(silent=True) or {}
    intake_id = (body.get("id") or "").strip()
    if not intake_id:
        return jsonify({"error": "id required"}), 400
    for a in db.list_attachments("manual_intake", intake_id):
        attachments_mod.delete(a["id"])
    db.delete_pending_intake(intake_id)
    intake_dir = attachments_mod.ATTACH_DIR / intake_id
    if intake_dir.exists():
        try:
            intake_dir.rmdir()
        except OSError:
            pass
    return jsonify({"ok": True})


# ---------------- viagogo draft-listing push (Kupat -> viagogo) ------------
# Kupat purchase emails get auto-matched to a viagogo event + auto-priced at
# 5x the FX-converted USD cost during intake (see mail_intake._push_kupat_to_
# viagogo), landing here as 'awaiting_approval'. Creating the actual draft
# listing always requires an explicit approve click — see viagogo_listing.
# create_draft_listing's docstring for why.

def _run_viagogo_approve(push_id, event_id, search_query, ticket_type, section,
                          available_tickets, website_price, face_value, row,
                          seat_from, seat_to, venue_for_map, kupat_section,
                          ticket_url=None, publish=False, proceeds=None,
                          pricer=None, upload_tickets=True,
                          listing_notes=None, split_preference=None):
    now = lambda: datetime.now(timezone.utc).isoformat()
    db.viagogo_push_update(push_id, {"status": "creating"}, now())
    try:
        ticket_pdfs = None
        ticket_problem = None
        ticket_skipped = bool(ticket_url) and not upload_tickets
        if ticket_url and upload_tickets:
            try:
                _push = db.viagogo_push_get(push_id) or {}
                _qty = _push.get("qty") or available_tickets or 1
                # Seat-match the tickets out of the order: essential when one
                # order was split into multiple listings (different sections/
                # rows share one ticket link), a no-op identity check otherwise.
                _seats = viagogo_listing.seats_range(_push.get("seats") or "")
                ticket_pdfs = viagogo_listing.download_ticket_pdfs_for(
                    ticket_url, qty=int(_qty),
                    seats=_seats or None,
                    row=(_push.get("row_label") or "").strip() or None,
                )
            except Exception as e:
                # Never upload questionable tickets; list without them and
                # say so loudly on the card.
                ticket_problem = f"ticket download/validation failed: {type(e).__name__}: {e}"
                traceback.print_exc()
        result = viagogo_listing.create_draft_listing(
            event_id=event_id, search_query=search_query, ticket_type=ticket_type,
            section=section, available_tickets=available_tickets,
            website_price=website_price, face_value=face_value, proceeds=proceeds,
            row=row, seat_from=seat_from, seat_to=seat_to,
            ticket_pdfs=ticket_pdfs, publish=publish,
            listing_notes=listing_notes, split_preference=split_preference,
        )
        if ticket_pdfs and not result.get("tickets_uploaded"):
            ticket_problem = (f"ticket upload failed: "
                              f"{result.get('ticket_upload_error') or 'unknown'}")
        listing_id = result.get("listing_id")
        if listing_id:
            db.viagogo_push_update(push_id, {"listing_id": str(listing_id)}, now())
        # Optional auto-pricer enrollment, straight from the approve card —
        # saves the round-trip to /pricer after every listing.
        pricer_note = None
        if pricer and pricer.get("enabled"):
            if not listing_id:
                pricer_note = ("auto-pricer NOT enrolled: listing id could not "
                               "be resolved — enable it on /pricer manually")
            else:
                try:
                    secs = sorted({viagogo_pricer._norm_section(s)
                                   for s in (pricer.get("compete_sections") or [])
                                   if s and s.strip()})
                    db.pricer_config_set(str(listing_id), {
                        "enabled": 1,
                        "floor_price": float(pricer["floor_price"]),
                        "compete_sections": json.dumps(secs) if secs else None,
                        "last_set_price": None, "last_set_at": None,
                        "paused": 0, "paused_reason": None, "paused_at": None,
                    }, now())
                except Exception as e:
                    pricer_note = f"auto-pricer enrollment failed: {type(e).__name__}: {e}"
        if venue_for_map and kupat_section and section:
            db.viagogo_section_map_set(venue_for_map, kupat_section, section, now())
        # Teach Hebrew→English name mapping so future emails auto-match.
        push_row = db.viagogo_push_get(push_id) or {}
        hebrew_event = push_row.get("event_name") or ""
        english_event = push_row.get("chosen_event_name") or ""
        if hebrew_event and english_event and hebrew_event != english_event:
            db.kupat_name_map_set(hebrew_event, english_event, now())
        # Success clears stale error text from earlier failed attempts; a
        # ticket problem replaces it so the card shows the listing exists
        # but NEEDS TICKETS before a sale can be fulfilled.
        notes = []
        if ticket_problem:
            notes.append(f"NO TICKETS ATTACHED — {ticket_problem}")
        elif ticket_skipped:
            # Deliberate, not a failure — but still say it, since a listing
            # can't be fulfilled until the tickets are on it.
            notes.append("tickets not uploaded by choice — attach them before it sells")
        if pricer_note:
            notes.append(pricer_note)
        # Notes/split that didn't stick are warnings, not failures — the
        # listing exists; fix them on inv.viagogo before uploading tickets.
        for w in result.get("option_warnings") or []:
            notes.append(f"OPTION NOT APPLIED — {w}")
        db.viagogo_push_update(push_id, {
            "status": "listed" if publish else "created",
            "viagogo_section": section,
            "error": ("; ".join(notes)[:500] if notes else None),
        }, now())
        # Heads-up ping now that the listing exists on viagogo.
        try:
            pr_cfg = db.pricer_config_get(str(listing_id)) if listing_id else None
            pr_secs = None
            if pr_cfg and pr_cfg.get("compete_sections"):
                try:
                    pr_secs = json.loads(pr_cfg["compete_sections"])
                except Exception:
                    pr_secs = None
            notify.notify_viagogo_listed({
                "event_name": push_row.get("chosen_event_name") or push_row.get("event_name"),
                "venue": push_row.get("chosen_venue") or push_row.get("venue"),
                "event_date": push_row.get("chosen_event_date") or push_row.get("event_date_iso"),
                "section": section,
                "row": row,
                "seats": push_row.get("seats"),
                "qty": available_tickets,
                "price": website_price,
                "currency": "USD",
                "published": publish,
                "tickets_uploaded": (result.get("tickets_uploaded") if ticket_pdfs else None),
                "ticket_note": (ticket_problem or
                                ("not uploaded by choice — attach before it sells"
                                 if ticket_skipped else None)),
                "tickets_skipped": ticket_skipped,
                "pricer_enabled": bool(pr_cfg and pr_cfg.get("enabled")),
                "pricer_floor": (pr_cfg or {}).get("floor_price"),
                "pricer_sections": pr_secs,
                "buyer_email": push_row.get("buyer_email"),
                "listing_id": listing_id,
            })
        except Exception:
            traceback.print_exc()
    except Exception as e:
        db.viagogo_push_update(push_id, {"status": "error", "error": f"{type(e).__name__}: {e}"}, now())
        traceback.print_exc()


@app.route("/api/viagogo-push")
def api_viagogo_push():
    from flask import request
    status = (request.args.get("status") or "").strip() or None
    rows = db.viagogo_push_all(status=status)
    for r in rows:
        try:
            r["candidates"] = json.loads(r["candidates_json"]) if r.get("candidates_json") else []
        except Exception:
            r["candidates"] = []
        venue_for_map = r.get("chosen_venue") or r.get("venue") or ""
        r["suggested_viagogo_section"] = (
            db.viagogo_section_map_get(venue_for_map, r.get("section") or "")
            if venue_for_map and r.get("section") else None
        )
    return jsonify({"rows": rows})


# Simple in-process cache: (event_id, ticket_type) -> list[str]
_sections_cache: dict = {}


@app.route("/api/viagogo-sections")
def api_viagogo_sections():
    from flask import request
    event_id = (request.args.get("event_id") or "").strip()
    search_query = (request.args.get("search_query") or "").strip()
    ticket_type = (request.args.get("ticket_type") or "E-Tickets").strip()
    if not event_id or not search_query:
        return jsonify({"error": "event_id and search_query required"}), 400
    cache_key = (event_id, ticket_type)
    if cache_key in _sections_cache:
        return jsonify({"sections": _sections_cache[cache_key], "cached": True})
    try:
        sections = viagogo_listing.fetch_sections(event_id, search_query, ticket_type)
        _sections_cache[cache_key] = sections
        return jsonify({"sections": sections})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/viagogo-push/approve", methods=["POST"])
def api_viagogo_push_approve():
    from flask import request
    body = request.get_json(silent=True) or {}
    push_id = (body.get("id") or "").strip()
    if not push_id:
        return jsonify({"error": "id required"}), 400
    push = db.viagogo_push_get(push_id)
    if not push:
        return jsonify({"error": "not found"}), 404
    if push["status"] not in ("awaiting_approval", "error"):
        return jsonify({"error": f"cannot approve from status '{push['status']}'"}), 400
    ticket_url = (push.get("ticket_url") or "").strip() or None
    event_id = (body.get("event_id") or push.get("chosen_event_id") or "").strip()
    if not event_id:
        return jsonify({"error": "event_id required (no matched event)"}), 400
    section = (body.get("section") or "").strip()
    if not section:
        return jsonify({"error": "section required"}), 400
    ticket_type = (body.get("ticket_type") or "E-Tickets").strip()
    try:
        available_tickets = int(body.get("available_tickets") or push.get("qty") or 0)
    except (TypeError, ValueError):
        available_tickets = 0
    if available_tickets <= 0:
        return jsonify({"error": "available_tickets must be > 0"}), 400
    try:
        website_price = float(body.get("website_price") or push.get("website_price_usd") or 0)
    except (TypeError, ValueError):
        website_price = 0
    if website_price <= 0:
        return jsonify({"error": "website_price must be > 0"}), 400
    try:
        fv = body.get("face_value")
        face_value = float(fv) if fv not in (None, "") else float(push.get("cost_usd_per_ticket") or 0)
    except (TypeError, ValueError):
        face_value = 0
    if face_value <= 0:
        return jsonify({"error": "face_value must be > 0"}), 400
    publish = bool(body.get("publish"))
    # Default ON: attaching tickets is the norm; the card can turn it off to
    # list now and upload manually later.
    upload_tickets = bool(body.get("upload_tickets", True))
    try:
        pr = body.get("proceeds")
        proceeds = float(pr) if pr not in (None, "") else None
    except (TypeError, ValueError):
        proceeds = None
    pricer_opts = None
    _p = body.get("pricer") or {}
    if _p.get("enabled"):
        try:
            _floor = float(_p.get("floor_price") or 0)
        except (TypeError, ValueError):
            _floor = 0
        if _floor <= 0:
            return jsonify({"error": "a floor price > 0 is required to enable auto-pricing"}), 400
        _secs = _p.get("compete_sections") or []
        if not isinstance(_secs, list):
            return jsonify({"error": "compete_sections must be a list"}), 400
        pricer_opts = {"enabled": True, "floor_price": _floor,
                       "compete_sections": [str(s) for s in _secs]}
    _notes_in = body.get("listing_notes") or []
    if not isinstance(_notes_in, list):
        return jsonify({"error": "listing_notes must be a list"}), 400
    listing_notes = [str(n) for n in _notes_in]
    _bad = [n for n in listing_notes if n not in viagogo_listing.LISTING_NOTE_OPTIONS]
    if _bad:
        return jsonify({"error": f"unknown listing notes: {_bad}"}), 400
    split_preference = (body.get("split_preference") or "").strip() or None
    if split_preference and split_preference not in viagogo_listing.SPLIT_TYPES:
        return jsonify({"error": f"unknown split_preference '{split_preference}'"}), 400
    row = (body.get("row") or push.get("row_label") or "").strip() or None
    seat_from = (body.get("seat_from") or "").strip() or None
    seat_to = (body.get("seat_to") or "").strip() or None
    if not seat_from and not seat_to and push.get("seats"):
        m = re.search(r"(\d+)\s*-\s*(\d+)", push["seats"])
        if m:
            seat_from, seat_to = m.group(1), m.group(2)
        elif push["seats"].strip().isdigit():
            seat_from = push["seats"].strip()
    search_query = push.get("chosen_event_name") or push.get("event_name") or push.get("venue") or ""
    venue_for_map = push.get("chosen_venue") or push.get("venue") or ""
    kupat_section = push.get("section") or ""
    threading.Thread(
        target=_run_viagogo_approve,
        args=(push_id, event_id, search_query, ticket_type, section, available_tickets,
              website_price, face_value, row, seat_from, seat_to, venue_for_map, kupat_section,
              ticket_url, publish, proceeds, pricer_opts, upload_tickets),
        kwargs={"listing_notes": listing_notes,
                "split_preference": split_preference},
        daemon=True,
    ).start()
    return jsonify({"ok": True, "status": "creating"})


@app.route("/api/viagogo-push/reject", methods=["POST"])
def api_viagogo_push_reject():
    from flask import request
    body = request.get_json(silent=True) or {}
    push_id = (body.get("id") or "").strip()
    if not push_id:
        return jsonify({"error": "id required"}), 400
    db.viagogo_push_update(push_id, {"status": "rejected"}, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True})


@app.route("/api/viagogo-push/set-event", methods=["POST"])
def api_viagogo_push_set_event():
    """Point a push at a viagogo event by pasted public URL (.../E-<id>) or
    bare id — for shows the picker search didn't surface as candidates.
    Body: {id, url}. Drives the live browser (~15s) to verify the seller
    picker can actually see the event."""
    from flask import request
    body = request.get_json(silent=True) or {}
    push_id = (body.get("id") or "").strip()
    url = (body.get("url") or "").strip()
    if not push_id or not url:
        return jsonify({"error": "id and url required"}), 400
    try:
        candidate = mail_intake.set_push_event_from_url(push_id, url)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 502
    return jsonify({"ok": True, "candidate": candidate})


@app.route("/api/kupat-name-map", methods=["POST"])
def api_kupat_name_map():
    """Teach a Hebrew→English name mapping and optionally re-run the search
    for an existing no_match push row. Body: {hebrew, english, push_id?}."""
    from flask import request
    import threading
    body = request.get_json(silent=True) or {}
    hebrew = (body.get("hebrew") or "").strip()
    english = (body.get("english") or "").strip()
    if not hebrew or not english:
        return jsonify({"error": "hebrew and english required"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()

    # A pasted viagogo event URL beats a name: it carries the exact event id
    # (…/Gordon-Fight-Night-Tickets/E-161493942) plus a searchable name in
    # the slug. Event-specific, so it is NOT saved as a name mapping — a
    # generic sender name (e.g. tickchak's "Tickchak4u") must not get
    # permanently bound to one event.
    force_term = force_event_id = None
    m = re.search(r"viagogo\.[^\s/]+/.*E-(\d+)", english, re.I)
    if m:
        force_event_id = m.group(1)
        slug = re.search(r"/([^/]+?)(?:-Tickets)?/E-\d+", english)
        force_term = (slug.group(1).replace("-", " ").strip() if slug else "")
        if not force_term:
            return jsonify({"error": "couldn't read an event name from that link"}), 400
    else:
        db.kupat_name_map_set(hebrew, english, now_iso)

    push_id = (body.get("push_id") or "").strip()
    if push_id:
        push = db.viagogo_push_get(push_id)
        if push and push.get("status") in ("no_match", "error"):
            # Re-run the search in background with the new mapping / link.
            def _retry(pid, fields, now, term, evid):
                import mail_intake as _mi
                _mi._push_kupat_to_viagogo_update(pid, fields, now,
                                                  force_term=term,
                                                  force_event_id=evid)
            fields = {k: push.get(k) for k in
                      ("event_name", "venue", "event_date_iso", "section",
                       "row_label", "seats", "qty", "cost", "cost_per_unit",
                       "ticket_url")}
            threading.Thread(
                target=_retry,
                args=(push_id, fields, now_iso, force_term, force_event_id),
                daemon=True,
            ).start()
    return jsonify({"ok": True})


@app.route("/api/inventory/manual-add", methods=["POST"])
def api_inventory_manual_add():
    from flask import request
    import uuid
    body = request.get_json(silent=True) or {}
    event_name = (body.get("event_name") or "").strip()
    if not event_name:
        return jsonify({"error": "event_name required"}), 400
    try:
        qty = int(body.get("qty") or 0)
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0:
        return jsonify({"error": "qty must be > 0"}), 400
    cost_per = body.get("cost_per_unit")
    try:
        cost_per = float(cost_per) if cost_per not in (None, "") else None
    except (TypeError, ValueError):
        cost_per = None
    event_date = (body.get("event_date") or "").strip()
    iso = (body.get("event_date_iso") or "").strip()
    if not iso and event_date:
        iso = _date_only(_parse_event_date(event_date)) or ""
    row = {
        "id": "pending-" + uuid.uuid4().hex[:12],
        "event_name": event_name,
        "event_date": event_date,
        "event_date_iso": iso,
        "venue": (body.get("venue") or "").strip(),
        "section": (body.get("section") or "").strip(),
        "row_label": (body.get("row") or "").strip(),
        "seats": (body.get("seats") or "").strip(),
        "qty": qty,
        "cost_per_unit": cost_per,
        "note": (body.get("note") or "").strip(),
        "email": (body.get("email") or "").strip(),
    }
    db.insert_manual_inventory(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/inventory/manual-delete", methods=["POST"])
def api_inventory_manual_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_manual_inventory(id_)
    # Drop the DB attachment rows but leave the files on disk so they
    # survive in OneDrive sync (per user preference).
    db.delete_attachments_for_owner("manual_inventory", id_)
    return jsonify({"ok": True})


@app.route("/api/inventory/manual-attach", methods=["POST"])
def api_inventory_manual_attach():
    from flask import request
    ticket_id = (request.form.get("ticket_id") or "").strip()
    if not ticket_id:
        return jsonify({"error": "ticket_id required"}), 400
    # Confirm the pending row actually exists so we don't accept uploads for
    # arbitrary owner_ids.
    existing = {m["id"] for m in db.all_manual_inventory()}
    if ticket_id not in existing:
        return jsonify({"error": "unknown ticket_id"}), 404
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "file required"}), 400
    # Probe size cheaply via stream seek; werkzeug's FileStorage wraps a
    # SpooledTemporaryFile so this is fine even for the 25 MB ceiling.
    f.stream.seek(0, 2)
    size = f.stream.tell()
    f.stream.seek(0)
    if size > attachments_mod.MAX_BYTES:
        return jsonify({"error": f"file too large (>{attachments_mod.MAX_BYTES} bytes)"}), 413
    row = attachments_mod.save_upload("manual_inventory", ticket_id, f)
    return jsonify({"ok": True, "attachment": row})


@app.route("/api/attachments/<att_id>/download")
def api_attachment_download(att_id):
    row = db.get_attachment(att_id)
    if not row:
        return jsonify({"error": "not found"}), 404
    try:
        path = attachments_mod.disk_path(row)
    except ValueError:
        return jsonify({"error": "invalid path"}), 400
    if not path.exists():
        return jsonify({"error": "file missing"}), 404
    return send_file(
        str(path),
        as_attachment=True,
        download_name=row.get("filename") or "file",
        mimetype=row.get("content_type") or None,
    )


@app.route("/api/attachments/delete", methods=["POST"])
def api_attachment_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    attachments_mod.delete(id_)
    return jsonify({"ok": True})


# --- Expenses + Maaser ---------------------------------------------------

def _parse_iso(s):
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _ensure_subscription_instances():
    """For each active expense subscription, ensure an `expenses` row exists for
    every month from started_at_iso through min(today, ended_at_iso). Idempotent
    via UNIQUE(subscription_id, month_key) — running this on every page load is
    cheap and keeps history accurate even if Kartis was off on the 1st."""
    today = date.today()
    now_iso = datetime.now(timezone.utc).isoformat()
    for sub in db.all_expense_subscriptions():
        start = _parse_iso(sub.get("started_at_iso"))
        if not start:
            continue
        end = _parse_iso(sub.get("ended_at_iso")) or today
        end = min(end, today)
        cursor = date(start.year, start.month, 1)
        last = date(end.year, end.month, 1)
        while cursor <= last:
            month_key = cursor.strftime("%Y-%m")
            row = {
                "id": f"sub-{sub['id']}-{month_key}",
                "date": cursor.strftime("%Y-%m-%d"),
                "date_iso": cursor.strftime("%Y-%m-%d"),
                "vendor": sub.get("name"),
                "category": sub.get("category"),
                "amount": sub.get("amount"),
                "notes": "auto from subscription",
            }
            db.upsert_subscription_instance(sub["id"], month_key, row, now_iso)
            # Advance to first of next month
            nxt = cursor.replace(day=28) + timedelta(days=4)
            cursor = nxt.replace(day=1)


def _maaser_summary():
    """Owed = sum over EVENT-months of max(0, profit − expenses) × 0.10.
    Maaser is calculated against the month each event takes place in,
    not the month its tickets were sold. Given = sum of maaser_payments.
    Outstanding = owed − given."""
    profit_resp = _profit_response()
    owed_by_month = {}
    net_by_month = {}
    for m in profit_resp["months_by_event"]:
        net = (m.get("profit") or 0) - (m.get("expenses") or 0)
        net_by_month[m["month"]] = round(net, 2)
        owed_by_month[m["month"]] = round(max(0, net) * 0.10, 2)
    payments = db.all_maaser()
    given_by_month = {}
    for p in payments:
        key = (p.get("date_iso") or "")[:7]
        if not key:
            continue
        given_by_month[key] = given_by_month.get(key, 0) + (p.get("amount") or 0)
    given_by_month = {k: round(v, 2) for k, v in given_by_month.items()}
    owed_lifetime = round(sum(owed_by_month.values()), 2)
    given_lifetime = round(sum((p.get("amount") or 0) for p in payments), 2)
    this_month_key = date.today().strftime("%Y-%m")
    return {
        "owed_by_month": owed_by_month,
        "given_by_month": given_by_month,
        "net_by_month": net_by_month,
        "summary": {
            "owed_lifetime": owed_lifetime,
            "given_lifetime": given_lifetime,
            "outstanding": round(owed_lifetime - given_lifetime, 2),
            "owed_this_month": owed_by_month.get(this_month_key, 0),
        },
    }


@app.route("/expenses")
def expenses_page():
    return render_template("expenses.html")


@app.route("/maaser")
def maaser_page():
    return render_template("maaser.html")


@app.route("/api/expenses")
def api_expenses():
    _ensure_subscription_instances()
    subs = db.all_expense_subscriptions()
    expenses = db.all_expenses()
    totals_by_month = {}
    for e in expenses:
        mk = (e.get("date_iso") or "")[:7]
        if not mk:
            continue
        totals_by_month[mk] = round(totals_by_month.get(mk, 0) + (e.get("amount") or 0), 2)
    today = date.today()
    this_month = today.strftime("%Y-%m")
    this_year = today.strftime("%Y")
    recurring_monthly = sum((s.get("amount") or 0) for s in subs if not s.get("ended_at_iso"))
    this_month_total = totals_by_month.get(this_month, 0)
    this_year_total = round(sum(v for k, v in totals_by_month.items() if k.startswith(this_year)), 2)
    all_time_total = round(sum(totals_by_month.values()), 2)
    return jsonify({
        "subscriptions": subs,
        "expenses": expenses,
        "totals_by_month": totals_by_month,
        "summary": {
            "recurring_monthly": round(recurring_monthly, 2),
            "this_month": round(this_month_total, 2),
            "this_year": this_year_total,
            "all_time": all_time_total,
        },
        "last_run": _last_run, "last_backup": _last_backup,
    })


@app.route("/api/expenses/sub-add", methods=["POST"])
def api_expenses_sub_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    try:
        amount = float(body.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400
    started = (body.get("started_at_iso") or "").strip()
    if not _parse_iso(started):
        return jsonify({"error": "started_at_iso required (YYYY-MM-DD)"}), 400
    row = {
        "id": "sub-" + uuid.uuid4().hex[:12],
        "name": name,
        "category": (body.get("category") or "").strip(),
        "amount": amount,
        "started_at_iso": started,
        "ended_at_iso": (body.get("ended_at_iso") or "").strip() or None,
        "notes": (body.get("notes") or "").strip(),
    }
    db.insert_expense_subscription(row, datetime.now(timezone.utc).isoformat())
    _ensure_subscription_instances()
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/expenses/sub-edit", methods=["POST"])
def api_expenses_sub_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    fields = {}
    for k in ("name", "category", "amount", "started_at_iso", "ended_at_iso", "notes"):
        if k in body:
            v = body[k]
            if k == "amount":
                try:
                    v = float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    continue
            elif isinstance(v, str):
                v = v.strip() or None
            fields[k] = v
    db.update_expense_subscription(id_, fields)
    _ensure_subscription_instances()
    return jsonify({"ok": True})


@app.route("/api/expenses/sub-delete", methods=["POST"])
def api_expenses_sub_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    # Past auto-generated rows are intentionally left in `expenses` as historical
    # record; only the subscription template is removed.
    db.delete_expense_subscription(id_)
    return jsonify({"ok": True})


@app.route("/api/expenses/add", methods=["POST"])
def api_expenses_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    date_iso = (body.get("date_iso") or "").strip()
    if not _parse_iso(date_iso):
        return jsonify({"error": "date_iso required (YYYY-MM-DD)"}), 400
    try:
        amount = float(body.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400
    row = {
        "id": "exp-" + uuid.uuid4().hex[:12],
        "date": date_iso,
        "date_iso": date_iso,
        "vendor": (body.get("vendor") or "").strip(),
        "category": (body.get("category") or "").strip(),
        "amount": amount,
        "subscription_id": None,
        "month_key": None,
        "notes": (body.get("notes") or "").strip(),
    }
    db.insert_expense(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/expenses/edit", methods=["POST"])
def api_expenses_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    fields = {}
    for k in ("date_iso", "vendor", "category", "amount", "notes"):
        if k in body:
            v = body[k]
            if k == "amount":
                try:
                    v = float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    continue
            elif isinstance(v, str):
                v = v.strip()
            fields[k] = v
    if "date_iso" in fields and fields["date_iso"]:
        fields["date"] = fields["date_iso"]
    db.update_expense(id_, fields)
    return jsonify({"ok": True})


@app.route("/api/expenses/delete", methods=["POST"])
def api_expenses_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_expense(id_)
    return jsonify({"ok": True})


@app.route("/owed")
def owed_page():
    return render_template("owed.html")


@app.route("/api/owed")
def api_owed():
    return jsonify({
        "items": db.all_owed_items(),
        "last_run": _last_run, "last_backup": _last_backup,
    })


@app.route("/api/owed/add", methods=["POST"])
def api_owed_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    date_iso = (body.get("date_iso") or "").strip()
    if not _parse_iso(date_iso):
        return jsonify({"error": "date_iso required (YYYY-MM-DD)"}), 400
    try:
        amount = float(body.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400
    direction = (body.get("direction") or "").strip()
    if direction not in ("charge", "withdrawal"):
        return jsonify({"error": "direction must be 'charge' or 'withdrawal'"}), 400
    row = {
        "id": "owe-" + uuid.uuid4().hex[:12],
        "date_iso": date_iso,
        "amount": amount,
        "description": (body.get("description") or "").strip(),
        "card_account": (body.get("card_account") or "").strip(),
        "direction": direction,
    }
    db.insert_owed_item(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/owed/edit", methods=["POST"])
def api_owed_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    fields = {}
    for k in ("date_iso", "amount", "description", "card_account", "direction"):
        if k in body:
            v = body[k]
            if k == "amount":
                try:
                    v = float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    continue
            elif k == "direction":
                v = (v or "").strip()
                if v not in ("charge", "withdrawal"):
                    continue
            elif isinstance(v, str):
                v = v.strip()
            fields[k] = v
    db.update_owed_item(id_, fields)
    return jsonify({"ok": True})


@app.route("/api/owed/delete", methods=["POST"])
def api_owed_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_owed_item(id_)
    return jsonify({"ok": True})


# --- Unmapped tickets (bought, not yet listed) ---------------------------
# Standalone hand-entry ledger for tickets we own but haven't listed, either
# because we haven't gotten to it or because no secondary market exists yet.
# The `listed` checkbox is a latch, not a delete -- checked rows sink to a
# LISTED block at the bottom of /unmapped so the purchase record survives.

_UNMAPPED_TEXT_FIELDS = (
    "event_name", "event_date", "venue", "section", "row_label", "seats",
    "purchase_source", "account", "email", "phone", "card", "link", "reason",
    "notes",
)


def _unmapped_extra_json(value):
    """Normalize the free-form extra-fields blob to a JSON object string.
    Accepts a dict from the page or a pre-serialized string; anything else
    (or an empty map) stores NULL so the column stays cheap to test."""
    if value in (None, ""):
        return None
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except ValueError:
            return None
    if not isinstance(value, dict):
        return None
    clean = {}
    for k, v in value.items():
        k = str(k).strip()
        if k:
            clean[k] = "" if v is None else str(v).strip()
    return json.dumps(clean, ensure_ascii=False) if clean else None


def _unmapped_payload(body, existing_qty=None):
    """Shared add/edit field coercion. Returns only the keys present in
    `body` so edit can do partial updates.

    Cost can arrive either way: `cost_per_unit` (per ticket) or `cost_total`
    (what the whole batch cost). Only cost_per_unit is stored -- a total is
    divided by qty here, so the division happens once, at full precision,
    against the qty that actually applies. `existing_qty` lets an edit that
    sends a total without a qty still divide by the row's current qty.
    """
    out = {}
    for k in _UNMAPPED_TEXT_FIELDS:
        if k in body:
            out[k] = (body.get(k) or "").strip()
    if "qty" in body:
        try:
            out["qty"] = int(body.get("qty") or 0) or None
        except (TypeError, ValueError):
            out["qty"] = None
    if "cost_per_unit" in body:
        v = body.get("cost_per_unit")
        try:
            out["cost_per_unit"] = float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            out["cost_per_unit"] = None
    if body.get("cost_total") not in (None, "") and not out.get("cost_per_unit"):
        # Batch cost split evenly per ticket, the same shape as the
        # event_group_costs split on /sales.
        try:
            batch = float(body["cost_total"])
        except (TypeError, ValueError):
            batch = None
        qty = out["qty"] if "qty" in out else existing_qty
        if batch is not None and qty:
            out["cost_per_unit"] = batch / qty
    if "extra" in body:
        out["extra_json"] = _unmapped_extra_json(body.get("extra"))
    # Derive the sortable ISO date from the free-text one when the page
    # didn't send it, same as /api/inventory/manual-add.
    if "event_date" in body or "event_date_iso" in body:
        iso = (body.get("event_date_iso") or "").strip()
        if not iso and out.get("event_date"):
            iso = _date_only(_parse_event_date(out["event_date"])) or ""
        # NULL, not "", so the "undated last" ordering actually catches it.
        out["event_date_iso"] = iso or None
    return out


@app.route("/unmapped")
def unmapped_page():
    return render_template("unmapped.html")


@app.route("/api/unmapped")
def api_unmapped():
    items = db.all_unmapped_tickets()
    atts = db.list_attachments_for_owners("unmapped", [i["id"] for i in items])
    for it in items:
        it["attachments"] = atts.get(it["id"], [])
        try:
            it["extra"] = json.loads(it.get("extra_json") or "{}")
        except ValueError:
            it["extra"] = {}
    return jsonify({"items": items, "last_run": _last_run, "last_backup": _last_backup})


@app.route("/api/unmapped/add", methods=["POST"])
def api_unmapped_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    fields = _unmapped_payload(body)
    if not fields.get("event_name"):
        return jsonify({"error": "event_name required"}), 400
    row = {k: fields.get(k) for k in db.UNMAPPED_FIELDS}
    row["id"] = "unm-" + uuid.uuid4().hex[:12]
    db.insert_unmapped_ticket(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/unmapped/edit", methods=["POST"])
def api_unmapped_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    existing = db.get_unmapped_ticket(id_)
    if not existing:
        return jsonify({"error": "unknown id"}), 404
    fields = _unmapped_payload(body, existing_qty=existing.get("qty"))
    if "event_name" in fields and not fields["event_name"]:
        return jsonify({"error": "event_name cannot be blank"}), 400
    db.update_unmapped_ticket(id_, fields, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True})


@app.route("/api/unmapped/listed", methods=["POST"])
def api_unmapped_listed():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.set_unmapped_listed(id_, bool(body.get("listed")), datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True})


@app.route("/api/unmapped/delete", methods=["POST"])
def api_unmapped_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_unmapped_ticket(id_)
    # Same convention as /api/inventory/manual-delete: drop the DB rows but
    # leave the PDFs on disk so OneDrive still has them.
    db.delete_attachments_for_owner("unmapped", id_)
    return jsonify({"ok": True})


@app.route("/api/unmapped/attach", methods=["POST"])
def api_unmapped_attach():
    from flask import request
    ticket_id = (request.form.get("ticket_id") or "").strip()
    if not ticket_id:
        return jsonify({"error": "ticket_id required"}), 400
    if ticket_id not in {t["id"] for t in db.all_unmapped_tickets()}:
        return jsonify({"error": "unknown ticket_id"}), 404
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "file required"}), 400
    f.stream.seek(0, 2)
    size = f.stream.tell()
    f.stream.seek(0)
    if size > attachments_mod.MAX_BYTES:
        return jsonify({"error": f"file too large (>{attachments_mod.MAX_BYTES} bytes)"}), 413
    row = attachments_mod.save_upload("unmapped", ticket_id, f)
    return jsonify({"ok": True, "attachment": row})


# --- Cashback ledger -----------------------------------------------------
# Manual log of credit-card cashback rewards. Purely informational — does
# not feed into Maaser (user opted out: cashback is treated as a rebate,
# not income). Pattern mirrors /owed exactly.
_CASHBACK_EDITABLE = {"date_iso", "amount", "card_name"}


@app.route("/cashback")
def cashback_page():
    return render_template("cashback.html")


@app.route("/api/cashback")
def api_cashback():
    return jsonify({
        "items": db.all_cashback_entries(),
        "known_cards": db.distinct_cashback_cards(),
    })


@app.route("/api/cashback/add", methods=["POST"])
def api_cashback_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    date_iso = (body.get("date_iso") or "").strip()
    if not _parse_iso(date_iso):
        return jsonify({"error": "date_iso required (YYYY-MM-DD)"}), 400
    try:
        amount = float(body.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400
    card_name = (body.get("card_name") or "").strip()
    if not card_name:
        return jsonify({"error": "card_name required"}), 400
    row = {
        "id": "cb-" + uuid.uuid4().hex[:12],
        "date_iso": date_iso,
        "amount": amount,
        "card_name": card_name,
    }
    db.insert_cashback_entry(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/cashback/edit", methods=["POST"])
def api_cashback_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    fields = {}
    for k in _CASHBACK_EDITABLE:
        if k in body:
            v = body[k]
            if k == "amount":
                try:
                    v = float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    continue
                if v is None or v <= 0:
                    continue
            elif isinstance(v, str):
                v = v.strip()
                if k in ("date_iso", "card_name") and not v:
                    continue
            fields[k] = v
    db.update_cashback_entry(id_, fields)
    return jsonify({"ok": True})


@app.route("/api/cashback/delete", methods=["POST"])
def api_cashback_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_cashback_entry(id_)
    return jsonify({"ok": True})


# ---------------- Kupat credits ---------------------------------------------

_CREDIT_EDITABLE = {"issued_date", "ils_amount", "original_usd_cost", "note"}


@app.route("/credits")
def credits_page():
    return render_template("credits.html")


@app.route("/api/credits")
def api_credits():
    return jsonify(kupat_credits.build_credits_view())


@app.route("/api/credits/add", methods=["POST"])
def api_credits_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    issued_date = (body.get("issued_date") or "").strip()
    if not _parse_iso(issued_date):
        return jsonify({"error": "issued_date required (YYYY-MM-DD)"}), 400
    try:
        ils_amount = float(body.get("ils_amount") or 0)
    except (TypeError, ValueError):
        ils_amount = 0
    if ils_amount <= 0:
        return jsonify({"error": "ils_amount must be > 0"}), 400
    original_usd_cost = body.get("original_usd_cost")
    if original_usd_cost in (None, ""):
        original_usd_cost = None
    else:
        try:
            original_usd_cost = float(original_usd_cost)
        except (TypeError, ValueError):
            return jsonify({"error": "original_usd_cost must be numeric"}), 400
        if original_usd_cost < 0:
            return jsonify({"error": "original_usd_cost must be >= 0"}), 400
    note = (body.get("note") or "").strip()
    row = {
        "id": "kc-" + uuid.uuid4().hex[:12],
        "issued_date": issued_date,
        "ils_amount": ils_amount,
        "original_usd_cost": original_usd_cost,
        "note": note,
    }
    db.insert_kupat_credit(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/credits/edit", methods=["POST"])
def api_credits_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    fields = {}
    for k in _CREDIT_EDITABLE:
        if k not in body:
            continue
        v = body[k]
        if k == "ils_amount":
            try:
                v = float(v)
            except (TypeError, ValueError):
                continue
            if v <= 0:
                continue
        elif k == "original_usd_cost":
            if v in (None, ""):
                v = None
            else:
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    continue
                if v < 0:
                    continue
        elif isinstance(v, str):
            v = v.strip()
            if k == "issued_date" and not v:
                continue
        fields[k] = v
    db.update_kupat_credit(id_, fields)
    return jsonify({"ok": True})


@app.route("/api/credits/delete", methods=["POST"])
def api_credits_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_kupat_credit(id_)
    return jsonify({"ok": True})


@app.route("/api/credits/spend/add", methods=["POST"])
def api_credits_spend_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    credit_id = (body.get("credit_id") or "").strip()
    if not credit_id:
        return jsonify({"error": "credit_id required"}), 400
    spend_date = (body.get("spend_date") or "").strip()
    if not _parse_iso(spend_date):
        return jsonify({"error": "spend_date required (YYYY-MM-DD)"}), 400
    try:
        ils_amount = float(body.get("ils_amount") or 0)
    except (TypeError, ValueError):
        ils_amount = 0
    if ils_amount <= 0:
        return jsonify({"error": "ils_amount must be > 0"}), 400

    credits_by_id = {c["id"]: c for c in db.all_kupat_credits()}
    credit = credits_by_id.get(credit_id)
    if not credit:
        return jsonify({"error": "credit not found"}), 404
    spent_so_far = sum(
        float(s["ils_amount"] or 0)
        for s in db.all_kupat_spends()
        if s["credit_id"] == credit_id
    )
    if spent_so_far + ils_amount > float(credit["ils_amount"] or 0) + 0.001:
        remaining = float(credit["ils_amount"] or 0) - spent_so_far
        return jsonify({
            "error": f"overspend: only {remaining:.2f} ILS remaining on this credit"
        }), 400

    rate, _, _ = kupat_credits.current_fx_rate()
    note = (body.get("note") or "").strip()
    row = {
        "id": "ks-" + uuid.uuid4().hex[:12],
        "credit_id": credit_id,
        "spend_date": spend_date,
        "ils_amount": ils_amount,
        "fx_rate": rate,
        "note": note,
    }
    db.insert_kupat_spend(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"], "fx_rate": rate})


@app.route("/api/credits/spend/delete", methods=["POST"])
def api_credits_spend_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_kupat_spend(id_)
    return jsonify({"ok": True})


@app.route("/api/maaser")
def api_maaser():
    summary = _maaser_summary()
    return jsonify({
        "payments": db.all_maaser(),
        **summary,
        "last_run": _last_run, "last_backup": _last_backup,
    })


@app.route("/api/maaser/add", methods=["POST"])
def api_maaser_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    date_iso = (body.get("date_iso") or "").strip()
    if not _parse_iso(date_iso):
        return jsonify({"error": "date_iso required (YYYY-MM-DD)"}), 400
    try:
        amount = float(body.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400
    row = {
        "id": "maaser-" + uuid.uuid4().hex[:12],
        "date": date_iso,
        "date_iso": date_iso,
        "recipient": (body.get("recipient") or "").strip(),
        "amount": amount,
        "notes": (body.get("notes") or "").strip(),
        # Stored as 0/1 to match the SQLite NOT NULL DEFAULT 0 column.
        "tax_deductible": 1 if body.get("tax_deductible") else 0,
    }
    db.insert_maaser(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/maaser/edit", methods=["POST"])
def api_maaser_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    fields = {}
    for k in ("date_iso", "recipient", "amount", "notes", "tax_deductible"):
        if k in body:
            v = body[k]
            if k == "amount":
                try:
                    v = float(v) if v not in (None, "") else None
                except (TypeError, ValueError):
                    continue
            elif k == "tax_deductible":
                # Accept bool / 0 / 1 / "0" / "1" / "true" / "false" — coerce to 0/1.
                if isinstance(v, str):
                    v = v.strip().lower()
                    v = 1 if v in ("1", "true", "yes", "on") else 0
                else:
                    v = 1 if v else 0
            elif isinstance(v, str):
                v = v.strip()
            fields[k] = v
    if "date_iso" in fields and fields["date_iso"]:
        fields["date"] = fields["date_iso"]
    db.update_maaser(id_, fields)
    return jsonify({"ok": True})


@app.route("/api/maaser/delete", methods=["POST"])
def api_maaser_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.delete_maaser(id_)
    return jsonify({"ok": True})


# ----- To-Do list ----------------------------------------------------------

@app.route("/todos")
def todos_page():
    return render_template("todos.html")


@app.route("/api/todos")
def api_todos():
    from flask import request
    status = (request.args.get("status") or "all").strip().lower()
    rows, stats = todos_mod.list_todos(status=status)
    return jsonify({
        "todos": rows,
        "stats": stats,
        "last_remind": _last_todo_remind,
        "linkable_sources": list(todos_mod.VALID_SOURCES),
        "urgencies": list(todos_mod.VALID_URGENCIES),
        "recurrences": list(todos_mod.VALID_RECURRENCES),
    })


@app.route("/api/todos/add", methods=["POST"])
def api_todos_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    new_id, err = todos_mod.create(body)
    if err:
        return jsonify({"error": err}), 400
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/todos/edit", methods=["POST"])
def api_todos_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    ok, err = todos_mod.edit(id_, body)
    if not ok:
        return jsonify({"error": err}), (404 if err == "not found" else 400)
    return jsonify({"ok": True})


@app.route("/api/todos/delete", methods=["POST"])
def api_todos_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.todo_delete(id_)
    return jsonify({"ok": True})


@app.route("/api/todos/toggle", methods=["POST"])
def api_todos_toggle():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    ok, payload = todos_mod.toggle(id_)
    if not ok:
        return jsonify({"error": payload}), 404
    return jsonify({"ok": True, "result": payload})


@app.route("/api/todos/subtask/add", methods=["POST"])
def api_todos_subtask_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    todo_id = (body.get("todo_id") or "").strip()
    title = (body.get("title") or "").strip()
    if not todo_id or not db.todo_get(todo_id):
        return jsonify({"error": "todo_id required"}), 400
    if not title:
        return jsonify({"error": "title required"}), 400
    existing = db.subtask_list(todo_id)
    row = {
        "id": todos_mod.new_subtask_id(),
        "todo_id": todo_id,
        "title": title[:300],
        "done": 0,
        "order_idx": len(existing),
    }
    db.subtask_insert(row, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "id": row["id"]})


@app.route("/api/todos/subtask/toggle", methods=["POST"])
def api_todos_subtask_toggle():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.subtask_toggle(id_)
    return jsonify({"ok": True})


@app.route("/api/todos/subtask/delete", methods=["POST"])
def api_todos_subtask_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    id_ = (body.get("id") or "").strip()
    if not id_:
        return jsonify({"error": "id required"}), 400
    db.subtask_delete(id_)
    return jsonify({"ok": True})


@app.route("/api/todos/suggestions")
def api_todos_suggestions():
    return jsonify({"suggestions": todos_mod.compute_suggestions()})


@app.route("/api/todos/suggestions/accept", methods=["POST"])
def api_todos_suggestions_accept():
    from flask import request
    body = request.get_json(silent=True) or {}
    key = (body.get("key") or "").strip()
    if not key:
        return jsonify({"error": "key required"}), 400
    new_id, err = todos_mod.accept_suggestion(key)
    if err:
        return jsonify({"error": err}), 400
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/todos/suggestions/dismiss", methods=["POST"])
def api_todos_suggestions_dismiss():
    from flask import request
    body = request.get_json(silent=True) or {}
    key = (body.get("key") or "").strip()
    if not key:
        return jsonify({"error": "key required"}), 400
    until = todos_mod.dismiss_suggestion(key, body.get("days") or 7)
    return jsonify({"ok": True, "until": until})


@app.route("/api/todos/remind-now", methods=["POST"])
def api_todos_remind_now():
    """Force-run the digest. Useful for manual testing without waiting for
    the 08:00 cron."""
    threading.Thread(target=run_todo_remind, daemon=True).start()
    return jsonify({"started": True})


@app.route("/api/sales/manual-record", methods=["POST"])
def api_sales_manual_record():
    from flask import request
    import uuid
    body = request.get_json(silent=True) or {}
    inv_source = (body.get("inv_source") or "").strip()
    inv_source_id = (body.get("inv_source_id") or "").strip()
    if not inv_source or not inv_source_id:
        return jsonify({"error": "inv_source and inv_source_id required"}), 400
    rows, _ = _build_unified_inventory()
    inv = next((r for r in rows if r.get("source") == inv_source and str(r.get("source_id")) == inv_source_id), None)
    if not inv:
        return jsonify({"error": "inventory row not found (may already be sold or hidden)"}), 404
    try:
        qty = int(body.get("qty") or 0)
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0 or qty > (inv.get("qty_unsold") or 0):
        return jsonify({"error": "qty must be 1..remaining"}), 400
    is_loss = bool(body.get("is_loss"))
    sale_price = 0.0 if is_loss else float(body.get("sale_price") or 0)
    sale_date = (body.get("sale_date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
    sale_iso = sale_date[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", sale_date) else sale_date
    platform = (body.get("platform") or ("loss" if is_loss else "manual")).strip()
    note = (body.get("note") or "").strip()
    # Optional override: which seat numbers were sold. Falls back to the
    # inventory row's full seat list if the caller leaves it blank.
    seats_sold = (body.get("seats") or "").strip() if "seats" in body else None
    cost_per = inv.get("cost_per_unit")
    if cost_per is None and (inv.get("qty_unsold") or 0) > 0:
        cost_per = (inv.get("cost") or 0) / (inv.get("qty_unsold") or 1)
    cost = round((cost_per or 0) * qty, 2)
    sale_id = "manual-" + uuid.uuid4().hex[:12]
    now_iso = datetime.now(timezone.utc).isoformat()
    db.insert_manual_sale({
        "id": sale_id,
        "inv_source": inv_source,
        "inv_source_id": inv_source_id,
        "event_name": inv.get("event_name") or "",
        "event_date": inv.get("event_date") or "",
        "event_date_iso": inv.get("event_date_iso") or "",
        "venue": inv.get("venue") or "",
        "section": inv.get("section") or "",
        "row_label": inv.get("row") or "",
        "seats": seats_sold if seats_sold is not None else (inv.get("seats") or ""),
        "qty": qty,
        "sale_price": sale_price,
        "cost": cost,
        "sale_date": sale_date,
        "sale_date_iso": sale_iso,
        "platform": platform,
        "is_loss": 1 if is_loss else 0,
        "note": note,
    }, now_iso)
    db.record_match(
        sale_source="manual", sale_id=sale_id,
        inv_source=inv_source, inv_source_id=inv_source_id,
        qty=qty, reason="manual" + ("/loss" if is_loss else ""), now_iso=now_iso,
    )
    return jsonify({"ok": True, "sale_id": sale_id})


@app.route("/api/sales/manual-delete", methods=["POST"])
def api_sales_manual_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    sale_id = (body.get("sale_id") or "").strip()
    if not sale_id:
        return jsonify({"error": "sale_id required"}), 400
    db.delete_manual_sale(sale_id)
    db.delete_match("manual", sale_id)
    return jsonify({"ok": True})


@app.route("/api/match-now", methods=["POST"])
def api_match_now():
    # Pending-match first so manual rows already linked to a Lysted/Viagogo
    # listing get flagged before the auto-matcher considers them.
    pending_listed = matcher.run_pending_match()
    matched, skipped = matcher.run_match_pass()
    return jsonify({"matched": matched, "unmatched": skipped, "pending_listed": pending_listed})


@app.route("/api/sales/unmatched")
def api_sales_unmatched():
    """External (Lysted/Viagogo/CrowdVolt) sales not yet paired with any
    inventory row. Powers the manual pair-with-sale fallback UI for cases
    where the auto-matcher's section/row scoring misses."""
    from flask import request
    q = (request.args.get("event") or "").strip()
    qn = _norm_event_name(q) if q else ""
    matched_keys = {(m["sale_source"], m["sale_id"]) for m in db.all_matches()}
    blocked = db.all_blocklist_keys()
    rows = []
    for s in matcher._collect_external_sales():
        key = (s["source"], s["id"])
        if key in matched_keys or key in blocked:
            continue
        if qn and qn not in _norm_event_name(s.get("event_name") or ""):
            continue
        rows.append({
            "sale_source": s["source"],
            "sale_id": s["id"],
            "event_name": s.get("event_name") or "",
            "event_date_iso": s.get("event_date_iso") or "",
            "section": s.get("section") or "",
            "row": s.get("row") or "",
            "qty": s.get("qty") or 0,
            "sale_date": s.get("sale_date") or "",
        })
    rows.sort(key=lambda r: r.get("sale_date") or "", reverse=True)
    return jsonify({"rows": rows})


@app.route("/api/sales/pair", methods=["POST"])
def api_sales_pair():
    """Manually pair an external sale with an inventory row when auto-match
    couldn't (e.g. section label mismatch). Writes inventory_matches and
    hides the inventory row when fully consumed."""
    from flask import request
    body = request.get_json(silent=True) or {}
    sale_source = (body.get("sale_source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    inv_source = (body.get("inv_source") or "").strip()
    inv_source_id = (body.get("inv_source_id") or "").strip()
    if not (sale_source and sale_id and inv_source and inv_source_id):
        return jsonify({"error": "sale_source, sale_id, inv_source, inv_source_id required"}), 400
    try:
        qty = int(body.get("qty") or 0)
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0:
        return jsonify({"error": "qty must be > 0"}), 400
    # Reject duplicate pair (inventory_matches PRIMARY KEY is sale-side).
    existing = {(m["sale_source"], m["sale_id"]) for m in db.all_matches()}
    if (sale_source, sale_id) in existing:
        return jsonify({"error": "sale already paired"}), 409
    rows, _ = _build_unified_inventory()
    inv = next((r for r in rows if r.get("source") == inv_source and str(r.get("source_id")) == inv_source_id), None)
    if not inv:
        return jsonify({"error": "inventory row not found"}), 404
    if qty > (inv.get("qty_unsold") or 0):
        return jsonify({"error": "qty exceeds remaining"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    db.record_match(
        sale_source=sale_source, sale_id=sale_id,
        inv_source=inv_source, inv_source_id=inv_source_id,
        qty=qty, reason="manual-pair", now_iso=now_iso,
    )
    if qty >= (inv.get("qty_unsold") or 0):
        db.hide_inventory(inv_source, inv_source_id, now_iso)
    return jsonify({"ok": True})


@app.route("/api/sales/unpair", methods=["POST"])
def api_sales_unpair():
    """Reverse a manual pair: drop the inventory_matches row and unhide the
    inventory row so it returns to the active list."""
    from flask import request
    body = request.get_json(silent=True) or {}
    sale_source = (body.get("sale_source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    if not (sale_source and sale_id):
        return jsonify({"error": "sale_source and sale_id required"}), 400
    target = next(
        (m for m in db.all_matches()
         if m.get("sale_source") == sale_source and m.get("sale_id") == sale_id),
        None,
    )
    if not target:
        return jsonify({"error": "match not found"}), 404
    db.delete_match(sale_source, sale_id)
    inv_src = target.get("inv_source")
    inv_id = target.get("inv_source_id")
    if inv_src and inv_id is not None:
        db.unhide_inventory(inv_src, str(inv_id))
    return jsonify({"ok": True})


_INV_EDITABLE = {"section", "row", "seats", "qty_unsold", "cost", "cost_per_unit", "list_price", "delivery_type", "status", "event_name", "venue", "event_date", "event_date_iso"}
_SALE_EDITABLE = {"event_name", "venue", "section", "row", "qty", "sale_price", "cost", "platform", "sale_date", "sale_date_iso", "event_date", "event_date_iso"}

# Status values that mean "this batch didn't sell" — triggers archive into
# inventory_unsold instead of a normal status override. Editing the Status
# field on the inventory page to any of these (case- and whitespace-
# insensitive) hides the row and surfaces it on the sales page under the
# "Didn't Sell" section.
_UNSOLD_RE = re.compile(r"^\s*(not[\s_-]?sold|didn'?t\s*sell|did\s*not\s*sell|unsold)\s*$", re.IGNORECASE)


@app.route("/api/inventory/edit", methods=["POST"])
def api_inventory_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    source_id = (body.get("source_id") or "").strip()
    field = (body.get("field") or "").strip()
    value = body.get("value")
    if not source or not source_id or field not in _INV_EDITABLE:
        return jsonify({"error": "invalid source/source_id/field"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()

    # "Didn't sell" trigger: archive into inventory_unsold by content
    # fingerprint so the tombstone survives source-id rotation on resync.
    if field == "status" and value and _UNSOLD_RE.match(str(value)):
        rows, _skipped = _build_unified_inventory()
        inv = next(
            (r for r in rows
             if r.get("source") == source and str(r.get("source_id")) == source_id),
            None,
        )
        if not inv:
            return jsonify({"error": "row not found in current inventory"}), 404
        fp = db.unsold_fingerprint(
            source, inv.get("event_name"), inv.get("event_date_iso"),
            inv.get("section"), inv.get("row"), inv.get("seats"), inv.get("qty_unsold"),
        )
        snap = {
            "fingerprint": fp,
            "source": source,
            "source_id": source_id,
            "event_name": inv.get("event_name"),
            "event_date": inv.get("event_date"),
            "event_date_iso": inv.get("event_date_iso"),
            "venue": inv.get("venue"),
            "section": inv.get("section"),
            "row_label": inv.get("row"),
            "seats": inv.get("seats"),
            "qty": inv.get("qty_unsold"),
            "cost": inv.get("cost"),
            "cost_per_unit": inv.get("cost_per_unit"),
            "list_price": inv.get("list_price"),
            "delivery_type": inv.get("delivery_type"),
            "note": None,
        }
        db.mark_inventory_unsold(snap, now_iso)
        # Drop any active status override on this row — the unsold archive
        # is now the source of truth for "this row is gone".
        db.set_inv_override(source, source_id, "status", None, now_iso)
        return jsonify({"ok": True, "archived": True, "fingerprint": fp})

    db.set_inv_override(source, source_id, field, value, now_iso)
    return jsonify({"ok": True})


@app.route("/api/sales/edit", methods=["POST"])
def api_sales_edit():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    sale_id = (body.get("sale_id") or "").strip()
    field = (body.get("field") or "").strip()
    value = body.get("value")
    if not source or not sale_id or field not in _SALE_EDITABLE:
        return jsonify({"error": "invalid source/sale_id/field"}), 400
    db.set_sale_override(source, sale_id, field, value, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True})


@app.route("/api/inventory/hide", methods=["POST"])
def api_inventory_hide():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    source_id = (body.get("source_id") or "").strip()
    if not source or not source_id:
        return jsonify({"error": "source and source_id required"}), 400
    db.hide_inventory(source, source_id, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "source": source, "source_id": source_id})


@app.route("/api/inventory/unhide", methods=["POST"])
def api_inventory_unhide():
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    source_id = (body.get("source_id") or "").strip()
    if not source or not source_id:
        return jsonify({"error": "source and source_id required"}), 400
    # If an auto-match caused this hide, also blocklist the sale so the
    # next matcher pass doesn't immediately re-pair them.
    match = db.find_match_for_inv(source, source_id)
    if match:
        db.add_blocklist(match["sale_source"], match["sale_id"], datetime.now(timezone.utc).isoformat())
        db.delete_match(match["sale_source"], match["sale_id"])
    db.unhide_inventory(source, source_id)
    return jsonify({"ok": True, "source": source, "source_id": source_id, "from_match": bool(match)})


@app.route("/api/inventory/unmark-unsold", methods=["POST"])
def api_inventory_unmark_unsold():
    """Restore a row from the Didn't Sell archive back to active inventory.

    Removes the content-fingerprint tombstone from inventory_unsold. The row
    will reappear on the next /api/inventory-all load if its source data is
    still present (Lysted/Viagogo/JeruJam still expose the underlying ticket).
    """
    from flask import request
    body = request.get_json(silent=True) or {}
    fp = (body.get("fingerprint") or "").strip()
    if not fp:
        return jsonify({"error": "fingerprint required"}), 400
    db.unmark_inventory_unsold(fp)
    return jsonify({"ok": True, "fingerprint": fp})


@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    threading.Thread(target=run_scraper, daemon=True).start()
    return jsonify({"started": True})


@app.route("/api/backup", methods=["POST"])
def api_backup():
    threading.Thread(target=run_backup, daemon=True).start()
    return jsonify({"started": True})


@app.route("/jerujam")
def jerujam():
    return render_template("jerujam.html")


@app.route("/api/jerujam")
def api_jerujam():
    tickets = db.all_jerujam_tickets()
    sales = db.all_jerujam_sales()
    expenses = db.all_jerujam_expenses()

    def is_sold(t):
        return (t.get("status") or "").strip().lower() == "sold"

    sold = [t for t in tickets if is_sold(t)]
    held = [t for t in tickets if not is_sold(t)]
    total_purchase_cost = sum((t.get("total_purchase_cost") or 0) for t in tickets)
    total_sales = sum((s.get("sale_price") or 0) for s in sales)
    total_seller_fees = sum((t.get("seller_fees") or 0) for t in tickets)
    sold_purchase_cost = sum((t.get("total_purchase_cost") or 0) for t in sold)
    held_purchase_cost = sum((t.get("total_purchase_cost") or 0) for t in held)

    by_status = {}
    for t in tickets:
        s = (t.get("status") or "(none)").strip() or "(none)"
        by_status[s] = by_status.get(s, 0) + 1

    totals = {
        "tickets": len(tickets),
        "sales": len(sales),
        "expenses": len(expenses),
        "qty_total": sum((t.get("quantity") or 0) for t in tickets),
        "qty_sold": sum((s.get("quantity") or 0) for s in sales),
        "total_purchase_cost": round(total_purchase_cost, 2),
        "sold_purchase_cost": round(sold_purchase_cost, 2),
        "held_purchase_cost": round(held_purchase_cost, 2),
        "total_sales": round(total_sales, 2),
        "total_seller_fees": round(total_seller_fees, 2),
        "net_profit": round(total_sales - sold_purchase_cost - total_seller_fees, 2),
        "by_status": by_status,
    }
    return jsonify({
        "tickets": tickets,
        "sales": sales,
        "expenses": expenses,
        "totals": totals,
        "last_import": _last_jerujam,
    })


@app.route("/api/jerujam/import", methods=["POST"])
def api_jerujam_import():
    threading.Thread(target=run_jerujam_import, daemon=True).start()
    return jsonify({"started": True})


@app.route("/export.xlsx")
def export_xlsx():
    wb = Workbook()
    ws_l = wb.active
    ws_l.title = "Lysted"
    ws_l.append([
        "Event", "Date", "Time", "Venue",
        "Listings", "Tickets", "Total Cost", "Total List", "P/L",
    ])
    for r in _enrich(db.all_inventory()):
        ws_l.append([
            r.get("event_name"), r.get("event_date"), r.get("event_time"),
            r.get("venue"),
            r.get("listings_count"), r.get("tickets_count"),
            r.get("total_cost"), r.get("total_list"), r.get("profit_loss"),
        ])

    ws_p = wb.create_sheet("Lysted Tickets")
    ws_p.append([
        "Order", "Order Date", "Event", "Event Date", "Venue",
        "Section", "Row", "Qty", "Seats",
        "Delivery", "Account", "Transaction ID",
        "Total Cost", "Cost/Unit", "Status",
    ])
    for r in db.all_lysted_purchases():
        ws_p.append([
            r.get("order_id"), r.get("order_date"),
            r.get("event_name"), r.get("event_date"), r.get("venue"),
            r.get("section"), r.get("row_label"), r.get("qty"), r.get("seats"),
            r.get("delivery_type"), r.get("account_email"), r.get("transaction_id"),
            r.get("total_cost"), r.get("cost_per_unit"), r.get("status"),
        ])

    ws_v = wb.create_sheet("Viagogo")
    ws_v.append([
        "Event", "Date", "Venue", "Section", "Ticket Type",
        "Visibility", "Face Value", "Available", "Cost (Avail x Face)",
        "Price", "Proceeds", "Sold",
    ])
    for r in _enrich_viagogo(db.all_viagogo()):
        ws_v.append([
            r.get("event_name"), r.get("event_date"), r.get("venue"),
            r.get("section"), r.get("ticket_type"),
            r.get("visibility"),
            r.get("face_value"), r.get("available"), r.get("cost"),
            r.get("price"), r.get("proceeds"), r.get("sold"),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"kartis-{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _diff_seats_set(prev_keys, curr_seats, key_fn):
    """Source-agnostic set diff. `key_fn` is the source's seat_key callable
    (or `ticketmaster.event_seat_key` for event-level watchers, where the
    perf code is part of the dedup key)."""
    prev = set(prev_keys)
    by_key = {}
    for s in curr_seats:
        by_key.setdefault(key_fn(s), s)
    curr = set(by_key)
    added = [by_key[k] for k in curr - prev]
    removed = list(prev - curr)
    return added, removed


def _check_one_watcher(w, now_iso):
    """One tick for one watcher. Returns (added_count, error_str_or_None).

    The very first tick on a fresh watcher (no recorded state and no prior
    check timestamp) is treated as a baseline: we capture whatever's
    currently available and DO NOT notify, otherwise the user gets a
    spam ping for seats that were already there at watch-create time.

    Event-level watchers (perf_code='ALL', ticketmaster only) aggregate seats
    across every active performance. The same baseline rule applies, plus
    a "tickets just opened" headline fires when the aggregate flips from
    zero to non-zero seats.

    Notification gating: master_muted (db setting), the watcher's own
    `muted` flag, and `notify_channels` all apply here. When notifications
    are suppressed the drop is still recorded in `tm_drops` so the
    dashboard history shows what was missed.
    """
    import json as _json
    wid = w["id"]
    label = w.get("label") or f"{w['event_code']}/{w['perf_code']}"
    src_name = w.get("source") or "ticketmaster"
    src = WATCHER_SOURCES.get(src_name, ticketmaster)

    event_level = src is ticketmaster and ticketmaster.is_event_level(w)
    if event_level:
        perf_url = ticketmaster.event_url(w["event_code"])
        key_fn = ticketmaster.event_seat_key
    else:
        perf_url = src.perf_url(w["event_code"], w["perf_code"])
        key_fn = src.seat_key

    try:
        if event_level:
            seats, perf_errors = ticketmaster.fetch_event_seats(w["event_code"])
            if perf_errors:
                print(f"[{now_iso}] {label}: per-perf errors {perf_errors}", flush=True)
        else:
            seats = src.fetch_selectable_seats(w["event_code"], w["perf_code"])
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        db.tm_update_watcher(wid, {
            "last_check_at": now_iso, "last_check_error": err,
        })
        return 0, err

    prev_keys = db.tm_get_seat_keys(wid)
    is_baseline = not prev_keys and not w.get("last_check_at")
    was_empty = (w.get("last_seat_count") or 0) == 0
    added, removed = _diff_seats_set(prev_keys, seats, key_fn)
    db.tm_replace_seat_state(wid, seats)
    # Unconfirmed (sold-out perf) seats: only perfs already under tracking may
    # ping — the tick that STARTS tracking a perf sees its whole standing
    # unsold-seat set as "added", which is inventory, not a drop. The per-perf
    # `_tracking` sentinel in prev state marks tracking as established.
    # (Keep in sync with watcher_only.py.)
    if any(s.get("unconfirmed") for s in added):
        _tracked = {k.split("|", 2)[1] for k in prev_keys if k.startswith("U|")}
        added = [s for s in added
                 if not s.get("_tracking")
                 and (not s.get("unconfirmed") or (s.get("_perf") or "") in _tracked)]
    db.tm_update_watcher(wid, {
        "last_check_at": now_iso,
        "last_check_error": None,
        "last_seat_count": len(seats),
    })

    if is_baseline:
        return 0, None

    if added:
        # Prefer a real seat for the label probe — status pseudo-seats carry a
        # synthetic block label that would force a labels refetch every flip.
        probe_block = next(
            (s.get("block") or s.get("b") for s in added
             if (s.get("block") or s.get("b")) and not (s.get("festival") or s.get("ga"))),
            None,
        )
        if event_level:
            probe_perf = next((s.get("_perf") for s in added if s.get("_perf")), None)
            try:
                labels = src.get_labels(w["event_code"], probe_perf, lang="iw", missing_block=probe_block) if probe_perf else None
            except Exception:
                labels = None
        else:
            try:
                labels = src.get_labels(w["event_code"], w["perf_code"], lang="iw", missing_block=probe_block)
            except Exception:
                labels = None

        # Apply the watcher's filters (min consecutive seats, exclude sections,
        # price range). Diff against `seats` (current full availability) so
        # adjacency reflects the live snapshot, not historical state.
        matched = watcher_filters.apply(added, seats, w.get("filters"), labels=labels)
        # Mixed venues (a seat map PLUS a GA lawn/pit — kupat Live Park, TM
        # GA allocations) also carry a GA status pseudo-seat whose
        # soldout↔available flips are noise next to real per-seat drops.
        # A watcher whose snapshot has actual seats only pings on those;
        # GA-ONLY events (no seat map at all) keep their status pings.
        # (Keep in sync with watcher_only.py.)
        if any(not (s.get("festival") or s.get("ga")) for s in seats):
            matched = [s for s in matched if not s.get("ga")]

        # Per-seat cool-down: drop physical seats that already pinged inside
        # the window, so a flapping VIP seat doesn't re-ping all day. Status
        # pseudo-seats (GA/festival flips) are never cooled. (Sync w/ watcher_only.)
        if matched and SEAT_COOLDOWN_SECONDS > 0:
            phys_keys = [key_fn(s) for s in matched if not (s.get("festival") or s.get("ga"))]
            cooled = db.tm_seat_cooldown_active(wid, phys_keys, SEAT_COOLDOWN_SECONDS, now_iso)
            if cooled:
                matched = [s for s in matched
                           if (s.get("festival") or s.get("ga")) or key_fn(s) not in cooled]

        master_muted = db.setting_get_bool("master_muted", default=False)
        watcher_muted = bool(w.get("muted"))
        channels_csv = (w.get("notify_channels") or "discord,email").strip().lower()
        enabled = {c for c in (s.strip() for s in channels_csv.split(",")) if c}
        if master_muted or watcher_muted:
            enabled = set()

        # Status flip — fires on every sold-out → available transition (since
        # last_seat_count returns to 0 between flips), satisfying the
        # "keep notifying" requirement for event-level watchers.
        status_flipped = event_level and was_empty and len(seats) > 0
        headline = f"🎟️ {w['event_code']} — tickets just opened" if status_flipped and matched else None
        # Festival/hub + kupat-GA watchers — and TM event-level per-perf
        # status seats — carry a status flag per tick (the seat key encodes
        # it), so any transition shows up as an `added` seat — phrase the
        # ping for the new status.
        # Scan matched (not added) so a GA pseudo-seat stripped above can't
        # slap a status headline onto a real-seat ping.
        _fest = next((s for s in matched if s.get("festival") or s.get("ga")), None)
        if _fest:
            _nm = label or w["event_code"]
            headline = {
                "soldout": f"❌ Sold out — {_nm}",
                "lasttickets": f"⚠️ Last tickets — {_nm}",
                "available": f"🎟️ Available again — {_nm}",
                "closed": f"⛔ Sales closed — {_nm}",
            }.get(_fest.get("status"), headline)

        # DICE: a buyable type appearing is a snipe opportunity — make the
        # ping unmistakable and one-tap (perf_url opens the DICE app). Loudest
        # when the watcher was empty (a genuine restock).
        if src_name == "dice" and matched:
            _dnm = ((labels or {}).get("meta") or {}).get("eventName") or label or w["event_code"]
            headline = f"🎯 BUY NOW — {_dnm}" + (" · restocked" if was_empty else "")

        # haku: every ping is registration news — spots back at a charity or
        # the sold-out prose changing. Say so instead of "N new seats".
        if src_name == "haku" and matched:
            _hnm = ((labels or {}).get("meta") or {}).get("eventName") or label or w["event_code"]
            if any(s.get("kind") == "rfar" for s in matched):
                headline = f"🏃 Charity entries — {_hnm}"
            else:
                headline = f"🏃 Registration update — {_hnm}"

        # tmdiscover: every matched seat is a date whose status box just
        # turned buyable (sold out / not-yet-open → last tickets or on sale),
        # so headline the flip instead of "N new seats". A date going the
        # other way is a silent removal and never reaches here.
        # (Keep in sync with watcher_only.py.)
        if src_name == "tmdiscover" and matched:
            _snm = ((labels or {}).get("meta") or {}).get("eventName") or label or w["event_code"]
            _last = all((s.get("status") or "") == "low_availability" for s in matched)
            _n = len(matched)
            _icon, _tail = ("⚠️", " (כרטיסים אחרונים)") if _last else ("🎟️", "")
            headline = f"{_icon} {_n} date{'s' if _n != 1 else ''} just opened{_tail} — {_snm}"

        if enabled and matched:
            result = notify.notify_drop(
                label=label, perf_url=perf_url,
                added_seats=matched, removed_count=len(removed),
                total_now=len(seats), labels=labels,
                channels=enabled,
                headline=headline,
                discord_override=discord_bot.webhook_for(w.get("discord_channel")),
            )
            # Stamp the physical seats we just pinged so they cool down.
            if SEAT_COOLDOWN_SECONDS > 0:
                db.tm_seat_cooldown_mark(
                    wid,
                    [key_fn(s) for s in matched if not (s.get("festival") or s.get("ga"))],
                    now_iso,
                )
        elif not matched and added:
            # All new seats filtered out — record the drop so the user sees
            # the filter is working, but skip the notification.
            result = {"discord": "skipped (filtered)", "email": "skipped (filtered)"}
        else:
            reason = "master-muted" if master_muted else ("watcher-muted" if watcher_muted else "channels-empty")
            result = {"discord": f"skipped ({reason})", "email": f"skipped ({reason})"}

        db.tm_record_drop(
            wid, len(added), len(removed),
            _json.dumps(added, ensure_ascii=False)[:8000],
            _json.dumps(result)[:1000],
            now_iso,
            notify_count=len(matched),
        )
    return len(added), None


def _purge_past_watchers():
    """Auto-delete watchers whose event date has already passed.

    For each watcher we read the cached label payload (no network call) and
    check meta.firstPerfMs / firstPerfText. Once today's local date is past
    the event's date, the watcher has nothing left to watch, so we drop it.

    Returns the number of watchers deleted. Soft-fails on any per-watcher
    error so a single broken cache file can't block the cleanup pass.
    """
    today = datetime.now().date()
    deleted = 0
    for w in db.tm_all_watchers():
        try:
            src_name = w.get("source") or "ticketmaster"
            src = WATCHER_SOURCES.get(src_name, ticketmaster)
            # Event-level watchers don't have a single perf — use the last
            # performance under the event as the date probe (so we only
            # purge after the LAST date of a multi-date show has passed).
            if src is ticketmaster and ticketmaster.is_event_level(w):
                try:
                    perfs = ticketmaster.list_performances(w["event_code"])
                    perfs_sorted = sorted(perfs, key=lambda p: p.get("performanceDate") or 0)
                    probe_perf = str(perfs_sorted[-1].get("performanceCode") or "") if perfs_sorted else None
                except Exception:
                    probe_perf = None
                if not probe_perf:
                    continue
                lbls = src.get_labels(w["event_code"], probe_perf, lang="iw")
            else:
                lbls = src.get_labels(w["event_code"], w["perf_code"], lang="iw")
            meta = (lbls or {}).get("meta") or {}
            event_date = None
            ms = meta.get("firstPerfMs")
            if ms:
                try:
                    event_date = datetime.fromtimestamp(int(ms) / 1000).date()
                except (TypeError, ValueError, OSError):
                    event_date = None
            if event_date is None:
                txt = (meta.get("firstPerfText") or "")[:10]
                try:
                    event_date = datetime.strptime(txt, "%Y-%m-%d").date()
                except ValueError:
                    event_date = None
            if event_date and event_date < today:
                db.tm_delete_watcher(w["id"])
                deleted += 1
        except Exception:
            # Don't let one broken watcher block the rest of the purge.
            traceback.print_exc()
    return deleted


def _kupat_session_for(watchers):
    """Shared kupat browser for a tick, or a no-op when no kupat watcher is
    in the list. (Keep in sync with watcher_only._kupat_session_for.)"""
    if any((w.get("source") or "") == "kupat" for w in watchers):
        return kupat.shared_session()
    return contextlib.nullcontext()


def run_tm_check():
    """Iterate active watchers, run one check each. Lock prevents overlap if
    a tick exceeds the interval. Honors the master_paused setting — when
    set, the tick records the timestamp but skips polling entirely."""
    if not _tm_lock.acquire(blocking=False):
        return
    _last_tm["running"] = True
    try:
        if db.setting_get_bool("master_paused", default=False):
            _last_tm.update(
                at=datetime.now(timezone.utc).isoformat(),
                checked=0, drops=0, errors=0, paused=True,
            )
            return
        # Drop any watchers whose event date has passed before we burn
        # network calls on them.
        _purge_past_watchers()
        watchers = db.tm_active_watchers()
        now_iso = datetime.now(timezone.utc).isoformat()
        drops = 0
        errors = 0
        # One Chromium for every kupat watcher in this tick instead of one
        # each — see kupat.shared_session. No-op when none are due.
        # (Keep in sync with watcher_only.tick.)
        with _kupat_session_for(watchers):
            for w in watchers:
                try:
                    added, err = _check_one_watcher(w, now_iso)
                    if err:
                        errors += 1
                    if added:
                        drops += 1
                except Exception:
                    errors += 1
                    traceback.print_exc()
        _last_tm.update(
            at=now_iso, checked=len(watchers), drops=drops, errors=errors, paused=False,
        )
    finally:
        _last_tm["running"] = False
        _tm_lock.release()


def run_dice_sniper():
    """Fast restock poll for dice watchers that are currently SOLD OUT.

    Runs every DICE_SNIPER_SECONDS. Only fetches dice watchers with zero
    current seats (a watcher that's already on sale has nothing to snipe and
    is left to the normal 60s tick), so the burst on DICE's anonymous API
    stays to the handful of events actually being waited on. Shares
    `_tm_lock` with run_tm_check so the two never double-fire the same
    watcher — whichever grabs the lock first captures the restock; the diff
    against stored seat state makes the loser a no-op."""
    if db.setting_get_bool("master_paused", default=False):
        return
    if not _tm_lock.acquire(blocking=False):
        return  # main tick (or a prior sniper run) still holding the lock
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        targets = [
            w for w in db.tm_active_watchers()
            if (w.get("source") or "") == "dice"
            and not w.get("paused")
            and (w.get("last_seat_count") or 0) == 0
        ]
        for w in targets:
            try:
                _check_one_watcher(w, now_iso)
            except Exception:
                traceback.print_exc()
    finally:
        _tm_lock.release()


def run_tm_sniper():
    """Fast poll for ticketmaster / tmdiscover watchers (TM_SNIPER_SECONDS).

    Same lock discipline as run_dice_sniper. No last_seat_count gate: a TM
    event-level watcher always holds per-perf status pseudo-seats so its
    count is never 0, and the drop worth racing is a seat appearing in a
    sold-out perf's feed while the rest of the event stays on sale."""
    if not TM_SNIPER_ENABLED:
        return
    if db.setting_get_bool("master_paused", default=False):
        return
    if not _tm_lock.acquire(blocking=False):
        return  # main tick (or a prior sniper run) still holding the lock
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        targets = [
            w for w in db.tm_active_watchers()
            if (w.get("source") or "") in TM_SNIPER_SOURCES
            and not w.get("paused")
        ]
        for w in targets:
            try:
                _check_one_watcher(w, now_iso)
            except Exception:
                traceback.print_exc()
    finally:
        _tm_lock.release()


@app.route("/watchers")
def watchers_page():
    return render_template("watchers.html")


# ---------------- Festival + GA Tracker pages -------------------------------
# Both pages share one snapshot store and one sync job. "Festival" = tickchak
# hub events (have capacity → released/sold/left). "GA Tracker" = kupat GA
# events (only tickets-left, no total). Sold-velocity is the drop in
# `available` between snapshots — works for both with or without a total.

# Rolling windows for the "tickets sold in the last …" box. (label, seconds).
FESTIVAL_WINDOWS = [
    ("1h", 3600), ("6h", 6 * 3600), ("24h", 24 * 3600),
    ("3d", 3 * 86400), ("7d", 7 * 86400),
]
FESTIVAL_SYNC_MINUTES = int(os.getenv("KARTIS_FESTIVAL_SYNC_MINUTES") or 10)
_last_sales_sync = {"at": None, "recorded": 0, "error": None, "running": False}
_sales_sync_lock = threading.Lock()


def _counted_fresh_labels(w):
    """Fresh labels for a count-tracked watcher, or None if it isn't one.
    tickchak festival is cheap HTTP; for kupat we only force the (browser)
    fetch once we've confirmed via cache that it's a GA event."""
    src = w.get("source") or ""
    if src == "tickchak":
        lbls = tickchak.get_labels(w["event_code"], w["perf_code"], lang="iw", force=True)
        return lbls if (lbls.get("meta") or {}).get("festival") else None
    if src == "kupat":
        cached = kupat.get_labels(w["event_code"], w["perf_code"], lang="iw")
        if not (cached.get("meta") or {}).get("ga"):
            return None
        return kupat.get_labels(w["event_code"], w["perf_code"], lang="iw", force=True)
    if src == "ticketmaster":
        # TM GA (unnumbered) shows carry counts in the labels meta (from
        # getAllGaBlock — pure HTTP, cheap). Confirm via cache before the
        # forced refresh so seated TM watchers don't refetch every sync.
        cached = ticketmaster.get_labels(w["event_code"], w["perf_code"], lang="iw")
        if not (cached.get("meta") or {}).get("ga"):
            return None
        return ticketmaster.get_labels(w["event_code"], w["perf_code"], lang="iw", force=True)
    return None


def run_sales_sync():
    """Snapshot each festival/GA event's availability so the Festival / GA
    Tracker pages can show how many sold over rolling windows."""
    if not _sales_sync_lock.acquire(blocking=False):
        return
    _last_sales_sync["running"] = True
    try:
        now = datetime.now(timezone.utc)
        now_iso = now.isoformat()
        recorded = 0
        for w in db.tm_all_watchers():
            try:
                lbls = _counted_fresh_labels(w)
            except Exception:
                continue
            if lbls is None:
                continue
            meta = lbls.get("meta") or {}
            avail = meta.get("availSeats")
            if avail is None:
                continue
            cap = meta.get("totalSeats")
            sold = max(0, cap - avail) if cap is not None else None
            db.sales_snapshot_insert(w["source"], w["event_code"], w["perf_code"],
                                     cap, avail, sold, now_iso)
            recorded += 1
        db.sales_snapshot_prune((now - timedelta(days=8)).isoformat())
        _last_sales_sync.update(at=now_iso, recorded=recorded, error=None)
    except Exception as e:
        _last_sales_sync.update(error=str(e))
        traceback.print_exc()
    finally:
        _last_sales_sync["running"] = False
        _sales_sync_lock.release()


def _sales_windows(series, now):
    """Per-window {sold, partial}. `sold` is the SUM of availability
    decreases inside the window — not the net change — so ticket
    releases/refunds (availability going back UP) don't cancel out real
    sales into a misleading zero. `series` is an ascending
    [(captured_at_iso, available), …] list (last ~7 days). A window with no
    baseline older than its start is flagged partial (history too short)."""
    earliest_t = datetime.fromisoformat(series[0][0]) if series else None
    windows = {}
    for key, secs in FESTIVAL_WINDOWS:
        start = now - timedelta(seconds=secs)
        base = None
        after = []
        for t_iso, a in series:
            t = datetime.fromisoformat(t_iso)
            if t <= start:
                base = (t, a)          # latest reading at/before the window start
            else:
                after.append((t, a))
        seq = ([base] if base else []) + after
        if len(seq) < 2:
            windows[key] = {"sold": None, "partial": True}
            continue
        sold = sum(max(0, seq[i - 1][1] - seq[i][1]) for i in range(1, len(seq)))
        windows[key] = {"sold": sold, "partial": base is None}
    return windows, earliest_t


def _sales_totals(shows):
    totals = {k: {"sold": 0, "partial": False} for k, _ in FESTIVAL_WINDOWS}
    for s in shows:
        for key, _ in FESTIVAL_WINDOWS:
            wv = s["windows"][key]
            if wv["sold"] is None:
                totals[key]["partial"] = True
            else:
                totals[key]["sold"] += wv["sold"]
                if wv["partial"]:
                    totals[key]["partial"] = True
    return totals


def _counted_shows(source_name, flag_key, status_key):
    """Build the per-show payload list for one page (festival or GA)."""
    now = datetime.now(timezone.utc)
    src = WATCHER_SOURCES.get(source_name)
    shows = []
    for w in db.tm_all_watchers():
        if (w.get("source") or "") != source_name:
            continue
        try:
            lbls = src.get_labels(w["event_code"], w["perf_code"], lang="iw")
        except Exception:
            continue
        meta = (lbls or {}).get("meta") or {}
        if not meta.get(flag_key):
            continue
        ec, pc = w["event_code"], w["perf_code"]
        total = meta.get("totalSeats")
        avail = meta.get("availSeats")
        series = db.sales_snapshot_series(source_name, ec, pc, (now - timedelta(days=7)).isoformat())
        cur_avail = avail if avail is not None else (series[-1][1] if series else None)
        windows, earliest_t = _sales_windows(series, now)
        sold = max(0, total - cur_avail) if (total is not None and cur_avail is not None) else None
        shows.append({
            "source": source_name,
            "event_code": ec, "perf_code": pc, "label": w.get("label"),
            "event_name": meta.get("eventName"), "when": meta.get("firstPerfText"),
            "status": meta.get(status_key),
            "total": total, "available": cur_avail, "sold": sold,
            "festival_types": lbls.get("blocks") or {},
            "windows": windows, "tracking_since": earliest_t.isoformat() if earliest_t else None,
            # Watcher controls (manage from the page itself).
            "id": w["id"],
            "notify_channels": w.get("notify_channels") or "",
            "muted": bool(w.get("muted")),
            "paused": bool(w.get("paused")),
            "last_check_error": w.get("last_check_error"),
        })
    shows.sort(key=lambda s: (s.get("when") or ""))
    return shows, now


@app.route("/festival")
def festival_page():
    return render_template("festival.html")


@app.route("/api/festival")
def api_festival():
    shows, now = _counted_shows("tickchak", "festival", "festivalStatus")
    return jsonify({
        "shows": shows, "totals": _sales_totals(shows),
        "windows": [k for k, _ in FESTIVAL_WINDOWS],
        "last_sync": _last_sales_sync, "now": now.isoformat(),
    })


@app.route("/ga")
def ga_page():
    return render_template("ga.html")


@app.route("/api/ga")
def api_ga():
    shows, now = _counted_shows("kupat", "ga", "gaStatus")
    tm_shows, _ = _counted_shows("ticketmaster", "ga", "gaStatus")
    shows = sorted(shows + tm_shows, key=lambda s: (s.get("when") or ""))
    return jsonify({
        "shows": shows, "totals": _sales_totals(shows),
        "windows": [k for k, _ in FESTIVAL_WINDOWS],
        "last_sync": _last_sales_sync, "now": now.isoformat(),
    })


@app.route("/market")
def market_page():
    return render_template("market.html")


@app.route("/pacha")
def pacha_page():
    return render_template("pacha.html")


@app.route("/api/pacha")
def api_pacha():
    """Everything the /pacha page needs in one call: each currently-listed
    event's state from pacha_seen_events (incl. the tier breakdown stored
    per tick), its sold-per-window velocity from the shared snapshot series,
    and the monitor status. Events that dropped off pacha-nyc.com (past
    shows) age out via last_seen_at."""
    import json as _json
    now = datetime.now(timezone.utc)
    series_map = db.sales_snapshot_series_bulk((now - timedelta(days=7)).isoformat())
    release_log = db.pacha_release_log_all()
    cutoff = (now - timedelta(hours=24)).isoformat()
    events = []
    for r in db.pacha_all_seen().values():
        if (r.get("last_seen_at") or "") < cutoff:
            continue  # dropped off the site (past / pulled)
        windows, earliest_t = _sales_windows(
            series_map.get(("pacha", r["event_id"], "0"), []), now)
        tiers = []
        if r.get("tiers_json"):
            try:
                tiers = _json.loads(r["tiers_json"])
            except ValueError:
                pass
        events.append({
            "event_id": r["event_id"], "name": r.get("name"),
            "slug": r.get("slug"), "date_text": r.get("date_text"),
            "start_date": r.get("start_date"),
            "on_sale": bool(r.get("on_sale")),
            "ga_price": r.get("ga_price"),
            "ga_sold_out": bool(r.get("ga_sold_out")),
            "ga_release": r.get("ga_release"),
            "ga_available": r.get("ga_available"),
            "ga_quantity": r.get("ga_quantity"),
            "total_available": r.get("total_available"),
            "sold_cum": r.get("sold_cum"),
            "sold_cum_since": r.get("sold_cum_since"),
            "buy_url": r.get("buy_url"),
            "tiers": tiers,
            "releases": [
                {"name": x["tier_name"], "price": x.get("price"),
                 "quantity": x.get("quantity"),
                 "available_last": x.get("available_last"),
                 "used_last": x.get("used_last"),
                 "first_seen_at": x.get("first_seen_at"),
                 "sold_out_at": x.get("sold_out_at")}
                for x in release_log.get(r["event_id"], [])
            ],
            "windows": windows,
            "first_seen_at": r.get("first_seen_at"),
            "last_seen_at": r.get("last_seen_at"),
        })
    events.sort(key=lambda e: e.get("start_date") or "9999")
    return jsonify({
        "events": events,
        "low_stock_threshold": PACHA_LOW_STOCK_THRESHOLD,
        "status": {**_last_pacha_events,
                   "enabled": PACHA_MONITOR_ENABLED,
                   "interval_minutes": PACHA_MONITOR_INTERVAL_MINUTES},
        "windows": [k for k, _ in FESTIVAL_WINDOWS],
        "now": now.isoformat(),
    })


@app.route("/marquee")
def marquee_page():
    return render_template("marquee.html")


# tao's two rooms, in the order the page stacks them. Keys match the
# venue_key tao_events derives (and the auto_source suffix on the tracked
# rows), so a third Tao room only needs a row in tao_events.VENUES.
_MARQUEE_VENUE_ORDER = ["marquee-new-york", "marquee-skydeck"]


def _tao_venues():
    """tao_events.VENUES via the EDM registry, so app.py stays decoupled
    from any single source module. Empty if tao is ever unregistered --
    /marquee then renders no rooms rather than raising."""
    mod = edm_events.CATALOG_SOURCES.get("tao")
    return getattr(mod, "VENUES", {}) if mod else {}


# Tao names its GA ladder "General Admission" / "General Admission - Tier N".
# Everything else on the page is a different PRODUCT, not a rung on that
# ladder: "GA Fast Pass", "VIP Table Share Experience", "Mastercard VIP".
# Mixing them destroys the climb figure -- an event selling GA at $30 next
# to a $200 VIP table reads as a 567% price climb -- so the ladder maths
# runs over GA rows only. tier_group is always NULL on this source, so the
# name is all there is to go on.
_GA_TIER_RE = re.compile(r"^\s*general\s+admission\b", re.I)
_GA_TIER_NUM_RE = re.compile(r"\btier\s*(\d+)", re.I)


def _is_ga_tier(name):
    n = (name or "")
    return bool(_GA_TIER_RE.match(n)) and "fast pass" not in n.lower()


def _tier_ladder(releases, now):
    """Turn one event's release-log rows into the demand signal Marquee
    actually publishes.

    Tao prints no ticket counts anywhere -- the quantity <select> tops out
    at an unpublished per-order limit -- so there is no "N of M left" and no
    units-sold velocity to be had. What Tao DOES publish is a numbered GA
    ladder (Tier 1 -> Tier 2 -> ...), and a tier only steps up once the one
    below it has sold out. So the ladder timings ARE the sales signal:
    how many releases have burned through, and how long each one lasted.

    Two things are deliberately reported as unknown rather than guessed:

    - A row whose first_seen_at EQUALS its sold_out_at was already gone the
      first time we looked. Its true lifetime is unknowable, so hours_open
      stays None and it's flagged already_gone -- calling that "sold out in
      0.0h" would invent the hottest possible reading out of no data.
    - hours_open for a release that was already open when tracking started
      is a floor, not a fact. The page words those as "at least".
    """
    ladder = []
    for r in releases:
        first, sold_out = r.get("first_seen_at"), r.get("sold_out_at")
        already_gone = bool(first and sold_out and first == sold_out)
        hours = None
        if first and not already_gone:
            try:
                end_t = datetime.fromisoformat(sold_out) if sold_out else now
                hours = round(
                    (end_t - datetime.fromisoformat(first)).total_seconds() / 3600, 1)
            except (TypeError, ValueError):
                hours = None
        name = r.get("tier_name")
        num = _GA_TIER_NUM_RE.search(name or "")
        ladder.append({
            "name": name,
            "is_ga": _is_ga_tier(name),
            "tier_num": int(num.group(1)) if num else None,
            "price": r.get("price"),
            "face_price": r.get("face_price"),
            "first_seen_at": first,
            "sold_out_at": sold_out,
            "hours_open": hours,
            "closed": bool(sold_out),
            "already_gone": already_gone,
        })

    ga = [x for x in ladder if x["is_ga"]]
    # Burn stats only count releases we actually watched open AND close.
    burned = [x for x in ga if x["closed"] and not x["already_gone"]
              and x["hours_open"] is not None]
    burn = [x["hours_open"] for x in burned]
    ga_open = [x for x in ga if not x["closed"]]
    # The climb walks the GA ladder in the order the rungs appeared.
    ga_priced = [x for x in ga if x["price"] is not None]
    ga_priced.sort(key=lambda x: (x["first_seen_at"] or "", x["tier_num"] or 0))
    lo = ga_priced[0]["price"] if ga_priced else None
    hi = ga_priced[-1]["price"] if ga_priced else None
    summary = {
        "releases": len(ga),
        "products": len(ladder) - len(ga),
        # Rungs that have closed, split by whether we saw them open.
        "cleared": len([x for x in ga if x["closed"]]),
        "cleared_timed": len(burned),
        "cleared_unseen": len([x for x in ga if x["already_gone"]]),
        "price_open": lo,
        "price_top": hi,
        # Hours the most recently sold-out release lasted -- the freshest
        # read on how fast this room is moving.
        "last_burn_hours": burned[-1]["hours_open"] if burned else None,
        "avg_burn_hours": round(sum(burn) / len(burn), 1) if burn else None,
        # How long the currently-selling release has been up.
        "current_open_hours": max((x["hours_open"] for x in ga_open
                                   if x["hours_open"] is not None), default=None),
    }
    # A climb needs two rungs to compare. With one priced GA release --
    # every event on the day the tracker starts -- lo == hi, and reporting
    # 0.0% would assert "this event's price has held" when the truth is
    # "we have not watched it long enough to know".
    summary["climb_pct"] = (round((hi - lo) / lo * 100, 1)
                            if len(ga_priced) >= 2 and lo and hi is not None
                            else None)
    return ladder, summary


@app.route("/api/marquee")
def api_marquee():
    """Everything /marquee needs in one call: every tracked Tao Group event
    grouped by room, its current release + price, and the tier ladder that
    stands in for the sold-per-window velocity /pacha shows.

    Deliberately NOT reporting counts: lead_available / total_available /
    sold_cum are structurally None for this source (see _tier_ladder), and
    surfacing them as 0 would read as "sold out" rather than "unknown".
    """
    import json as _json
    now = datetime.now(timezone.utc)
    release_log = db.edm_release_log_all()
    seen = db.edm_all_seen()
    by_venue = {}
    for t in db.edm_tracked_all():
        if t.get("source") != "tao":
            continue
        r = seen.get(t["event_id"]) or {}
        # venue_key comes off auto_source ("tao:marquee-skydeck"); a
        # hand-added row has none, so fall back to the venue name.
        vkey = (t.get("auto_source") or "").split(":", 1)[-1]
        if vkey not in _tao_venues():
            vkey = next((k for k, v in _tao_venues().items()
                         if v["name"] == (r.get("venue") or "")), "other")
        tiers = []
        if r.get("tiers_json"):
            try:
                tiers = _json.loads(r["tiers_json"])
            except ValueError:
                pass
        ladder, summary = _tier_ladder(release_log.get(t["event_id"], []), now)
        by_venue.setdefault(vkey, []).append({
            "event_id": t["event_id"], "event_key": t["event_key"],
            "venue_key": vkey,
            "venue": r.get("venue") or _tao_venues().get(vkey, {}).get("name"),
            "name": r.get("name") or t.get("label"),
            "label": t.get("label"),
            "date_text": r.get("date_text"), "start_date": r.get("start_date"),
            "page_url": r.get("page_url") or t.get("url"),
            "paused": bool(t.get("paused")),
            "auto": bool(t.get("auto_source")),
            "on_sale": bool(r.get("on_sale")),
            "sold_out": bool(r.get("sold_out")),
            "min_price": r.get("min_price"),
            "lead_tier": r.get("lead_tier"),
            "lead_price": r.get("lead_price"),
            "tiers": tiers,
            "ladder": ladder,
            **summary,
            "first_seen_at": r.get("first_seen_at"),
            "last_seen_at": r.get("last_seen_at"),
            "last_error": r.get("last_error"),
            "tracked_only": not r,
        })

    venues = []
    extras = sorted(set(by_venue) - set(_MARQUEE_VENUE_ORDER))
    for key in _MARQUEE_VENUE_ORDER + extras:
        evs = by_venue.get(key)
        if not evs:
            continue
        evs.sort(key=lambda e: (e.get("start_date") or "9999", e.get("name") or ""))
        live = [e for e in evs if e["on_sale"]]
        priced = [e["lead_price"] for e in live if e["lead_price"] is not None]
        burns = sorted(e["last_burn_hours"] for e in evs
                       if e["last_burn_hours"] is not None)
        venues.append({
            "key": key,
            "name": _tao_venues().get(key, {}).get("name") or key,
            "calendar_url": _tao_venues().get(key, {}).get("calendar_url"),
            "events": evs,
            "totals": {
                "events": len(evs),
                "on_sale": len(live),
                "sold_out": len([e for e in evs if e["sold_out"]]),
                "cleared": sum(e["cleared"] for e in evs),
                "avg_price": round(sum(priced) / len(priced), 2) if priced else None,
                "median_burn_hours": burns[len(burns) // 2] if burns else None,
            },
        })
    return jsonify({
        "venues": venues,
        "catalogs": dict(_last_edm_catalogs),
        "status": {**_last_edm_events,
                   "enabled": EDM_MONITOR_ENABLED,
                   "interval_minutes": EDM_MONITOR_INTERVAL_MINUTES,
                   "catalog_sync_minutes": EDM_CATALOG_SYNC_MINUTES},
        "now": now.isoformat(),
    })


@app.route("/edm")
def edm_page():
    return render_template("edm.html")


@app.route("/api/edm")
def api_edm():
    """Everything the /edm page needs in one call: every tracked event's
    latest state, its tier breakdown, its release history and (where the
    site publishes counts) its sold-per-window velocity. Unlike /api/pacha
    nothing is aged out — an event is here because it was explicitly
    tracked, so it stays visible until it's removed."""
    import json as _json
    now = datetime.now(timezone.utc)
    series_map = db.sales_snapshot_series_bulk((now - timedelta(days=7)).isoformat())
    release_log = db.edm_release_log_all()
    seen = db.edm_all_seen()
    events = []
    for t in db.edm_tracked_all():
        # tao (Marquee NY / Skydeck) has its own dedicated page at /marquee,
        # where the tier-ladder view its count-less data actually supports
        # replaces the velocity columns here. The monitor still polls it --
        # this is a view filter, not a tracking change.
        if t.get("source") == "tao":
            continue
        r = seen.get(t["event_id"]) or {}
        windows, _earliest = _sales_windows(
            series_map.get((t["source"], t["event_key"], "0"), []), now)
        tiers = []
        if r.get("tiers_json"):
            try:
                tiers = _json.loads(r["tiers_json"])
            except ValueError:
                pass
        events.append({
            "event_id": t["event_id"], "source": t["source"],
            "event_key": t["event_key"], "url": t.get("url"),
            "label": t.get("label"), "paused": bool(t.get("paused")),
            "added_at": t.get("added_at"),
            "name": r.get("name"), "venue": r.get("venue"),
            "date_text": r.get("date_text"), "start_date": r.get("start_date"),
            "page_url": r.get("page_url") or t.get("url"),
            "currency": r.get("currency") or "USD",
            "on_sale": bool(r.get("on_sale")),
            "sold_out": bool(r.get("sold_out")),
            "min_price": r.get("min_price"),
            "lead_tier": r.get("lead_tier"),
            "lead_price": r.get("lead_price"),
            "lead_available": r.get("lead_available"),
            "total_available": r.get("total_available"),
            "total_sold": r.get("total_sold"),
            "sold_cum": r.get("sold_cum"),
            "sold_cum_since": r.get("sold_cum_since"),
            "tiers": tiers,
            "releases": [
                {"name": x["tier_name"], "group": x.get("tier_group"),
                 "price": x.get("price"), "face_price": x.get("face_price"),
                 "available_last": x.get("available_last"),
                 "first_seen_at": x.get("first_seen_at"),
                 "sold_out_at": x.get("sold_out_at")}
                for x in release_log.get(t["event_id"], [])
            ],
            "windows": windows,
            "first_seen_at": r.get("first_seen_at"),
            "last_seen_at": r.get("last_seen_at"),
            "last_error": r.get("last_error"),
            "tracked_only": not r,   # added but not yet successfully fetched
        })
    events.sort(key=lambda e: (e.get("start_date") or "9999", e.get("name") or ""))
    return jsonify({
        "events": events,
        "low_stock_threshold": EDM_LOW_STOCK_THRESHOLD,
        "sources": sorted(set(edm_events.SOURCES) - set(edm_events.CATALOG_SOURCES)),
        # Catalog sources (tao) are surfaced on /marquee, not here.
        "catalog_sources": [],
        "catalogs": {},
        "status": {**_last_edm_events,
                   "enabled": EDM_MONITOR_ENABLED,
                   "interval_minutes": EDM_MONITOR_INTERVAL_MINUTES},
        "windows": [k for k, _ in FESTIVAL_WINDOWS],
        "now": now.isoformat(),
    })


@app.route("/api/edm/add", methods=["POST"])
def api_edm_add():
    """Track another posh / leap / eventim event URL. Validates by actually
    fetching it, so a bad URL fails here instead of erroring every tick."""
    from flask import request
    body = request.get_json(silent=True) or {}
    url = (body.get("url") or "").strip()
    if not url:
        return jsonify({"error": "url required"}), 400
    try:
        source, key = edm_events.parse_target(url)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    try:
        ev = edm_events.fetch_one(source, key)
    except Exception as e:
        return jsonify({"error": f"could not fetch that event: {e}"}), 502
    event_id = db.edm_tracked_add(source, key, url,
                                  datetime.now(timezone.utc).isoformat(),
                                  label=(body.get("label") or "").strip() or None)
    return jsonify({"event_id": event_id, "source": source,
                    "name": ev.get("name"), "min_price": ev.get("min_price"),
                    "on_sale": ev.get("on_sale")})


@app.route("/api/edm/remove", methods=["POST"])
def api_edm_remove():
    from flask import request
    body = request.get_json(silent=True) or {}
    event_id = (body.get("event_id") or "").strip()
    if not event_id:
        return jsonify({"error": "event_id required"}), 400
    n = db.edm_tracked_remove(event_id)
    if not n:
        return jsonify({"error": f"not tracked: {event_id}"}), 404
    return jsonify({"removed": event_id})


@app.route("/api/edm/pause", methods=["POST"])
def api_edm_pause():
    from flask import request
    body = request.get_json(silent=True) or {}
    event_id = (body.get("event_id") or "").strip()
    if not event_id:
        return jsonify({"error": "event_id required"}), 400
    paused = bool(body.get("paused", True))
    n = db.edm_tracked_set_paused(event_id, paused)
    if not n:
        return jsonify({"error": f"not tracked: {event_id}"}), 404
    return jsonify({"event_id": event_id, "paused": paused})


@app.route("/api/edm-events/status")
def api_edm_events_status():
    return jsonify({
        **_last_edm_events,
        "enabled": EDM_MONITOR_ENABLED,
        "interval_minutes": EDM_MONITOR_INTERVAL_MINUTES,
        "low_stock_threshold": EDM_LOW_STOCK_THRESHOLD,
        "tracked_total": len(db.edm_tracked_all()),
        "catalog_sync_minutes": EDM_CATALOG_SYNC_MINUTES,
        "catalogs": dict(_last_edm_catalogs),
    })


@app.route("/api/edm-events/run-now", methods=["POST"])
def api_edm_events_run_now():
    """Manual tick — synchronous, so the response carries the diff summary.
    Works even when the scheduled job is disabled on this machine."""
    if _last_edm_events["running"]:
        return jsonify({"error": "edm monitor already running"}), 429
    run_edm_events()
    return jsonify(dict(_last_edm_events))


@app.route("/dice")
def dice_page():
    return render_template("dice.html")


def _dice_tracked_codes():
    """Every DICE event we track, from both halves: manual /market entries
    and drop watchers. Returns {event_code: source_of_tracking_label}."""
    codes = {}
    for r in db.market_manual_all():
        if r["source"] == "dice":
            codes[str(r["code"])] = "market"
    for w in db.tm_all_watchers():
        if (w.get("source") or "") == "dice":
            codes.setdefault(str(w["event_code"]), "watcher")
            codes[str(w["event_code"])] = "both" if codes[str(w["event_code"])] == "market" else codes[str(w["event_code"])]
    return codes


def _dice_payload(force=False):
    """Shared by GET /api/dice and the refresh endpoint. force=True hits
    the DICE API for every event (and logs changes); otherwise serves the
    1h labels cache."""
    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    log_map = db.dice_tier_log_all()
    events = []
    for code, tracked_by in _dice_tracked_codes().items():
        labels = dice.get_labels(code, "0", force=force)
        meta = (labels or {}).get("meta") or {}
        blocks = (labels or {}).get("blocks") or {}
        if force:
            try:
                db.dice_tier_log_update(code, blocks, now_iso)
            except Exception:
                pass
        types = sorted((
            {"name": b.get("name"), "price": b.get("price"),
             "currency": b.get("currency"), "status": b.get("status"),
             "tier_index": b.get("tier_index"), "tier_name": b.get("tier_name")}
            for b in blocks.values()
        ), key=lambda t: (t["price"] is None, t["price"] or 0))
        on_sale = [t for t in types if t["status"] == "on-sale"]
        history = log_map.get(code, [])
        # Change count = log states beyond each type's first sighting.
        first_states = {}
        changes = 0
        last_change_at = None
        for h in history:
            if h["type_name"] in first_states:
                changes += 1
                last_change_at = h["first_seen_at"]
            else:
                first_states[h["type_name"]] = True
        events.append({
            "event_code": code,
            "tracked_by": tracked_by,
            "name": meta.get("eventName"),
            "venue": " · ".join(v for v in ((meta.get("venueName") or "").strip(),
                                            (meta.get("venueCity") or "").strip()) if v),
            "date_text": meta.get("firstPerfText"),
            "first_date_ms": meta.get("firstPerfMs"),
            "url": dice.perf_url(code),
            "status": meta.get("status"),
            "event_status": meta.get("eventStatus"),
            "sale_end": meta.get("saleEnd"),
            "currency": next((t["currency"] for t in types if t["currency"]), "USD"),
            "min_price": min((t["price"] for t in on_sale if t["price"] is not None), default=None),
            "types": types,
            "types_on_sale": len(on_sale),
            "history": history,
            "changes": changes,
            "last_change_at": last_change_at,
            "fetched_at": labels.get("_fetched_at"),
        })
    events.sort(key=lambda e: e.get("first_date_ms") or float("inf"))
    return {"events": events, "now": now_iso}


@app.route("/api/dice")
def api_dice():
    """Everything the /dice page needs: each tracked dice.fm event's
    current ticket types (1h-cached labels), min on-sale price, and the
    tier/price change log accumulated by the sweeps + refreshes."""
    return jsonify(_dice_payload(force=False))


@app.route("/api/dice/refresh", methods=["POST"])
def api_dice_refresh():
    """Force-fetch every tracked event from the DICE API right now and log
    any tier/price/status changes. Synchronous — a handful of events is a
    couple of seconds."""
    return jsonify(_dice_payload(force=True))


def _dice_vault_accounts():
    """DICE identifies accounts by buyer name + phone, not email, so the
    /dice page enriches each account_email with the matching vault entry's
    plaintext label + phone last-4 (vault rows whose platform mentions dice,
    keyed by username == the account's email). Secrets stay PIN-gated —
    only the non-secret columns are read here."""
    out = {}
    try:
        with db.connect() as conn:
            rows = conn.execute(
                "SELECT label, username, phone FROM vault_accounts "
                "WHERE lower(platform) LIKE '%dice%'"
            ).fetchall()
    except Exception:
        return out
    for r in rows:
        email = (r["username"] or "").strip().lower()
        if not email:
            continue
        digits = re.sub(r"\D", "", r["phone"] or "")
        out[email] = {
            "name": (r["label"] or "").strip(),
            "phone_last4": digits[-4:] if len(digits) >= 4 else "",
        }
    return out


@app.route("/api/dice/purchases")
def api_dice_purchases():
    """The /dice page's Purchases & Holdings section: every auto-recorded
    DICE purchase (from forwarded confirmation emails) grouped by event,
    with per-account bought/transferred/held rows, plus the transfer log.
    Rows are written by mail_intake's dice branch — this is read-only."""
    purchases = db.dice_purchases_all()
    transfers = db.dice_transfers_all()
    sale_links = db.dice_sale_links_by_purchase()
    groups = {}
    for p in purchases:
        key = p.get("event_slug") or ("name:" + (p.get("event_name") or "").lower())
        g = groups.setdefault(key, {
            "event_slug": p.get("event_slug") or "",
            "event_name": p.get("event_name") or "",
            "event_date_iso": p.get("event_date_iso") or "",
            "venue": p.get("venue") or "",
            "url": f"https://dice.fm/event/{p['event_slug']}" if p.get("event_slug") else "",
            "purchases": [],
            "qty": 0, "transferred": 0, "held": 0, "spend": 0.0,
        })
        held = (p.get("qty") or 0) - (p.get("qty_transferred") or 0)
        links = sale_links.get(p["id"], [])
        sold = sum(l.get("qty") or 0 for l in links)
        # avail deliberately ignores transfers — delivery info is unreliable,
        # so sold (user-matched resale-platform sales) is the deduction.
        avail = (p.get("qty") or 0) - sold
        g["purchases"].append({**p, "held": held, "sold": sold, "avail": avail,
                               "sale_links": links})
        g["qty"] += p.get("qty") or 0
        g["transferred"] += p.get("qty_transferred") or 0
        g["held"] += held
        g["sold"] = g.get("sold", 0) + sold
        g["avail"] = g.get("avail", 0) + avail
        g["spend"] += p.get("price_total") or 0.0
        # Prefer a dated/venued row's metadata over an undated one's.
        if not g["event_date_iso"] and p.get("event_date_iso"):
            g["event_date_iso"] = p["event_date_iso"]
        if not g["venue"] and p.get("venue"):
            g["venue"] = p["venue"]
    events = sorted(groups.values(),
                    key=lambda g: (not g["event_date_iso"], g["event_date_iso"], g["event_name"]))
    problem_transfers = [t for t in transfers if t.get("match_status") != "matched"]
    return jsonify({
        "events": events,
        "transfers": transfers,
        "accounts": _dice_vault_accounts(),
        "problem_count": len(problem_transfers),
        "now": datetime.now(timezone.utc).isoformat(),
    })


DICE_LISTED_PLATFORMS = ("viagogo", "crowdvolt", "other")


@app.route("/api/dice/purchases/<purchase_id>/listed", methods=["POST"])
def api_dice_purchase_listed(purchase_id):
    """Record where a purchase's tickets are listed for resale. Body:
    {"viagogo": 2, "crowdvolt": 4, "other": 0} — zero/missing platforms are
    dropped; all zeros clears the marker. Counts are a user note, not
    inventory math, so the only hard rule is non-negative ints."""
    from flask import request
    body = request.get_json(silent=True) or {}
    listed = {}
    for plat in DICE_LISTED_PLATFORMS:
        v = body.get(plat)
        if v in (None, "", 0, "0"):
            continue
        try:
            n = int(v)
        except (TypeError, ValueError):
            return jsonify({"error": f"{plat} must be a whole number"}), 400
        if n < 0:
            return jsonify({"error": f"{plat} must be ≥ 0"}), 400
        if n:
            listed[plat] = n
    payload = json.dumps(listed) if listed else None
    if not db.dice_purchase_set_listed(purchase_id, payload):
        return jsonify({"error": "no such purchase"}), 404
    return jsonify({"ok": True, "listed": listed})


@app.route("/api/dice/sale-candidates")
def api_dice_sale_candidates():
    """Recent scraped sales (viagogo / lysted / crowdvolt) for the /dice
    match-a-sale picker, each with qty_linked showing how much of it is
    already attached to a purchase. With ?event_name=<dice event> the list
    is ranked by fuzzy name similarity (same normalization the DICE
    transfer matcher uses) so the right sale is on top."""
    from flask import request
    sales = db.dice_sale_candidates()
    target = (request.args.get("event_name") or "").strip()
    if target:
        from difflib import SequenceMatcher
        import dice_email
        nt = dice_email._norm_name(target)
        def score(s):
            ns = dice_email._norm_name(s.get("event_name") or "")
            if not nt or not ns:
                return 0.0
            if nt == ns:
                return 1.0
            base = SequenceMatcher(None, nt, ns).ratio()
            # Containment (either direction) is a strong signal — resale
            # listings often add venue/tour suffixes around the artist name.
            if nt in ns or ns in nt:
                base = max(base, 0.9)
            return base
        for s in sales:
            s["score"] = round(score(s), 3)
        sales.sort(key=lambda s: (-s["score"], s.get("sale_date_iso") or ""))
    return jsonify({"sales": sales})


@app.route("/api/dice/purchases/<purchase_id>/link-sale", methods=["POST"])
def api_dice_link_sale(purchase_id):
    """Attach a scraped sale to a purchase. Body: {"source": "viagogo",
    "sale_id": "...", "qty": 2}. avail on the page becomes qty − Σ linked."""
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip().lower()
    sale_id = (body.get("sale_id") or "").strip()
    if source not in db._DICE_SALE_TABLES or not sale_id:
        return jsonify({"error": "source (viagogo/lysted/crowdvolt) and sale_id required"}), 400
    try:
        qty = int(body.get("qty"))
    except (TypeError, ValueError):
        return jsonify({"error": "qty must be a whole number"}), 400
    if qty <= 0:
        return jsonify({"error": "qty must be ≥ 1"}), 400
    try:
        db.dice_sale_link_add(purchase_id, source, sale_id, qty,
                              datetime.now(timezone.utc).isoformat())
    except Exception as e:
        return jsonify({"error": f"already linked? {type(e).__name__}"}), 409
    return jsonify({"ok": True})


@app.route("/api/dice/sale-links/<link_id>/delete", methods=["POST"])
def api_dice_unlink_sale(link_id):
    if not db.dice_sale_link_delete(link_id):
        return jsonify({"error": "no such link"}), 404
    return jsonify({"ok": True})


@app.route("/api/market")
def api_market():
    """Every tracked market row + its sold-per-window velocity. One bulk
    snapshot scan feeds the same _sales_windows math the /ga and /festival
    pages use; rows that also have a watcher simply ride a denser series."""
    now = datetime.now(timezone.utc)
    series_map = db.sales_snapshot_series_bulk((now - timedelta(days=7)).isoformat())
    # Pacha rows carry a lifetime sold-since-tracking counter (snapshots are
    # pruned at ~7d, so the windows can't provide this).
    pacha_seen = db.pacha_all_seen()
    edm_seen = db.edm_all_seen()
    # User-snoozed noise events (water parks etc.) — filtered out of shows
    # and totals, returned separately so the page can list/unhide them.
    hidden_rows = db.market_hidden_active(now.isoformat())
    hidden_keys = {(h["source"], h["event_code"]) for h in hidden_rows}
    # Existing drop watchers, so each row can show WATCH vs already-watching.
    watcher_keys = {
        (wt.get("source") or "ticketmaster", wt["event_code"], str(wt["perf_code"]))
        for wt in db.tm_all_watchers()
    }

    def _watched(r):
        if (r["source"], r["event_code"], str(r["perf_code"])) in watcher_keys:
            return True
        # TM event-level watchers (perf 'ALL') cover every perf of the event;
        # tickchak/dice market rows use perf '0' while their watchers may
        # carry a different perf code — match those on event alone.
        if r["source"] == "ticketmaster":
            return ("ticketmaster", r["event_code"], "ALL") in watcher_keys
        if r["source"] in ("tickchak", "dice"):
            return any(k[0] == r["source"] and k[1] == r["event_code"] for k in watcher_keys)
        return False

    shows = []
    for r in db.market_all():
        if (r["source"], r["event_code"]) in hidden_keys:
            continue
        series = series_map.get((r["source"], r["event_code"], r["perf_code"]), [])
        windows, earliest_t = _sales_windows(series, now)
        avail = r.get("available")
        if avail is None and series:
            avail = series[-1][1]
        # sold_cum / tracked_from come from whichever monitor owns the row:
        # pacha keys on its own event_id, the EDM trackers on "<source>:<key>".
        if r["source"] == "pacha":
            cum_row = pacha_seen.get(r["event_code"])
        elif r["source"] in edm_events.SOURCES:
            cum_row = edm_seen.get(f"{r['source']}:{r['event_code']}")
        else:
            cum_row = None
        shows.append({
            "sold_cum": cum_row.get("sold_cum") if cum_row else None,
            "tracked_from": cum_row.get("sold_cum_since") if cum_row else None,
            "source": r["source"], "event_code": r["event_code"],
            "perf_code": r["perf_code"], "name": r.get("name"),
            "venue": r.get("venue"), "date_text": r.get("date_text"),
            "first_date_ms": r.get("first_date_ms"), "url": r.get("url"),
            "status": r.get("status"), "total": r.get("capacity"),
            "available": avail, "min_price": r.get("min_price"),
            "currency": r.get("currency") or "ILS",
            "manual": bool(r.get("manual")),
            "watched": _watched(r),
            "windows": windows,
            "tracking_since": earliest_t.isoformat() if earliest_t else None,
            "last_seen_at": r.get("last_seen_at"),
            "last_error": r.get("last_error"),
        })
    return jsonify({
        "shows": shows, "totals": _sales_totals(shows),
        "windows": [k for k, _ in FESTIVAL_WINDOWS],
        "hidden": hidden_rows,
        "last_sweep": {k: v for k, v in _last_market.items()},
        "now": now.isoformat(),
    })


_MARKET_HIDE_DURATIONS = {
    "1d": timedelta(days=1),
    "1w": timedelta(weeks=1),
    "1m": timedelta(days=30),
    "always": None,
}


@app.route("/api/market/watch", methods=["POST"])
def api_market_watch():
    """One-click drop watcher from a /market row. Body: {"source",
    "event_code", "perf_code"?}. TM rows become EVENT-LEVEL watchers
    (all dates + the sold-out radar); kupat rows watch that presentation;
    tickchak/dice rows watch the event. zappa/pacha have no watcher source.
    Reuses _add_one_watcher via each source's shorthand, so labels, default
    filters, per-event Discord channels, and the baseline tick all apply."""
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    event_code = str(body.get("event_code") or "").strip()
    perf_code = str(body.get("perf_code") or "").strip()
    if not source or not event_code:
        return jsonify({"error": "source and event_code are required"}), 400
    if source == "ticketmaster":
        shorthand = f"{event_code}/ALL"
    elif source == "kupat":
        if not perf_code or perf_code == "0":
            return jsonify({"error": "kupat rows need a perf_code"}), 400
        shorthand = f"{event_code}/{perf_code}"
    elif source in ("tickchak", "dice"):
        shorthand = event_code
    else:
        return jsonify({"error": f"{source} has no drop-checker source to watch with"}), 400
    try:
        w, warning = _add_one_watcher(shorthand)
    except ValueError as e:
        msg = str(e)
        return jsonify({"error": msg}), (409 if "already watching" in msg else 400)
    return jsonify({"ok": True, "id": w["id"], "label": w.get("label"), "warning": warning})


@app.route("/api/market/hide", methods=["POST"])
def api_market_hide():
    """Snooze an event off the /market page. Body: {"source", "event_code",
    "duration": "1d"|"1w"|"1m"|"always", "name"?}. Hides every date of that
    event; the sweep keeps snapshotting it, so unhiding restores full
    velocity history."""
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    event_code = str(body.get("event_code") or "").strip()
    duration = (body.get("duration") or "").strip()
    if not source or not event_code:
        return jsonify({"error": "source and event_code are required"}), 400
    if duration not in _MARKET_HIDE_DURATIONS:
        return jsonify({"error": f"duration must be one of {sorted(_MARKET_HIDE_DURATIONS)}"}), 400
    now = datetime.now(timezone.utc)
    delta = _MARKET_HIDE_DURATIONS[duration]
    until_iso = (now + delta).isoformat() if delta else None
    db.market_hide(source, event_code, (body.get("name") or "").strip() or None,
                   until_iso, now.isoformat())
    return jsonify({"hidden": {"source": source, "event_code": event_code,
                               "until": until_iso or "always"}})


@app.route("/api/market/unhide", methods=["POST"])
def api_market_unhide():
    """Body: {"source", "event_code"} — bring a snoozed event back."""
    from flask import request
    body = request.get_json(silent=True) or {}
    source = (body.get("source") or "").strip()
    event_code = str(body.get("event_code") or "").strip()
    if not source or not event_code:
        return jsonify({"error": "source and event_code are required"}), 400
    db.market_unhide(source, event_code)
    return jsonify({"unhidden": {"source": source, "event_code": event_code}})


@app.route("/api/market/add", methods=["POST"])
def api_market_add():
    """Manually track a tickchak event/hub or a dice.fm event. Body:
    {"url": ...}. Hub URLs (home.tickchak.co.il/<slug>) expand to their
    member events every sweep; event URLs track that one event. The new
    entry is swept inline so its row appears immediately. Note
    tickchak.parse_url's caveat: numeric /e/ ids and slugs are distinct
    namespaces — paste the public URL you'd buy from."""
    from flask import request
    url = ((request.get_json(silent=True) or {}).get("url") or "").strip()
    if not url:
        return jsonify({"error": "url is required"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    source = "tickchak"
    m = re.search(r"home\.tickchak\.co\.il/([^/?#]+)", url)
    if m:
        kind, code = "hub", m.group(1)
    elif "dice.fm" in url.lower():
        try:
            code, _ = dice.parse_url(url)
        except Exception as e:
            return jsonify({"error": f"not a dice.fm URL I understand: {e}"}), 400
        source, kind = "dice", "event"
    else:
        try:
            code, _ = tickchak.parse_url(url)
        except Exception as e:
            return jsonify({"error": f"not a tickchak URL I understand: {e}"}), 400
        kind = "event"
    db.market_manual_add(source, kind, code, url, now_iso)
    try:
        entities = market.sweep_one_manual(kind, code, source)
    except Exception as e:
        return jsonify({"added": {"kind": kind, "code": code},
                        "warning": f"added, but first fetch failed: {e}"})
    return jsonify({"added": {"kind": kind, "code": code},
                    "entities": len(entities)})


@app.route("/api/market/remove", methods=["POST"])
def api_market_remove():
    """Stop tracking a manual entry. Body: {"id": market_manual.id}. The
    market_events rows it produced stay (history) but stop being refreshed."""
    from flask import request
    id_ = (request.get_json(silent=True) or {}).get("id")
    if not id_:
        return jsonify({"error": "id is required"}), 400
    db.market_manual_remove(id_)
    return jsonify({"removed": id_})


@app.route("/api/market/manual")
def api_market_manual():
    return jsonify({"entries": db.market_manual_all()})


@app.route("/vault")
def vault_page():
    return render_template("vault.html")


def _vault_authed():
    from flask import request
    return vault.token_valid(request.headers.get("X-Vault-Token"))


@app.route("/api/vault/pin", methods=["POST"])
def api_vault_pin():
    """Body: {"pin"}. Sets the PIN on first use, verifies afterwards.
    Returns {"token"} on success; 429 while locked out."""
    from flask import request
    pin = str((request.get_json(silent=True) or {}).get("pin") or "").strip()
    if not (pin.isdigit() and 4 <= len(pin) <= 6):
        return jsonify({"error": "PIN must be 4-6 digits"}), 400
    rem = vault.lockout_remaining()
    if rem:
        return jsonify({"error": f"locked out — try again in {rem}s"}), 429
    if not vault.pin_is_set():
        vault.set_pin(pin, datetime.now(timezone.utc).isoformat())
    elif not vault.check_pin(pin):
        locked = vault.record_attempt(False)
        if locked:
            return jsonify({"error": f"too many attempts — locked for {locked}s"}), 429
        return jsonify({"error": "wrong PIN"}), 401
    vault.record_attempt(True)
    return jsonify({"token": vault.issue_token()})


@app.route("/api/vault")
def api_vault_list():
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"accounts": vault.list_accounts(),
                    "pin_set": vault.pin_is_set()})


@app.route("/api/vault/pin-status")
def api_vault_pin_status():
    return jsonify({"pin_set": vault.pin_is_set()})


@app.route("/api/vault/secret")
def api_vault_secret():
    """?id=N — decrypted secret fields of one account (copy buttons +
    edit-form prefill)."""
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    try:
        id_ = int(request.args.get("id", ""))
    except ValueError:
        return jsonify({"error": "id is required"}), 400
    secrets_ = vault.get_secrets(id_)
    if secrets_ is None:
        return jsonify({"error": "not found"}), 404
    return jsonify(secrets_)


@app.route("/api/vault/add", methods=["POST"])
def api_vault_add():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json(silent=True) or {}
    if not (body.get("platform") or "").strip():
        return jsonify({"error": "platform is required"}), 400
    id_ = vault.add_account(body, datetime.now(timezone.utc).isoformat())
    return jsonify({"added": id_})


@app.route("/api/vault/update", methods=["POST"])
def api_vault_update():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json(silent=True) or {}
    id_ = body.pop("id", None)
    if not id_:
        return jsonify({"error": "id is required"}), 400
    vault.update_account(int(id_), body, datetime.now(timezone.utc).isoformat())
    return jsonify({"updated": id_})


@app.route("/api/vault/delete", methods=["POST"])
def api_vault_delete():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    id_ = (request.get_json(silent=True) or {}).get("id")
    if not id_:
        return jsonify({"error": "id is required"}), 400
    vault.delete_account(int(id_))
    return jsonify({"deleted": id_})


@app.route("/tm")
def tm_page():
    return render_template("tm.html")


@app.route("/api/tm")
def api_tm_list():
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"accounts": vault.tm_list()})


@app.route("/api/tm/secret")
def api_tm_secret():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    try:
        id_ = int(request.args.get("id", ""))
    except ValueError:
        return jsonify({"error": "id is required"}), 400
    secrets_ = vault.tm_get_secrets(id_)
    if secrets_ is None:
        return jsonify({"error": "not found"}), 404
    return jsonify(secrets_)


@app.route("/api/tm/add", methods=["POST"])
def api_tm_add():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json(silent=True) or {}
    id_ = vault.tm_add(body, datetime.now(timezone.utc).isoformat())
    return jsonify({"added": id_})


@app.route("/api/tm/bulk-add", methods=["POST"])
def api_tm_bulk_add():
    """Body: {"rows": [{account_no, email, ...}, ...]} — add many at once
    (the /tm bulk-paste box). Rows with no email AND no password are
    skipped; returns how many were added."""
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    rows = (request.get_json(silent=True) or {}).get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"error": "rows must be a list"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    added = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        if not ((row.get("email") or "").strip() or (row.get("password") or "").strip()):
            continue
        vault.tm_add(row, now_iso)
        added += 1
    return jsonify({"added": added, "skipped": len(rows) - added})


@app.route("/api/tm/update", methods=["POST"])
def api_tm_update():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    body = request.get_json(silent=True) or {}
    id_ = body.pop("id", None)
    if not id_:
        return jsonify({"error": "id is required"}), 400
    vault.tm_update(int(id_), body, datetime.now(timezone.utc).isoformat())
    return jsonify({"updated": id_})


@app.route("/api/tm/delete", methods=["POST"])
def api_tm_delete():
    from flask import request
    if not _vault_authed():
        return jsonify({"error": "unauthorized"}), 401
    id_ = (request.get_json(silent=True) or {}).get("id")
    if not id_:
        return jsonify({"error": "id is required"}), 400
    vault.tm_delete(int(id_))
    return jsonify({"deleted": id_})


@app.route("/vgsales")
def vgsales_page():
    return render_template("vgsales.html")


def _vg_windows(rows, now, tracking_since_iso):
    """Per-window {sold, count, partial} from discrete sale rows (ascending,
    baseline rows already excluded). Unlike _sales_windows there's no delta
    math — a row IS a sale; `sold` sums qty, `count` counts orders. A window
    that started before we began tracking the event is partial (same `*`
    semantics as the availability-based pages)."""
    out = {}
    for key, secs in FESTIVAL_WINDOWS:
        start_iso = (now - timedelta(seconds=secs)).isoformat()
        inside = [r for r in rows if r["observed_at"] >= start_iso]
        partial = not tracking_since_iso or tracking_since_iso > start_iso
        out[key] = {"sold": sum(r.get("qty") or 0 for r in inside),
                    "count": len(inside), "partial": partial}
    return out


def _vg_median(values):
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return None
    mid = len(vals) // 2
    return vals[mid] if len(vals) % 2 else round((vals[mid - 1] + vals[mid]) / 2, 2)


_VG_DATE_TEXT_FMT = "%A, %B %d, %Y"   # "Sunday, August 23, 2026" (popup header)


def _vg_date_iso(listed_row, seen_row):
    """Best-effort sortable date: the scrape mirror's ISO when we're listed,
    else parse the popup header's locale text."""
    if listed_row and listed_row.get("event_date_iso"):
        return listed_row["event_date_iso"]
    txt = (seen_row or {}).get("event_date")
    if txt:
        try:
            return datetime.strptime(txt.strip(), _VG_DATE_TEXT_FMT).date().isoformat()
        except ValueError:
            pass
    return None


@app.route("/api/vgsales")
def api_vgsales():
    """Everything the /vgsales page needs in one call: every tracked event
    (fresh listings + watchlist + anything with sales observed in the last
    7d), its sales-per-window velocity, per-section 7d aggregates, the raw
    recent rows for the expanded view, and the tracker status. All prices
    USD (MarketDataV3 answers in our account currency)."""
    now = datetime.now(timezone.utc)
    fresh = (now - timedelta(hours=viagogo_market_sales.LISTING_FRESH_HOURS)).isoformat()
    listed = {str(r["event_id"]): r for r in db.viagogo_listing_event_ids(fresh)}
    watch = {str(w["event_id"]): w for w in db.vg_watchlist_all()}
    seen = {e["event_id"]: e for e in db.vg_sales_events_all()}
    sales7 = db.vg_market_sales_since((now - timedelta(days=7)).isoformat())
    totals = db.vg_market_sales_totals()

    event_ids = set(listed) | set(watch) | {
        eid for eid, rows in sales7.items() if rows}
    events = []
    for eid in event_ids:
        lr, wr, sr = listed.get(eid), watch.get(eid), seen.get(eid)
        date_iso = _vg_date_iso(lr, sr)
        if not (lr or wr) and date_iso and date_iso[:10] < (now - timedelta(days=1)).date().isoformat():
            continue  # delisted + past — history only, off the page
        rows = sales7.get(eid, [])
        live = [r for r in rows if not r["baseline"]]
        tracking_since = (sr or {}).get("first_seen_at")
        tot = totals.get(eid, {"tickets": 0, "sales": 0})
        sections = {}
        for r in live:
            s = sections.setdefault(r.get("section") or "?", {
                "section": r.get("section") or "?", "sold": 0, "count": 0,
                "prices": [], "last_at": None})
            s["sold"] += r.get("qty") or 0
            s["count"] += 1
            s["prices"].append(r.get("price"))
            s["last_at"] = max(s["last_at"] or "", r["observed_at"])
        section_list = sorted(
            ({**s, "median": _vg_median(s.pop("prices"))} for s in sections.values()),
            key=lambda s: -s["sold"])
        events.append({
            "event_id": eid,
            "name": (lr or {}).get("name") or (sr or {}).get("name")
                    or (wr or {}).get("label") or f"event {eid}",
            "venue": (lr or {}).get("venue") or (sr or {}).get("venue"),
            "date_iso": date_iso,
            "date_text": (sr or {}).get("event_date"),
            "listed": bool(lr), "watch": bool(wr),
            "our_listings": (lr or {}).get("listings") or 0,
            "url": (wr or {}).get("url"),
            "windows": _vg_windows(live, now, tracking_since),
            "total_sold": tot["tickets"], "total_sales": tot["sales"],
            "median_price": _vg_median([r.get("price") for r in live]),
            "last_sale_at": live[-1]["observed_at"] if live else None,
            "sections": section_list,
            "recent": list(reversed(rows[-20:])),
            "tracking_since": tracking_since,
            "last_fetch_at": (sr or {}).get("last_fetch_at"),
            "last_error": (sr or {}).get("last_error"),
        })
    events.sort(key=lambda e: e.get("date_iso") or "9999")
    return jsonify({
        "events": events,
        "status": {**_last_vgsales, "enabled": VGSALES_ENABLED,
                   "interval_minutes": viagogo_market_sales.INTERVAL_MINUTES},
        "windows": [k for k, _ in FESTIVAL_WINDOWS],
        "now": now.isoformat(),
    })


def _parse_vg_event_id(s):
    """Accept a bare numeric id, a public www.viagogo.com/...(E-<id>) URL,
    or an inv URL with an eventId query param."""
    s = (s or "").strip()
    m = re.search(r"/E-(\d+)", s)
    if m:
        return m.group(1)
    m = re.search(r"[?&]eventId=(\d+)", s)
    if m:
        return m.group(1)
    if re.fullmatch(r"\d{6,}", s):
        return s
    return None


@app.route("/api/vgsales/watch", methods=["POST"])
def api_vgsales_watch():
    """Track an event we hold no listing on. Body: {"url": <public URL or
    bare event id>, "label": optional}. First fetch happens inline so the
    row appears immediately; failure still adds the watch (warning, not
    500) — the next tick retries."""
    from flask import request
    body = request.get_json(silent=True) or {}
    url = (body.get("url") or "").strip()
    eid = _parse_vg_event_id(url)
    if not eid:
        return jsonify({"error": "couldn't find a viagogo event id in that"
                                 " (paste a .../E-<id> URL or the bare id)"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    db.vg_watchlist_add(eid, (body.get("label") or "").strip() or None,
                        url if url != eid else None, now_iso)
    try:
        res = viagogo_market_sales.fetch_one(eid, lock_timeout=30)
    except Exception as e:
        return jsonify({"added": eid,
                        "warning": f"added, but first fetch failed: {e}"})
    return jsonify({"added": eid, "name": res["header"].get("name"),
                    "window": res["window"]})


@app.route("/api/vgsales/unwatch", methods=["POST"])
def api_vgsales_unwatch():
    """Stop watching. Observed sales stay (history) but stop refreshing —
    unless we're listed on the event, which keeps it tracked."""
    from flask import request
    eid = ((request.get_json(silent=True) or {}).get("event_id") or "").strip()
    if not eid:
        return jsonify({"error": "event_id is required"}), 400
    db.vg_watchlist_remove(eid)
    return jsonify({"removed": eid})


@app.route("/tools")
def tools_page():
    return render_template("tools.html")


@app.route("/api/kupat-pdf", methods=["POST"])
def api_kupat_pdf():
    """Drive kupat_pdf.render_pdfs from the dashboard form. Synchronous —
    the request body is the URL, the response is the manifest (or error).
    A future iteration could background this via a job queue if 30-second
    requests start being a problem, but for one-off ticket prints the
    synchronous flow keeps the UI dead simple."""
    from flask import request
    body = request.get_json(silent=True) or {}
    url = (body.get("url") or "").strip()
    if not url:
        return jsonify({"error": "URL required"}), 400
    try:
        manifest = kupat_pdf.render_pdfs(url)
    except kupat_pdf.KupatPdfError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected: {e}"}), 500
    return jsonify(manifest)


@app.route("/api/tickchak-pdf", methods=["POST"])
def api_tickchak_pdf():
    """Redact + split an uploaded Tickchak ticket PDF. Multipart upload
    (field name "file") rather than JSON because the PDF can be a few
    hundred KB. Returns the same manifest shape as /api/kupat-pdf."""
    from flask import request
    f = request.files.get("file")
    if not f or not (f.filename or "").lower().endswith(".pdf"):
        return jsonify({"error": "upload a .pdf file"}), 400
    # fitz reads from disk reliably; spool the upload to a tempfile so
    # we don't have to worry about Flask's stream pointer state.
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    try:
        f.save(tmp.name)
        tmp.close()
        manifest = tickchak_pdf.redact_and_split(tmp.name)
    except tickchak_pdf.TickchakPdfError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected: {e}"}), 500
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass
    return jsonify(manifest)


_PACHA_TMP = Path(tempfile.gettempdir()) / "kartis_pacha"


@app.route("/api/pacha-fetch", methods=["POST"])
def api_pacha_fetch():
    """Fetch Pacha tickets from Gmail and bundle the first-page PDFs into a zip
    for browser download. The dashboard runs on the VPS, which can't write to
    the user's PC — so we hand the tickets back through the browser instead of
    saving to disk. Synchronous, lock-guarded. Body: {days: int}. Returns the
    summary plus a download_url, or {error}."""
    from flask import request
    import time
    body = request.get_json(silent=True) or {}
    try:
        days = int(body.get("days") or pacha_tickets.DEFAULT_LOOKBACK_DAYS)
    except (TypeError, ValueError):
        return jsonify({"error": "days must be a number"}), 400
    days = max(1, min(days, 365))
    if not _pacha_lock.acquire(blocking=False):
        return jsonify({"error": "a Pacha fetch is already running"}), 429
    try:
        zip_bytes, summary = pacha_tickets.fetch_zip(days=days)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500
    finally:
        _pacha_lock.release()

    if zip_bytes:
        _PACHA_TMP.mkdir(parents=True, exist_ok=True)
        # Sweep zips older than an hour so the temp dir doesn't grow.
        cutoff = time.time() - 3600
        for old in _PACHA_TMP.glob("*.zip"):
            try:
                if old.stat().st_mtime < cutoff:
                    old.unlink()
            except OSError:
                pass
        token = uuid.uuid4().hex
        (_PACHA_TMP / f"{token}.zip").write_bytes(zip_bytes)
        groups = summary.get("groups") or []
        if len(groups) == 1:
            g = groups[0]
            summary["download_name"] = f"{g['event']} {g['date'] or ''}".strip() + " tickets.zip"
        else:
            summary["download_name"] = "Pacha tickets.zip"
        summary["download_url"] = f"/api/pacha-download/{token}"
    return jsonify(summary)


@app.route("/api/pacha-download/<token>")
def api_pacha_download(token):
    """Serve a previously-built Pacha ticket zip (see api_pacha_fetch)."""
    from flask import request
    if not re.fullmatch(r"[0-9a-f]{32}", token or ""):
        return jsonify({"error": "bad token"}), 400
    path = _PACHA_TMP / f"{token}.zip"
    if not path.exists():
        return jsonify({"error": "download expired — fetch again"}), 404
    name = request.args.get("name") or "Pacha tickets.zip"
    return send_file(str(path), as_attachment=True, download_name=name,
                     mimetype="application/zip")


# systemd unit for the scraper's Chrome on the VPS. Override via env if the
# unit is named differently. Exposed on the Tools page so finalizing a
# Lysted/Viagogo/CrowdVolt re-login (done in the noVNC window) needs no SSH.
CHROME_SERVICE = os.getenv("KARTIS_CHROME_SERVICE", "kartis-chrome")
_chrome_restart_lock = threading.Lock()


@app.route("/api/chrome/restart", methods=["POST"])
def api_chrome_restart():
    """Restart the scraper's Chrome service (`sudo systemctl restart
    kartis-chrome`). After re-logging into a source in the noVNC window, Chrome
    must be bounced so the scraper re-attaches with the fresh session — this
    lets the user do that from the dashboard instead of SSHing in. Only
    meaningful on the Linux/systemd VPS; returns a friendly error elsewhere.
    Uses `sudo -n` so a missing NOPASSWD rule fails fast instead of hanging."""
    import shutil
    if not _chrome_restart_lock.acquire(blocking=False):
        return jsonify({"error": "a restart is already running"}), 429
    try:
        if not shutil.which("systemctl"):
            return jsonify({"error": "systemctl not found — this control only works on the VPS"}), 400
        try:
            proc = subprocess.run(
                ["sudo", "-n", "systemctl", "restart", CHROME_SERVICE],
                capture_output=True, text=True, timeout=60,
            )
        except subprocess.TimeoutExpired:
            return jsonify({"error": "restart timed out after 60s"}), 504
        if proc.returncode != 0:
            msg = (proc.stderr or proc.stdout or "").strip() or f"exit {proc.returncode}"
            return jsonify({"error": f"systemctl failed: {msg}"}), 500
        return jsonify({"ok": True, "service": CHROME_SERVICE})
    finally:
        _chrome_restart_lock.release()


@app.route("/api/chrome/open-logins", methods=["POST"])
def api_chrome_open_logins():
    """Open the Lysted/Viagogo/CrowdVolt login tabs in the box's Chrome (via the
    CDP /json/new endpoint on :9222) so they're already waiting when the user
    opens the noVNC window to re-sign-in. Dedupes against tabs already open —
    repeat clicks (or a source you're still logged into) won't spam duplicates.
    Reuses login.DEFAULT_URLS so the tab list stays in one place."""
    import urllib.request
    from urllib.parse import urlparse
    import login as login_mod

    cdp_main = os.getenv("KARTIS_CDP_URL", "http://localhost:9222")
    # CrowdVolt lives in its own Chrome when configured (kartis-chrome-cv,
    # residential proxy) — its login tab must open THERE or the session
    # lands in the wrong profile.
    cdp_cv = os.getenv("KARTIS_CDP_URL_CROWDVOLT", "").strip() or cdp_main
    if not login_mod.is_chrome_running():
        return jsonify({"error": "Chrome isn't running — click Restart Chrome first."}), 503

    def base_domain(url_or_host):
        host = urlparse(url_or_host).hostname if "//" in url_or_host else url_or_host
        parts = (host or "").split(".")
        return ".".join(parts[-2:]) if len(parts) >= 2 else (host or "")

    def cdp_for(url):
        return cdp_cv if base_domain(url) == "crowdvolt.com" else cdp_main

    # Domains already open per browser, so we don't reopen a source the
    # user is on/into.
    open_domains = {}
    for endpoint in {cdp_main, cdp_cv}:
        found = set()
        try:
            with urllib.request.urlopen(f"{endpoint}/json/list", timeout=8) as r:
                for t in json.load(r):
                    if t.get("type") == "page":
                        found.add(base_domain(t.get("url", "")))
        except Exception:
            pass
        open_domains[endpoint] = found

    opened, skipped, errors = [], [], []
    for url in login_mod.DEFAULT_URLS:
        endpoint = cdp_for(url)
        if base_domain(url) in open_domains[endpoint]:
            skipped.append(url)
            continue
        try:
            req = urllib.request.Request(f"{endpoint}/json/new?{url}", method="PUT")
            with urllib.request.urlopen(req, timeout=8) as r:
                (opened if r.status == 200 else errors).append(url)
        except Exception as e:
            errors.append(f"{url} ({type(e).__name__})")

    if errors and not opened and not skipped:
        return jsonify({"error": "; ".join(errors)}), 502
    return jsonify({"ok": True, "opened": opened, "skipped": skipped, "errors": errors})


@app.route("/api/watchers")
def api_watchers():
    # Sweep past-event watchers on every page load so the UI doesn't have to
    # wait for the next periodic tick to clean them out.
    _purge_past_watchers()
    watchers = db.tm_all_watchers()
    drops = db.tm_recent_drops(limit=200)
    settings = db.all_settings()
    # Enrich each watcher with venue capacity from its cached labels — pure
    # disk read, no network. Lets the UI show "53 / 4125 (98.7% sold)"
    # without holding a venue total in tm_watchers.
    for w in watchers:
        src_name = w.get("source") or "ticketmaster"
        src = WATCHER_SOURCES.get(src_name, ticketmaster)
        try:
            # Event-level watchers span multiple perfs and don't have a single
            # venue-capacity number — skip the capacity enrichment for them.
            if src is ticketmaster and ticketmaster.is_event_level(w):
                w["total_seats"] = None
                w["available_seats"] = None
                continue
            lbls = src.get_labels(w["event_code"], w["perf_code"], lang="iw")
            meta = (lbls or {}).get("meta") or {}
            w["total_seats"] = meta.get("totalSeats")
            # Sources that expose a per-tick available QUANTITY (rather
            # than just a per-type list count) — currently tickchak —
            # surface it here so the dashboard can show real "X / Y"
            # ratios. None for sources where last_seat_count is already
            # the right number.
            w["available_seats"] = meta.get("availSeats")
            # Festival/hub tickchak events: surface the status flag and the
            # per-type released/sold/left breakdown so the watchers table
            # can show a badge + an expandable type list.
            if meta.get("festival"):
                w["festival_status"] = meta.get("festivalStatus")
                w["festival_types"] = lbls.get("blocks") or {}
        except Exception:
            w["total_seats"] = None
            w["available_seats"] = None
    return jsonify({
        "watchers": watchers,
        "drops": drops,
        "last_tm": _last_tm,
        "interval_seconds": TM_CHECK_INTERVAL_SECONDS,
        "check_enabled": TM_CHECK_ENABLED,
        "settings": {
            "master_paused": settings.get("master_paused", "false") in ("1", "true", "True"),
            "master_muted": settings.get("master_muted", "false") in ("1", "true", "True"),
        },
        "sources": list(WATCHER_SOURCES.keys()),
    })


def _add_one_watcher(url, label=None):
    """Shared logic between single-add and bulk-add. Returns (watcher_dict, warning) on success
    or raises ValueError with a user-readable message on failure."""
    url = (url or "").strip()
    src_name, src = _detect_source(url)
    try:
        event_code, perf_code = src.parse_url(url)
    except Exception as e:
        raise ValueError(f"{src_name}: {e}")
    for existing in db.tm_all_watchers():
        if (
            (existing.get("source") or "ticketmaster") == src_name
            and existing["event_code"] == event_code
            and existing["perf_code"] == perf_code
        ):
            raise ValueError(f"already watching {src_name} {event_code}/{perf_code}")
    wid = "tmw-" + uuid.uuid4().hex[:12]
    final_label = (label or "").strip()
    if not final_label:
        try:
            # Event-level watchers don't have a single perf for labels —
            # probe the first real perf so we still get the event name +
            # venue, then tag it as covering all dates.
            if src is ticketmaster and ticketmaster.is_event_level({"perf_code": perf_code}):
                perfs = ticketmaster.list_performances(event_code)
                probe_perf = str((perfs[0] or {}).get("performanceCode") or "")
                lbls = src.get_labels(event_code, probe_perf, lang="iw") if probe_perf else None
                base = src.event_summary(lbls) if lbls else event_code
                final_label = f"{base} — all dates" if base else f"{event_code} — all dates"
            else:
                lbls = src.get_labels(event_code, perf_code, lang="iw")
                final_label = src.event_summary(lbls) or f"{event_code}/{perf_code}"
        except Exception:
            final_label = f"{event_code}/{perf_code}"
    db.tm_insert_watcher({
        "id": wid, "label": final_label, "source": src_name,
        "event_code": event_code, "perf_code": perf_code,
        "paused": 0, "muted": 0, "notify_channels": "discord,email",
        # Each source picks its own default filter. Seated sources
        # (ticketmaster, kupat) default to {min_group_size: 2} (exclude
        # singles); GA sources (tickchak) override to {min_group_size: 1}
        # since adjacency doesn't apply to ticket-type buckets.
        "filters": json.dumps(getattr(src, "DEFAULT_FILTERS", {"min_group_size": 2})),
        # Per-event Discord routing: auto-derive a dateless channel name from
        # the label so all dates of one artist share a channel. Only when the
        # bot is configured — otherwise everything stays on the shared drops
        # webhook and the column stays NULL.
        "discord_channel": (discord_bot.slugify_label(final_label) or None) if discord_bot.configured() else None,
    }, datetime.now(timezone.utc).isoformat())
    w = db.tm_get_watcher(wid)
    warning = None
    try:
        _check_one_watcher(w, datetime.now(timezone.utc).isoformat())
    except Exception as e:
        traceback.print_exc()
        warning = str(e)
    return db.tm_get_watcher(wid), warning


@app.route("/api/watchers/add", methods=["POST"])
def api_watchers_add():
    from flask import request
    body = request.get_json(silent=True) or {}
    try:
        w, warning = _add_one_watcher(body.get("url"), body.get("label"))
    except ValueError as e:
        # Use 409 for dedupe, 400 for parse errors — caller distinguishes.
        msg = str(e)
        return jsonify({"error": msg}), (409 if "already watching" in msg else 400)
    return jsonify({"ok": True, "id": w["id"], "warning": warning})


@app.route("/api/watchers/bulk-add", methods=["POST"])
def api_watchers_bulk_add():
    """Accept a textarea of URLs (one per line) and add each. Returns a
    per-line outcome so the UI can show what worked + what didn't."""
    from flask import request
    body = request.get_json(silent=True) or {}
    text = body.get("urls") or ""
    results = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        try:
            w, warning = _add_one_watcher(line)
            results.append({"url": line, "ok": True, "id": w["id"], "label": w.get("label"), "warning": warning})
        except ValueError as e:
            results.append({"url": line, "ok": False, "error": str(e)})
        except Exception as e:
            traceback.print_exc()
            results.append({"url": line, "ok": False, "error": f"{type(e).__name__}: {e}"})
    added = sum(1 for r in results if r["ok"])
    return jsonify({"added": added, "results": results})


@app.route("/api/watchers/update", methods=["POST"])
def api_watchers_update():
    """Change per-watcher settings: muted, notify_channels, label, filters.
    The UI uses this for the inline mute toggle, channel dropdown, and
    the filters modal save button."""
    from flask import request
    body = request.get_json(silent=True) or {}
    wid = (body.get("id") or "").strip()
    if not wid or not db.tm_get_watcher(wid):
        return jsonify({"error": "watcher not found"}), 404
    fields = {}
    if "muted" in body:
        fields["muted"] = 1 if body["muted"] else 0
    if "notify_channels" in body:
        chans = body["notify_channels"]
        if isinstance(chans, list):
            chans = ",".join(c.strip().lower() for c in chans if c)
        chans = (chans or "").strip().lower()
        valid = {c for c in chans.split(",") if c in {"discord", "email"}}
        fields["notify_channels"] = ",".join(sorted(valid))
    if "label" in body:
        new_label = (body["label"] or "").strip()
        if new_label:
            fields["label"] = new_label
    if "discord_channel" in body:
        # Empty string clears it (back to the shared drops channel). Names
        # are normalized the same way the auto-default is, so typing
        # "Peer Tasi" and the derived "peer-tasi" land on one channel.
        raw = (body["discord_channel"] or "").strip()
        fields["discord_channel"] = discord_bot.slugify_label(raw) or None if raw else None
    if "filters" in body:
        # Validate + normalize the filter object before storing.
        f = body["filters"] or {}
        if not isinstance(f, dict):
            return jsonify({"error": "filters must be an object"}), 400
        out_f = {}
        try:
            mgs = int(f.get("min_group_size") or 1)
        except (TypeError, ValueError):
            mgs = 1
        if mgs > 1:
            out_f["min_group_size"] = mgs
        excl = f.get("exclude_sections") or []
        if isinstance(excl, list) and excl:
            out_f["exclude_sections"] = [str(x) for x in excl if x]
        for k in ("min_price", "max_price"):
            v = f.get(k)
            if v in (None, ""):
                continue
            try:
                out_f[k] = float(v)
            except (TypeError, ValueError):
                pass
        fields["filters"] = json.dumps(out_f) if out_f else None
    if not fields:
        return jsonify({"error": "no fields to update"}), 400
    db.tm_update_watcher(wid, fields)
    return jsonify({"ok": True, "fields": fields})


@app.route("/api/watchers/sections")
def api_watchers_sections():
    """Returns the list of {code, name, price} for a watcher's blocks/sections,
    so the filter modal can populate its multi-select. Pulls from the cached
    labels — fast, no extra network calls in the common case."""
    from flask import request
    wid = (request.args.get("id") or "").strip()
    w = db.tm_get_watcher(wid) if wid else None
    if not w:
        return jsonify({"error": "watcher not found"}), 404
    src_name = w.get("source") or "ticketmaster"
    src = WATCHER_SOURCES.get(src_name, ticketmaster)
    try:
        # Event-level watchers don't have a single perf — probe the first
        # performance under the event so the filter modal still gets a
        # representative section/price list. Most multi-date shows share
        # the same venue + block layout across perfs.
        if src is ticketmaster and ticketmaster.is_event_level(w):
            perfs = ticketmaster.list_performances(w["event_code"])
            probe_perf = str((perfs[0] or {}).get("performanceCode") or "")
            labels = src.get_labels(w["event_code"], probe_perf, lang="iw") if probe_perf else {}
        else:
            labels = src.get_labels(w["event_code"], w["perf_code"], lang="iw")
    except Exception as e:
        return jsonify({"error": f"label fetch failed: {e}"}), 500
    blocks = (labels or {}).get("blocks") or {}
    rows = [
        {"code": code, "name": info.get("name") or code, "price": info.get("price")}
        for code, info in blocks.items()
    ]
    # Priced sections first (sorted by price ascending), then unpriced ones
    # alphabetically — keeps the relevant inventory on top for the user.
    rows.sort(key=lambda r: (r["price"] is None, r["price"] or 0, r["code"]))
    return jsonify({"sections": rows, "filters": json.loads(w.get("filters")) if w.get("filters") else None})


@app.route("/api/watchers/settings", methods=["POST"])
def api_watchers_settings():
    """Master pause / master mute. Body: {key: 'master_paused'|'master_muted', value: bool}."""
    from flask import request
    body = request.get_json(silent=True) or {}
    key = (body.get("key") or "").strip()
    if key not in {"master_paused", "master_muted"}:
        return jsonify({"error": "key must be master_paused or master_muted"}), 400
    value = "true" if body.get("value") else "false"
    db.setting_set(key, value, datetime.now(timezone.utc).isoformat())
    return jsonify({"ok": True, "key": key, "value": value})


@app.route("/api/watchers/stop-everything", methods=["POST"])
def api_watchers_stop_everything():
    """One-click flip both master_paused and master_muted to true. Used by
    the big STOP button on the dashboard."""
    now_iso = datetime.now(timezone.utc).isoformat()
    db.setting_set("master_paused", "true", now_iso)
    db.setting_set("master_muted", "true", now_iso)
    return jsonify({"ok": True})


@app.route("/api/watchers/resume-everything", methods=["POST"])
def api_watchers_resume_everything():
    now_iso = datetime.now(timezone.utc).isoformat()
    db.setting_set("master_paused", "false", now_iso)
    db.setting_set("master_muted", "false", now_iso)
    return jsonify({"ok": True})


@app.route("/api/watchers/toggle", methods=["POST"])
def api_watchers_toggle():
    from flask import request
    body = request.get_json(silent=True) or {}
    wid = (body.get("id") or "").strip()
    paused = 1 if body.get("paused") else 0
    if not wid:
        return jsonify({"error": "id required"}), 400
    db.tm_update_watcher(wid, {"paused": paused})
    return jsonify({"ok": True})


@app.route("/api/watchers/delete", methods=["POST"])
def api_watchers_delete():
    from flask import request
    body = request.get_json(silent=True) or {}
    wid = (body.get("id") or "").strip()
    if not wid:
        return jsonify({"error": "id required"}), 400
    db.tm_delete_watcher(wid)
    return jsonify({"ok": True})


@app.route("/api/watchers/test-notify", methods=["POST"])
def api_watchers_test_notify():
    from flask import request
    body = request.get_json(silent=True) or {}
    label = (body.get("label") or "Kartis test").strip()
    result = notify.send_test(label=label)
    return jsonify({"ok": True, "result": result})


@app.route("/api/watchers/check-now", methods=["POST"])
def api_watchers_check_now():
    """Force a single tick for one watcher (or all if no id given). Useful
    when the user wants to verify a watcher right after adding it without
    waiting for the next scheduled tick."""
    from flask import request
    body = request.get_json(silent=True) or {}
    wid = (body.get("id") or "").strip()
    now_iso = datetime.now(timezone.utc).isoformat()
    if wid:
        w = db.tm_get_watcher(wid)
        if not w:
            return jsonify({"error": "watcher not found"}), 404
        added, err = _check_one_watcher(w, now_iso)
        return jsonify({"ok": True, "added": added, "error": err})
    threading.Thread(target=run_tm_check, daemon=True).start()
    return jsonify({"started": True})


# --- Viagogo auto-pricer API ----------------------------------------------

@app.route("/api/pricer/status")
def api_pricer_status():
    return jsonify({
        "last": _last_pricer,
        "master_enabled": viagogo_pricer.master_enabled(),
        "dry_run": viagogo_pricer.dry_run_enabled(),
        "undercut": viagogo_pricer.undercut_amount(),
        "max_drop_pct": viagogo_pricer.max_drop_pct(),
        "drop_window_hours": viagogo_pricer.drop_window_hours(),
        "drop_cap_enabled": viagogo_pricer.drop_cap_enabled(),
        "ignore_singles": viagogo_pricer.ignore_single_competitors(),
        "interval_minutes": PRICER_INTERVAL_MINUTES,
        "fast_interval_minutes": PRICER_FAST_INTERVAL_MINUTES,
        "fast_events": viagogo_pricer.fast_event_tokens(),
        "configs": db.pricer_config_all(),
        "log": db.pricer_log_recent(50),
    })


@app.route("/api/pricer/config", methods=["POST"])
def api_pricer_config():
    """Per-listing pricer config. Body: {listing_id, enabled?, floor_price?,
    allow_raise?, resume?}. Enabling requires a floor > 0. resume clears a
    manual-change pause and re-adopts the live price on the next tick."""
    from flask import request
    body = request.get_json(silent=True) or {}
    listing_id = str(body.get("listing_id") or "").strip()
    if not listing_id:
        return jsonify({"error": "listing_id required"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    fields = {}
    if body.get("resume"):
        fields.update(paused=0, paused_reason=None, paused_at=None,
                      last_set_price=None, last_set_at=None)
    if "floor_price" in body:
        try:
            floor = float(body["floor_price"]) if body["floor_price"] not in (None, "") else None
        except (TypeError, ValueError):
            return jsonify({"error": "floor_price must be a number"}), 400
        fields["floor_price"] = floor
    if "allow_raise" in body:
        fields["allow_raise"] = 1 if body.get("allow_raise") else 0
    if "no_drop_cap" in body:
        fields["no_drop_cap"] = 1 if body.get("no_drop_cap") else 0
    if "compete_sections" in body:
        secs = body.get("compete_sections")
        if secs in (None, []):
            fields["compete_sections"] = None
        else:
            if not isinstance(secs, list) or not all(isinstance(s, str) for s in secs):
                return jsonify({"error": "compete_sections must be a list of section names"}), 400
            cleaned = sorted({viagogo_pricer._norm_section(s) for s in secs if s.strip()})
            fields["compete_sections"] = json.dumps(cleaned) if cleaned else None
    for fp_key in ("compete_include", "compete_exclude"):
        if fp_key in body:
            fps = body.get(fp_key)
            if fps in (None, []):
                fields[fp_key] = None
                continue
            if not isinstance(fps, list) or not all(isinstance(f, dict) for f in fps):
                return jsonify({"error": f"{fp_key} must be a list of {{s,r,q}} objects"}), 400
            cleaned = [{"s": viagogo_pricer._norm_section(f.get("s")),
                        "r": (f.get("r") or "").strip(),
                        "q": f.get("q")} for f in fps]
            fields[fp_key] = json.dumps(cleaned)
    if "enabled" in body:
        enabled = 1 if body.get("enabled") else 0
        if enabled:
            existing = db.pricer_config_get(listing_id) or {}
            floor = fields.get("floor_price", existing.get("floor_price"))
            if not floor or floor <= 0:
                return jsonify({"error": "a floor price > 0 is required to enable"}), 400
            # (re)enabling always adopts the current live price as baseline
            fields.update(last_set_price=None, last_set_at=None,
                          paused=0, paused_reason=None, paused_at=None)
        fields["enabled"] = enabled
    if not fields:
        return jsonify({"error": "nothing to update"}), 400
    db.pricer_config_set(listing_id, fields, now_iso)
    return jsonify({"ok": True, "config": db.pricer_config_get(listing_id)})


@app.route("/api/pricer/settings", methods=["POST"])
def api_pricer_settings():
    """Global pricer settings. Body: {master_enabled?, dry_run?, undercut?}."""
    from flask import request
    body = request.get_json(silent=True) or {}
    now_iso = datetime.now(timezone.utc).isoformat()
    out = {}
    if "master_enabled" in body:
        v = "true" if body.get("master_enabled") else "false"
        db.setting_set("pricer_master_enabled", v, now_iso)
        out["master_enabled"] = v
    if "dry_run" in body:
        v = "true" if body.get("dry_run") else "false"
        db.setting_set("pricer_dry_run", v, now_iso)
        out["dry_run"] = v
    if "undercut" in body:
        try:
            u = float(body["undercut"])
        except (TypeError, ValueError):
            return jsonify({"error": "undercut must be a number"}), 400
        if not (0 < u <= 50):
            return jsonify({"error": "undercut must be between 0 and 50"}), 400
        db.setting_set("pricer_undercut", f"{u:.2f}", now_iso)
        out["undercut"] = u
    if "ignore_singles" in body:
        v = "true" if body.get("ignore_singles") else "false"
        db.setting_set("pricer_ignore_singles", v, now_iso)
        out["ignore_singles"] = v
    if "max_drop_pct" in body:
        try:
            pct = float(body["max_drop_pct"])
        except (TypeError, ValueError):
            return jsonify({"error": "max_drop_pct must be a number"}), 400
        if not (1 <= pct <= 90):
            return jsonify({"error": "max_drop_pct must be between 1 and 90"}), 400
        db.setting_set("pricer_max_drop_pct", f"{pct:g}", now_iso)
        out["max_drop_pct"] = pct
    if "drop_cap_enabled" in body:
        v = "true" if body.get("drop_cap_enabled") else "false"
        db.setting_set("pricer_drop_cap_enabled", v, now_iso)
        out["drop_cap_enabled"] = v
    if "fast_events" in body:
        # Free text: comma/newline separated event names or ids. Stored
        # verbatim (viagogo_pricer.fast_event_tokens does the parsing) so
        # the box round-trips what the user typed.
        raw = body.get("fast_events")
        if isinstance(raw, (list, tuple)):
            raw = ", ".join(str(t).strip() for t in raw if str(t).strip())
        raw = (raw or "").strip()
        if len(raw) > 500:
            return jsonify({"error": "fast_events too long"}), 400
        db.setting_set("pricer_fast_events", raw, now_iso)
        out["fast_events"] = viagogo_pricer.fast_event_tokens()
    if not out:
        return jsonify({"error": "nothing to update"}), 400
    return jsonify({"ok": True, **out})


@app.route("/api/pricer/run-now", methods=["POST"])
def api_pricer_run_now():
    if _last_pricer["running"]:
        return jsonify({"error": "pricer already running"}), 429
    threading.Thread(target=run_pricer, kwargs={"force_full": True},
                     daemon=True).start()
    return jsonify({"started": True})


@app.route("/api/pricer/market/<event_id>")
def api_pricer_market(event_id):
    """Last market snapshot for an event (written by every tick). Powers the
    /pricer market panel's section picker."""
    snap = db.market_snapshot_get(event_id)
    if not snap:
        return jsonify({"event_id": event_id, "fetched_at": None, "rows": []})
    try:
        rows = json.loads(snap["rows_json"])
    except (TypeError, ValueError):
        rows = []
    return jsonify({"event_id": event_id, "fetched_at": snap["fetched_at"],
                    "rows": rows})


@app.route("/api/pricer/market/<event_id>/refresh", methods=["POST"])
def api_pricer_market_refresh(event_id):
    """Live MarketDataV3 fetch (~30s, needs the CDP Chrome). 429 while a
    pricer tick holds the run lock."""
    if _last_pricer["running"]:
        return jsonify({"error": "pricer tick running — try again in a minute"}), 429
    try:
        rows = viagogo_pricer.refresh_market_snapshot(event_id)
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 502
    return jsonify({"event_id": event_id,
                    "fetched_at": datetime.now(timezone.utc).isoformat(),
                    "rows": rows})


# --- CrowdVolt auto-pricer API --------------------------------------------

@app.route("/api/cvpricer/status")
def api_cvpricer_status():
    rows = db.cv_listings_all()
    cfgs = db.cv_pricer_config_all()
    for r in rows:
        cfg = cfgs.get(r["ask_uqid"]) or {}
        r["pricer_enabled"] = bool(cfg.get("enabled"))
        r["pricer_floor"] = cfg.get("floor_price")
        r["pricer_paused"] = bool(cfg.get("paused"))
        r["pricer_paused_reason"] = cfg.get("paused_reason")
        r["pricer_last_set_price"] = cfg.get("last_set_price")
        r["pricer_no_drop_cap"] = bool(cfg.get("no_drop_cap"))
    return jsonify({
        "last": _last_cv_pricer,
        "master_enabled": crowdvolt_pricer.master_enabled(),
        "dry_run": crowdvolt_pricer.dry_run_enabled(),
        "undercut": crowdvolt_pricer.undercut_amount(),
        "max_drop_pct": crowdvolt_pricer.max_drop_pct(),
        "drop_window_hours": crowdvolt_pricer.drop_window_hours(),
        "drop_cap_enabled": crowdvolt_pricer.drop_cap_enabled(),
        "ignore_singles": crowdvolt_pricer.ignore_single_competitors(),
        "interval_minutes": CV_PRICER_INTERVAL_MINUTES,
        "rows": rows,
        "configs": cfgs,
        "log": db.cv_pricer_log_recent(50),
    })


@app.route("/api/cvpricer/config", methods=["POST"])
def api_cvpricer_config():
    """Per-ask pricer config. Body: {ask_uqid, enabled?, floor_price?,
    resume?, no_drop_cap?}. Enabling requires a floor (set now or before)."""
    from flask import request
    body = request.get_json(silent=True) or {}
    ask_uqid = str(body.get("ask_uqid") or "").strip()
    if not ask_uqid:
        return jsonify({"error": "ask_uqid required"}), 400
    now_iso = datetime.now(timezone.utc).isoformat()
    fields = {}
    if "floor_price" in body:
        try:
            floor = float(body["floor_price"])
        except (TypeError, ValueError):
            return jsonify({"error": "floor_price must be a number"}), 400
        if floor <= 0:
            return jsonify({"error": "floor_price must be positive"}), 400
        fields["floor_price"] = floor
    if "enabled" in body:
        enabled = bool(body["enabled"])
        if enabled:
            existing = db.cv_pricer_config_get(ask_uqid) or {}
            if not (fields.get("floor_price") or existing.get("floor_price")):
                return jsonify({"error": "set a floor price before enabling"}), 400
        fields["enabled"] = 1 if enabled else 0
    if body.get("resume"):
        fields.update(paused=0, paused_reason=None, paused_at=None)
    if "no_drop_cap" in body:
        fields["no_drop_cap"] = 1 if body["no_drop_cap"] else 0
    if not fields:
        return jsonify({"error": "nothing to update"}), 400
    db.cv_pricer_config_set(ask_uqid, fields, now_iso)
    return jsonify({"ok": True, "config": db.cv_pricer_config_get(ask_uqid)})


@app.route("/api/cvpricer/settings", methods=["POST"])
def api_cvpricer_settings():
    from flask import request
    body = request.get_json(silent=True) or {}
    now_iso = datetime.now(timezone.utc).isoformat()
    out = {}
    if "master_enabled" in body:
        v = "true" if body.get("master_enabled") else "false"
        db.setting_set("cv_pricer_master_enabled", v, now_iso)
        out["master_enabled"] = v
    if "dry_run" in body:
        v = "true" if body.get("dry_run") else "false"
        db.setting_set("cv_pricer_dry_run", v, now_iso)
        out["dry_run"] = v
    if "undercut" in body:
        try:
            u = float(body["undercut"])
        except (TypeError, ValueError):
            return jsonify({"error": "undercut must be a number"}), 400
        if not (0 < u <= 50):
            return jsonify({"error": "undercut must be between 0 and 50"}), 400
        db.setting_set("cv_pricer_undercut", f"{u:.2f}", now_iso)
        out["undercut"] = u
    if "ignore_singles" in body:
        v = "true" if body.get("ignore_singles") else "false"
        db.setting_set("cv_pricer_ignore_singles", v, now_iso)
        out["ignore_singles"] = v
    if "max_drop_pct" in body:
        try:
            pct = float(body["max_drop_pct"])
        except (TypeError, ValueError):
            return jsonify({"error": "max_drop_pct must be a number"}), 400
        if not (1 <= pct <= 90):
            return jsonify({"error": "max_drop_pct must be between 1 and 90"}), 400
        db.setting_set("cv_pricer_max_drop_pct", f"{pct:g}", now_iso)
        out["max_drop_pct"] = pct
    if "drop_cap_enabled" in body:
        v = "true" if body.get("drop_cap_enabled") else "false"
        db.setting_set("cv_pricer_drop_cap_enabled", v, now_iso)
        out["drop_cap_enabled"] = v
    if not out:
        return jsonify({"error": "nothing to update"}), 400
    return jsonify({"ok": True, **out})


@app.route("/api/cvpricer/run-now", methods=["POST"])
def api_cvpricer_run_now():
    if _last_cv_pricer["running"]:
        return jsonify({"error": "cv pricer already running"}), 429
    threading.Thread(target=run_cv_pricer, daemon=True).start()
    return jsonify({"started": True})


@app.route("/api/cvpricer/market/<event_uqid>")
def api_cvpricer_market(event_uqid):
    """Last book snapshot for an event (written by every tick)."""
    snap = db.cv_market_snapshot_get(event_uqid)
    if not snap:
        return jsonify({"event_uqid": event_uqid, "fetched_at": None, "rows": []})
    try:
        rows = json.loads(snap["rows_json"])
    except (TypeError, ValueError):
        rows = []
    return jsonify({"event_uqid": event_uqid, "fetched_at": snap["fetched_at"],
                    "rows": rows})


@app.route("/api/cvpricer/market/<event_uqid>/refresh", methods=["POST"])
def api_cvpricer_market_refresh(event_uqid):
    """Live book fetch (~10s, needs the CDP Chrome)."""
    if _last_cv_pricer["running"]:
        return jsonify({"error": "cv pricer tick running — try again in a minute"}), 429
    try:
        rows = crowdvolt_pricer.refresh_book_snapshot(event_uqid)
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 502
    return jsonify({"event_uqid": event_uqid,
                    "fetched_at": datetime.now(timezone.utc).isoformat(),
                    "rows": rows})


# --- CrowdVolt relink job queue ----------------------------------------
# kartis.homes cannot reach into the desktop's LAN, and a tab here cannot
# drive localhost:9222 either (mixed content + CDP's Host check). So the
# "Re-link" button does not perform the relink - it PARKS a request that the
# desktop agent (cv_agent.py) picks up on its next poll. Server proposes,
# desktop disposes.
CVAUTH_JOB_KEY = "cvauth_job"
CVAUTH_AGENT_SEEN_KEY = "cvauth_agent_seen"
# Two agent poll intervals plus slack. Below this the UI says the desktop is
# offline rather than spinning on a button nothing will ever answer.
CVAUTH_AGENT_ONLINE_SECONDS = 150


def _cvauth_secret_error(body=None):
    """None when the caller proved it holds KARTIS_CVAUTH_SECRET, else the
    (response, status) to return. Header first so the secret stays out of
    query strings and access logs; body is accepted for the import path."""
    import hmac
    from flask import request
    secret = (os.environ.get("KARTIS_CVAUTH_SECRET") or "").strip()
    if not secret:
        return jsonify({"error": "KARTIS_CVAUTH_SECRET is not set on the "
                                 "server - cvauth import is disabled"}), 503
    supplied = (request.headers.get("X-Kartis-Secret")
                or (body or {}).get("secret") or "")
    if not hmac.compare_digest(str(supplied), secret):
        return jsonify({"error": "bad secret"}), 403
    return None


def _cvauth_job_get():
    raw = db.setting_get(CVAUTH_JOB_KEY)
    try:
        job = json.loads(raw) if raw else {}
    except ValueError:
        job = {}
    return {"state": "idle", "requested_at": None, "started_at": None,
            "finished_at": None, "error": None, **job}


def _cvauth_job_set(**fields):
    job = {**_cvauth_job_get(), **fields}
    db.setting_set(CVAUTH_JOB_KEY, json.dumps(job),
                   datetime.now(timezone.utc).isoformat())
    return job


def _cvauth_agent_state():
    seen = db.setting_get(CVAUTH_AGENT_SEEN_KEY)
    online, age = False, None
    if seen:
        try:
            stamp = datetime.fromisoformat(seen)
            # A naive stamp would parse fine and then blow up on the
            # subtraction with a TypeError, 500-ing the status route (and so
            # the whole badge) over a cosmetic field. Treat it as unknown.
            if stamp.tzinfo is None:
                stamp = stamp.replace(tzinfo=timezone.utc)
            age = (datetime.now(timezone.utc) - stamp).total_seconds()
            online = age <= CVAUTH_AGENT_ONLINE_SECONDS
        except (ValueError, TypeError):
            pass
    return {"agent_seen": seen, "agent_online": online,
            "agent_seen_age_s": None if age is None else round(age)}


@app.route("/api/cvauth/request", methods=["POST"])
def api_cvauth_request():
    """Park a relink request for the desktop agent. Body: {"cancel": true}
    to withdraw one - a request nothing is listening for would otherwise sit
    pending forever."""
    from flask import request
    body = request.get_json(silent=True) or {}
    now = datetime.now(timezone.utc).isoformat()
    if body.get("cancel"):
        job = _cvauth_job_set(state="idle", error=None, finished_at=now)
    else:
        job = _cvauth_job_set(state="pending", requested_at=now,
                              started_at=None, finished_at=None, error=None)
    return jsonify({"job": job, **_cvauth_agent_state()})


@app.route("/api/cvauth/job")
def api_cvauth_job():
    """Agent poll. Doubles as the heartbeat that tells the UI a desktop is
    actually listening."""
    err = _cvauth_secret_error()
    if err:
        return err
    db.setting_set(CVAUTH_AGENT_SEEN_KEY,
                   datetime.now(timezone.utc).isoformat(),
                   datetime.now(timezone.utc).isoformat())
    return jsonify(_cvauth_job_get())


@app.route("/api/cvauth/job/result", methods=["POST"])
def api_cvauth_job_result():
    """Agent reports progress. Body: {"state": "running"|"error"|"done",
    "error": "..."}. Lets the button show why a relink failed on a machine
    the user may not be sitting at."""
    from flask import request
    body = request.get_json(silent=True) or {}
    err = _cvauth_secret_error(body)
    if err:
        return err
    state = str(body.get("state") or "").strip()
    if state not in ("running", "error", "done"):
        return jsonify({"error": "state must be running, error or done"}), 400
    now = datetime.now(timezone.utc).isoformat()
    fields = {"state": state, "error": (body.get("error") or None)}
    if state == "running":
        fields["started_at"] = now
    else:
        fields["finished_at"] = now
    return jsonify(_cvauth_job_set(**fields))


@app.route("/api/cvsales/import", methods=["POST"])
def api_cvsales_import():
    """Ingest CrowdVolt sell_delivered rows captured by a signed-in browser.

    This box cannot always reach CrowdVolt itself. cv_refresh_token dies (it
    is not re-persisted when CrowdVolt rotates it), and
    scraper._fetch_crowdvolt_sales_api runs CACHE-ONLY - no playwright - so
    it cannot re-harvest. It raises, the UI fallback finds nothing because
    CrowdVolt retired the Completed table, and the tick records
    crowdvolt_sales: 0 with no error anywhere. Meanwhile a browser that IS
    signed in can read the same endpoint perfectly.

    Body: {"rows": [ <raw /api/buy_sell_history/sell_delivered rows> ]}

    RAW rows on purpose: mapping stays in scraper._map_crowdvolt_api_sale so
    there is exactly one parser for CrowdVolt's shape, and a browser-sourced
    row lands byte-identical to a scraped one - same id (the order number),
    so the two paths converge on one row instead of double-counting.

    Gated by the edge (Caddy basic auth) like every other mutating route
    here; it carries no credential, only sales data.
    """
    from flask import request
    body = request.get_json(silent=True) or {}
    rows = body.get("rows")
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "body needs a non-empty 'rows' list"}), 400
    if len(rows) > 2000:
        return jsonify({"error": "too many rows (max 2000)"}), 413

    mapped, bad = [], 0
    for r in rows:
        if not isinstance(r, dict):
            bad += 1
            continue
        try:
            m = scraper._map_crowdvolt_api_sale(r)
        except Exception:
            m = None
        if m:
            mapped.append(m)
        else:
            bad += 1          # no order_number, or an unparseable shape
    if not mapped:
        return jsonify({"error": "no rows mapped - wrong shape?",
                        "received": len(rows), "unmappable": bad}), 400

    before = {x["id"] for x in db.all_crowdvolt_sales()}
    db.upsert_crowdvolt_sales(mapped, datetime.now(timezone.utc).isoformat())
    after = {x["id"] for x in db.all_crowdvolt_sales()}
    inserted = sorted(after - before)
    print(f"[kartis] cvsales: imported {len(mapped)} row(s), "
          f"{len(inserted)} new, {bad} unmappable")
    return jsonify({
        "received": len(rows), "mapped": len(mapped), "unmappable": bad,
        "inserted": len(inserted), "updated": len(mapped) - len(inserted),
        "new_order_ids": inserted[:50],
        "crowdvolt_sales_total": len(after),
    })


@app.route("/api/cvauth/status")
def api_cvauth_status():
    """Is a CrowdVolt session cached, and how stale? Drives the /inventory
    badge. Returns no cookie values."""
    import cv_auth
    return jsonify({**cv_auth.session_status(),
                    "import_configured": bool(
                        (os.environ.get("KARTIS_CVAUTH_SECRET") or "").strip()),
                    "job": _cvauth_job_get(),
                    **_cvauth_agent_state()})


@app.route("/api/cvauth/import", methods=["POST"])
def api_cvauth_import():
    """Receive a cv_refresh_token harvested on the desktop.

    CrowdVolt's refresh cookie is HttpOnly and only readable out of a
    signed-in Chrome over CDP. This box has none, and it cannot reach into
    the LAN of a machine that does, so the desktop pushes and the server
    receives. scripts/cv_relink.py is the sender.

    Body: {"secret": "...", "cookies": {"cv_refresh_token": "..."}}

    Gated on KARTIS_CVAUTH_SECRET *in addition to* Caddy's basic auth: the
    payload is a live ~30-day credential, so it does not ride on the edge
    gate alone. Missing secret fails closed (503) rather than accepting
    anything - an unset env var must never mean "open".
    """
    from flask import request
    import cv_auth

    body = request.get_json(silent=True) or {}
    err = _cvauth_secret_error(body)
    if err:
        return err
    try:
        status = cv_auth.import_cookies(body.get("cookies") or {})
    except cv_auth.CvAuthError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:                      # disk full, perms, ...
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 500
    print(f"[kartis] cvauth: imported cookies {status['cookie_names']}")
    # A successful import IS the job completing - the agent should not have
    # to report separately for the happy path.
    _cvauth_job_set(state="done", error=None,
                    finished_at=datetime.now(timezone.utc).isoformat())
    return jsonify({"imported": True, **status})


@app.route("/api/cvauth/verify", methods=["POST"])
def api_cvauth_verify():
    """Spend the cached cookie on one real /api/auth/refresh call.

    The import endpoint can only check the payload's SHAPE; this proves
    CrowdVolt still honours it. Worth its own route so the relink script can
    confirm end to end instead of leaving you to find out at the next tick.
    """
    import cv_auth
    try:
        cv_auth.CvAuth().token(force=True)      # cache-only: no browser here
    except cv_auth.CvAuthError as e:
        return jsonify({"ok": False, "error": str(e)}), 502
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 502
    return jsonify({"ok": True, **cv_auth.session_status()})


@app.route("/api/pacha-events/status")
def api_pacha_events_status():
    return jsonify({
        **_last_pacha_events,
        "enabled": PACHA_MONITOR_ENABLED,
        "interval_minutes": PACHA_MONITOR_INTERVAL_MINUTES,
        "seen_total": db.pacha_seen_count(),
    })


@app.route("/api/pacha-events/run-now", methods=["POST"])
def api_pacha_events_run_now():
    """Manual tick — runs synchronously so the response carries the diff
    summary (also the E2E test hook). Works even when the scheduled job is
    disabled on this machine."""
    if _last_pacha_events["running"]:
        return jsonify({"error": "pacha monitor already running"}), 429
    run_pacha_events()
    return jsonify(dict(_last_pacha_events))


@app.route("/api/il-events/status")
def api_il_events_status():
    return jsonify({
        **_last_il_events,
        "enabled": IL_EVENTS_ENABLED,
        "interval_minutes": IL_EVENTS_INTERVAL_MINUTES,
        "seen_total": db.site_events_seen_counts(),
    })


@app.route("/api/il-events/run-now", methods=["POST"])
def api_il_events_run_now():
    """Manual tick — runs synchronously so the response carries the diff
    summary (also the E2E test hook). Works even when the scheduled job is
    disabled on this machine."""
    if _last_il_events["running"]:
        return jsonify({"error": "il-events monitor already running"}), 429
    run_il_events()
    return jsonify(dict(_last_il_events))


@app.route("/api/market/status")
def api_market_status():
    return jsonify({
        **_last_market,
        "enabled": MARKET_ENABLED,
        "interval_minutes": MARKET_INTERVAL_MINUTES,
    })


@app.route("/api/market/run-now", methods=["POST"])
def api_market_run_now():
    """Manual sweep — synchronous (takes a few minutes: kupat's catalog
    rides a headless browser). Works even when the scheduled job is
    disabled on this machine."""
    if _last_market["running"]:
        return jsonify({"error": "market sweep already running"}), 429
    run_market_sweep()
    return jsonify(dict(_last_market))


@app.route("/api/vgsales/status")
def api_vgsales_status():
    return jsonify({
        **_last_vgsales,
        "enabled": VGSALES_ENABLED,
        "interval_minutes": viagogo_market_sales.INTERVAL_MINUTES,
    })


@app.route("/api/vgsales/run-now", methods=["POST"])
def api_vgsales_run_now():
    """Manual tick — synchronous (~2s per tracked event under the browser
    lock). Works even when the scheduled job is disabled on this machine."""
    if _last_vgsales["running"]:
        return jsonify({"error": "vgsales tick already running"}), 429
    run_vgsales(force=True)
    return jsonify(dict(_last_vgsales))


def _fresh_thread_job(fn):
    """Run a scheduler job in a brand-new disposable thread each tick.

    APScheduler's executor REUSES pool threads. When a sync_playwright
    session dies mid-run (e.g. the CDP Chrome restarts under it —
    TargetClosedError), it can leave a running asyncio loop behind in that
    thread, and every later sync_playwright() scheduled onto it fails with
    "Playwright Sync API inside the asyncio loop" — this silently killed
    the pricer + market sweep from 2026-07-07 13:10 until the next service
    restart. A fresh thread per tick makes that poisoning impossible.
    Only playwright-touching jobs need this."""
    def _runner():
        t = threading.Thread(target=fn, daemon=True, name=f"fresh-{fn.__name__}")
        t.start()
        t.join()
    _runner.__name__ = f"{fn.__name__}_fresh"
    return _runner


def _reconcile_orphaned_pushes():
    """A viagogo approve runs in a background thread; if the process dies
    mid-run (a deploy restart, a crash), its push is left stuck in
    'creating' forever with no thread to finish it. On startup, kick any
    such row back to awaiting_approval so the user can simply re-approve —
    the approve flow is idempotent enough (a half-created listing shows up
    on /Listings and won't be duplicated blindly; the user reviews first)."""
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        for row in db.viagogo_push_all(status="creating"):
            db.viagogo_push_update(row["id"], {
                "status": "awaiting_approval",
                "error": ("interrupted mid-create (server restart) — re-approve; "
                          "if a listing was already made on viagogo, reject this instead"),
            }, now_iso)
            print(f"[viagogo-push] reset orphaned 'creating' row {row['id']} → awaiting_approval")
    except Exception:
        traceback.print_exc()


_reconcile_orphaned_pushes()

scheduler = BackgroundScheduler(daemon=True)
scheduler.add_job(_fresh_thread_job(run_scraper), "interval", hours=1, id="scrape")
scheduler.add_job(run_backup, "cron", hour=3, minute=0, id="backup")
if TM_CHECK_ENABLED:
    # Fresh thread: kupat watchers ride sync_playwright (in-page harvest), so
    # an unwrapped tick inherits any poisoned pool thread — seen 2026-08-12,
    # every kupat check failing "Sync API inside the asyncio loop" for a day.
    scheduler.add_job(_fresh_thread_job(run_tm_check), "interval", seconds=TM_CHECK_INTERVAL_SECONDS, id="tm_check")
    if DICE_SNIPER_ENABLED:
        scheduler.add_job(run_dice_sniper, "interval", seconds=DICE_SNIPER_SECONDS,
                          id="dice_sniper", max_instances=1)
    if TM_SNIPER_ENABLED:
        scheduler.add_job(run_tm_sniper, "interval", seconds=TM_SNIPER_SECONDS,
                          id="tm_sniper", max_instances=1)
else:
    print("[tm_check] disabled via TM_CHECK_ENABLED=0 — drop-checking runs elsewhere (e.g. the VPS watcher)")
# mail_intake drives the CDP browser too (viagogo event search on kupat/
# tickchak/TM-IL purchase emails) — fresh thread, same reason as the scrape.
scheduler.add_job(_fresh_thread_job(run_mail_intake), "interval", minutes=INTAKE_INTERVAL_MINUTES, id="mail_intake")
if PACHA_MONITOR_ENABLED:
    # First tick shortly after boot (baseline on a fresh DB), then every N min.
    scheduler.add_job(run_pacha_events, "interval",
                      minutes=PACHA_MONITOR_INTERVAL_MINUTES, id="pacha_events",
                      start_date=datetime.now() + timedelta(minutes=2))
else:
    print("[pacha] disabled via KARTIS_PACHA_MONITOR_ENABLED=0 — the monitor runs elsewhere (e.g. the VPS)")
if IL_EVENTS_ENABLED:
    # Offset from the pacha job so the two monitors' first ticks don't stack.
    scheduler.add_job(run_il_events, "interval",
                      minutes=IL_EVENTS_INTERVAL_MINUTES, id="il_events",
                      start_date=datetime.now() + timedelta(minutes=4))
else:
    print("[il-events] disabled via KARTIS_IL_EVENTS_ENABLED=0 — the monitor runs elsewhere (e.g. the VPS)")
if EDM_MONITOR_ENABLED:
    # Offset again so pacha / il_events / edm first ticks don't collide.
    scheduler.add_job(run_edm_events, "interval",
                      minutes=EDM_MONITOR_INTERVAL_MINUTES, id="edm_events",
                      start_date=datetime.now() + timedelta(minutes=3))
else:
    print("[edm] disabled via KARTIS_EDM_MONITOR_ENABLED=0 — the monitor runs elsewhere (e.g. the VPS)")
scheduler.add_job(run_todo_remind, "cron", hour=8, minute=0, id="todo_remind")
# Festival/GA sales snapshots — fire one immediately (next_run_time) so the
# Festival / GA Tracker pages have a baseline right after a restart, then
# every N minutes.
# Fresh thread: kupat GA label refreshes ride sync_playwright too.
scheduler.add_job(_fresh_thread_job(run_sales_sync), "interval", minutes=FESTIVAL_SYNC_MINUTES,
                  id="sales_sync", next_run_time=datetime.now())
# Auto-pricer: offset from the top of the hour (start_date pushes the first
# run out) so its browser use doesn't collide with the hourly scrape. Cheap
# no-op while pricer_master_enabled is off.
scheduler.add_job(_fresh_thread_job(run_pricer), "interval",
                  minutes=PRICER_FAST_INTERVAL_MINUTES,
                  id="pricer",
                  start_date=datetime.now() + timedelta(minutes=7))
# CrowdVolt pricer: also drives the CDP Chrome (its own tab) — offset clear
# of scrape (t+0) / pricer (t+7). Cheap no-op while its master switch is off
# (still refreshes the /cvpricer listings mirror).
scheduler.add_job(_fresh_thread_job(run_cv_pricer), "interval", minutes=CV_PRICER_INTERVAL_MINUTES,
                  id="cv_pricer",
                  start_date=datetime.now() + timedelta(minutes=10))
if MARKET_ENABLED:
    # Market-wide availability sweep. Offset well clear of the hourly scrape
    # (t+0) and the pricer (t+7) — this one also launches a Chromium.
    scheduler.add_job(_fresh_thread_job(run_market_sweep), "interval", minutes=MARKET_INTERVAL_MINUTES,
                      id="market_sweep",
                      start_date=datetime.now() + timedelta(minutes=20))
else:
    print("[market] disabled via KARTIS_MARKET_ENABLED=0 — the sweep runs elsewhere")
if VGSALES_ENABLED:
    # Viagogo market-sales tracker — drives the CDP Chrome, so fresh thread
    # and staggered clear of scrape (t+0) / pricer (t+7) / market (t+20).
    scheduler.add_job(_fresh_thread_job(run_vgsales), "interval",
                      minutes=viagogo_market_sales.INTERVAL_MINUTES,
                      id="vgsales",
                      start_date=datetime.now() + timedelta(minutes=13))
else:
    print("[vgsales] disabled via KARTIS_VGSALES_ENABLED=0 — the tracker runs elsewhere")

# One-shot: archive any pre-existing inventory_overrides rows whose status
# value was already typed as "not sold" (or a variant). Idempotent so
# subsequent restarts find no candidates and do nothing.
try:
    _migrated = _migrate_legacy_unsold_overrides()
    if _migrated:
        print(f"[unsold-migration] archived {_migrated} legacy 'not sold' status override(s)")
except Exception:
    traceback.print_exc()

scheduler.start()

threading.Thread(target=run_backup, daemon=True).start()


if __name__ == "__main__":
    app.run(debug=True, port=5000, use_reloader=False)
